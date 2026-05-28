"""
Build docs/reminders.xlsx — every "do X by Y" / "watch for X" / "every quarter X"
the MyRX project has accumulated.

Sheets:
  1. Date-Specific Deadlines  — drop-dead dates with auto-counting days-remaining
  2. Event-Triggered Actions  — "when X happens, do Y"
  3. Recurring Tasks          — monthly / quarterly / yearly cadence
  4. How To Use This File     — re-snapshot protocol + status conventions

Run from repo root:
    python scripts/build_reminders_xlsx.py
"""
from __future__ import annotations
import datetime as _dt
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "reminders.xlsx"

# ─────────── styles ────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="F8FAFC", size=11)
SUBHEAD_FILL = PatternFill("solid", fgColor="334155")
SUBHEAD_FONT = Font(bold=True, color="F8FAFC", size=10)

STATUS_FILLS = {
    "Pending":   PatternFill("solid", fgColor="FEF3C7"),  # amber-100
    "Done":      PatternFill("solid", fgColor="DCFCE7"),  # green-100
    "Snoozed":   PatternFill("solid", fgColor="DBEAFE"),  # blue-100
    "Overdue":   PatternFill("solid", fgColor="FEE2E2"),  # red-100
    "Watch":     PatternFill("solid", fgColor="E0E7FF"),  # indigo-100
}

THIN = Side(border_style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header_row(ws, row_idx, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row_idx].height = 28

def style_data_rows(ws, start_row, end_row, cols, status_col=None):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = BORDER
            cell.font = Font(size=10)
        if status_col is not None:
            status_val = str(ws.cell(row=r, column=status_col).value or "")
            if status_val in STATUS_FILLS:
                ws.cell(row=r, column=status_col).fill = STATUS_FILLS[status_val]
                ws.cell(row=r, column=status_col).font = Font(size=10, bold=True)

def autosize(ws, widths):
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

# ═══════════════ Workbook ═════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)

# ─────────── Sheet 1: Date-Specific Deadlines ─────────────────────────────────
ws = wb.create_sheet("Date-Specific Deadlines")
ws["A1"] = "Date-Specific Deadlines"
ws["A1"].font = Font(bold=True, size=14)
ws.merge_cells("A1:F1")
ws["A2"] = ("Drop-dead dates. Days remaining auto-counts down vs TODAY(). "
            "After each item lands, change its Status cell to 'Done'.")
ws["A2"].font = Font(size=10, italic=True, color="475569")
ws["A2"].alignment = Alignment(wrap_text=True)
ws.merge_cells("A2:F2")
ws.row_dimensions[2].height = 32

cols = ["Date", "Days Remaining", "What to do", "How", "Status", "Notes"]
ws.append([])  # spacer row 3
ws.append(cols)  # row 4
style_header_row(ws, 4, len(cols))

deadlines = [
    [_dt.date(2026, 6, 10),
     None,  # placeholder, formula goes here per-row
     "DMARC ramp Step 2: bump pct from 25 → 100",
     "Run CF API PUT with content `v=DMARC1; p=quarantine; pct=100; adkim=r; aspf=r; rua=mailto:dmarc_rua@onsecureserver.net;`. ~2 weeks after Step 1 (May 27).",
     "Pending",
     "Pre-check: review any DMARC aggregate reports for unexpected fails."],

    [_dt.date(2026, 6, 24),
     None,
     "DMARC ramp Step 3: flip to p=reject (final state)",
     "CF API PUT with content `v=DMARC1; p=reject; adkim=r; aspf=r; rua=mailto:dmarc_rua@onsecureserver.net;`. ~2 weeks after Step 2.",
     "Pending",
     "Final hardened state. Spoofed mail bounces, never reaches inbox. Don't proceed if any quarantine-period complaints surfaced."],

    [_dt.date(2026, 7, 6),
     None,
     "Initiate myrxfit.com transfer from GoDaddy → Cloudflare Registrar",
     "Follow 8-step walkthrough in docs/vendors.xlsx → 'Domain Transfer Countdown' sheet. Saves ~$10/yr.",
     "Pending",
     "ICANN 60-day lock expires this date per GoDaddy's confirmation screen. Zero downtime — DNS already at Cloudflare."],

    [_dt.date(2026, 7, 25),
     None,
     "SendGrid Trial ends — pick paid plan or alternative",
     "Choices: SendGrid Essentials ($20/mo, 50k sends), Pro ($90/mo, 100k), or migrate outbound to AWS SES (~$0.10 per 1k). Decide based on actual coach invite + OTP volume by then.",
     "Pending",
     "Without action: outbound app email stops. Set calendar 5 days before."],

    [_dt.date(2027, 5, 18),
     None,
     "Zoho Mail Lite annual renewal ($12)",
     "Auto-renews if card on file. Otherwise: Mail Admin → Subscription → Renew.",
     "Pending",
     "Could upgrade to Workplace if user count grows past 1 (currently solo)."],
]

for i, row in enumerate(deadlines):
    excel_row = 5 + i
    # write date in A
    ws.cell(row=excel_row, column=1, value=row[0]).number_format = "yyyy-mm-dd"
    # formula in B (days remaining vs TODAY)
    ws.cell(row=excel_row, column=2, value=f"=MAX(0,A{excel_row}-TODAY())")
    # rest of columns
    for c_idx, val in enumerate(row[2:], start=3):
        ws.cell(row=excel_row, column=c_idx, value=val)

style_data_rows(ws, 5, 4 + len(deadlines), len(cols), status_col=5)
autosize(ws, {1: 14, 2: 14, 3: 38, 4: 52, 5: 11, 6: 42})
ws.freeze_panes = "A5"

# ─────────── Sheet 2: Event-Triggered Actions ──────────────────────────────────
ws2 = wb.create_sheet("Event-Triggered Actions")
ws2["A1"] = "Event-Triggered Actions"
ws2["A1"].font = Font(bold=True, size=14)
ws2.merge_cells("A1:E1")
ws2["A2"] = ("\"When X happens, do Y.\" These aren't dated — they fire when their "
             "trigger occurs. Mark 'Done' once the trigger has fired AND the action completed.")
ws2["A2"].font = Font(size=10, italic=True, color="475569")
ws2["A2"].alignment = Alignment(wrap_text=True)
ws2.merge_cells("A2:E2")
ws2.row_dimensions[2].height = 32

ev_cols = ["Trigger (when…)", "Then (do…)", "Reference", "Status", "Notes"]
ws2.append([])
ws2.append(ev_cols)
style_header_row(ws2, 4, len(ev_cols))

triggers = [
    ["Apple Developer Program enrollment approved",
     "Complete iOS reflection checklist before first TestFlight build",
     "CLAUDE.md → 'iOS reflection checklist' section",
     "Watch", "$99/yr. Apple verification can take 1-4 weeks for LLC."],

    ["Strava API app registered (instant, no approval)",
     "Build Strava integration (OAuth2 + REST)",
     "CLAUDE.md → 'Strava' build order #1",
     "Watch", "First integration to build because no approval delay."],

    ["Fitbit Personal-tier app registered (instant)",
     "Build Fitbit integration (OAuth2 + Web API)",
     "CLAUDE.md → 'Fitbit Web API' build order #2",
     "Watch", "Personal tier rate-limited; production tier needs separate approval."],

    ["Apple Developer Program active → HealthKit entitlement",
     "Build Apple HealthKit integration (mirror Samsung Health Data SDK architecture)",
     "CLAUDE.md → 'Apple HealthKit' build order #3",
     "Watch", "Native iOS framework. Requires com.apple.developer.healthkit entitlement."],

    ["Garmin Health API approval received (~2-4 wks)",
     "Build Garmin integration (OAuth1.0a + webhooks)",
     "CLAUDE.md → 'Garmin Health API' build order #5",
     "Watch", "Apply at https://developer.garmin.com — fill application from docs/integrations/developer-program-applications.md."],

    ["Whoop API v1 approval received (~1-2 wks)",
     "Build Whoop integration (OAuth2 + webhooks)",
     "CLAUDE.md → 'Whoop API v1' build order #6",
     "Watch", "Apply at https://developer.whoop.com."],

    ["Polar AccessLink approval received (~1-2 wks)",
     "Build Polar integration (OAuth2)",
     "CLAUDE.md → 'Polar AccessLink' build order #7",
     "Watch", "Apply via Polar Business team."],

    ["A client reports 'I never got your reply'",
     "Pause DMARC ramp progression. Check SendGrid Activity Feed + Zoho's outgoing log for that recipient. Inspect headers if available. If a legitimate path is failing DMARC, find root cause before continuing the ramp.",
     "SendGrid: https://app.sendgrid.com/email_activity\nZoho: Sent folder → original message → View Source",
     "Watch", "Most likely cause if it happens: a vendor we forgot is still sending mail as @myrxfit.com unaligned. Investigate before tightening DMARC further."],

    ["Before swapping email vendor / changing outbound infrastructure",
     "Send test from EACH outbound path → verify dkim=pass + dmarc=pass in recipient headers, BEFORE making the change permanent. Update docs/vendors.xlsx + CLAUDE.md email section in same turn.",
     "Existing pattern: Test 1/2/3 emails earlier this session.",
     "Watch", "Saves the kind of 4-hour Resend-to-SendGrid debug saga that birthed this whole reminders file."],

    ["When CLAUDE.md exceeds ~5000 lines",
     "Extract stable locked sections (Animation Patterns, Wearable scars, etc.) into /docs/*.md reference files and link from CLAUDE.md.",
     "Maintainability — context window costs scale with file size.",
     "Watch", "Currently the file is large but locked sections aren't actually being re-read by every Claude turn — they're being summarized."],

    ["Resend account dormant >60 days",
     "Resend auto-deactivates free-tier accounts. Confirm closure at resend.com → Settings → Delete Account for hygiene.",
     "Resend was removed from pipeline on 2026-05-27 but the account itself wasn't closed.",
     "Watch", "Low priority but cleaner to officially close."],

    ["First time setting up production Stripe webhook",
     "Change STRIPE_WEBHOOK_SECRET_LIVE in Supabase Edge Function secrets. Test with Stripe CLI: `stripe listen --forward-to https://xtxzfhoxyyrlxslgzvty.supabase.co/functions/v1/stripe-webhook?source=coach_subs`",
     ".env.example documents the secret name",
     "Watch", "Currently TEST mode only. Live keys not pulled."],

    ["Before Play Store production release",
     "Add production keystore SHA256 to public/.well-known/assetlinks.json so Android App Links deep-linking works.",
     "Currently only debug keystore SHA256 is registered.",
     "Watch", "If you skip this, magic-link emails open the browser instead of the mobile app on Android."],
]

for row in triggers:
    ws2.append(row)
style_data_rows(ws2, 5, 4 + len(triggers), len(ev_cols), status_col=4)
autosize(ws2, {1: 38, 2: 48, 3: 32, 4: 11, 5: 38})
ws2.freeze_panes = "A5"

# ─────────── Sheet 3: Recurring Tasks ──────────────────────────────────────────
ws3 = wb.create_sheet("Recurring Tasks")
ws3["A1"] = "Recurring Tasks"
ws3["A1"].font = Font(bold=True, size=14)
ws3.merge_cells("A1:E1")
ws3["A2"] = ("Tasks that repeat on a schedule. Update 'Last Done' each time you "
             "complete one — 'Next Due' auto-derives from it.")
ws3["A2"].font = Font(size=10, italic=True, color="475569")
ws3["A2"].alignment = Alignment(wrap_text=True)
ws3.merge_cells("A2:E2")
ws3.row_dimensions[2].height = 32

rec_cols = ["Cadence", "What to do", "Why", "Last Done", "Next Due (auto)"]
ws3.append([])
ws3.append(rec_cols)
style_header_row(ws3, 4, len(rec_cols))

# (Cadence, task, why, last_done_date, days_between)
recurring = [
    ["Monthly (1st week)",
     "Review DMARC aggregate reports forwarded to dmarc_rua@onsecureserver.net",
     "Catches new senders or alignment regressions before they become real problems",
     None, 30],

    ["Quarterly",
     "Rotate Cloudflare + SendGrid + Twilio API keys",
     "Security hygiene — limits blast radius if a credential leaks (incl. from chat transcripts)",
     None, 90],

    ["Quarterly",
     "Review legal docs in docs/legal/*.docx for accuracy (prices, vendor names, contact addresses)",
     "Stale legal text is a liability — esp. Privacy Policy + Refund Policy",
     None, 90],

    ["Yearly",
     "Review docs/vendors.xlsx for stale entries, cost optimization, plan upgrades",
     "Catches abandoned tools still billing, opportunities to consolidate",
     None, 365],

    ["Yearly (around Sep)",
     "Apple Developer Program renewal ($99)",
     "iOS app distribution stops if not renewed",
     None, 365],

    ["Per mobile release",
     "Bump app version in mobile/app.json + run iOS reflection checklist before TestFlight",
     "Apple rejects unchanged versions; reflection catches platform-specific bugs",
     None, None],
]

for i, (cadence, task, why, last_done, days) in enumerate(recurring):
    excel_row = 5 + i
    ws3.cell(row=excel_row, column=1, value=cadence)
    ws3.cell(row=excel_row, column=2, value=task)
    ws3.cell(row=excel_row, column=3, value=why)
    if last_done:
        ws3.cell(row=excel_row, column=4, value=last_done).number_format = "yyyy-mm-dd"
    if days:
        ws3.cell(row=excel_row, column=5,
                 value=f'=IF(ISBLANK(D{excel_row}),"⚠ enter Last Done",D{excel_row}+{days})').number_format = "yyyy-mm-dd"
    else:
        ws3.cell(row=excel_row, column=5, value="n/a (per-event)")
style_data_rows(ws3, 5, 4 + len(recurring), len(rec_cols))
autosize(ws3, {1: 22, 2: 42, 3: 44, 4: 14, 5: 18})
ws3.freeze_panes = "A5"

# ─────────── Sheet 4: How To Use This File ─────────────────────────────────────
ws4 = wb.create_sheet("How To Use This File")
ws4["A1"] = "How To Use This File"
ws4["A1"].font = Font(bold=True, size=14)
ws4.merge_cells("A1:B1")

instructions = [
    ("Status conventions",  ""),
    ("Pending",   "Hasn't happened yet but on the calendar (amber background)"),
    ("Done",      "Action completed (green background) — leave the row for history"),
    ("Snoozed",   "Intentionally pushed (blue background) — update Notes to say why"),
    ("Overdue",   "Date passed without action (red background) — fix or revisit"),
    ("Watch",     "Event-triggered, hasn't fired yet (indigo background) — passive monitoring"),
    ("", ""),
    ("Updating dates",  ""),
    ("Days Remaining",  "Auto-computed via =MAX(0, dateCell - TODAY()). Don't overwrite the formula."),
    ("Next Due (recurring sheet)", "Auto-derived from Last Done + cadence. Update Last Done after completing the task and Next Due updates itself."),
    ("Date formatting", "All date cells are yyyy-mm-dd. Excel auto-recognizes them as dates for the formulas."),
    ("", ""),
    ("Regenerating this file",  ""),
    ("Source",  "scripts/build_reminders_xlsx.py — Python + openpyxl"),
    ("Command", "From repo root: python scripts/build_reminders_xlsx.py"),
    ("Note",    "Regenerating wipes any manual edits to status cells. If you've marked things Done that aren't in the source, copy your overrides into the script first, then regenerate."),
    ("", ""),
    ("Cross-references", ""),
    ("docs/vendors.xlsx",   "Vendor inventory — referenced by GoDaddy unlock + Zoho renewal items"),
    ("docs/legal/*.docx",   "Legal docs — referenced by quarterly review item"),
    ("CLAUDE.md",           "Project bible — referenced by iOS, integrations, and dev-environment items"),
]
for i, (lbl, desc) in enumerate(instructions):
    row = 3 + i
    if lbl and not desc:
        # section header
        cell = ws4.cell(row=row, column=1, value=lbl)
        cell.font = Font(bold=True, size=11, color="1F2937")
        ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    else:
        ws4.cell(row=row, column=1, value=lbl).font = Font(size=10, bold=True)
        ws4.cell(row=row, column=2, value=desc).font = Font(size=10)
        ws4.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
ws4.column_dimensions["A"].width = 28
ws4.column_dimensions["B"].width = 80

# ─────────── Save ──────────────────────────────────────────────────────────────
OUT.parent.mkdir(parents=True, exist_ok=True)
try:
    wb.save(OUT)
    print(f"Wrote {OUT}")
except PermissionError:
    alt = OUT.with_name("reminders_v2.xlsx")
    wb.save(alt)
    print(f"[WARN] {OUT.name} is open (Excel locked it).")
    print(f"       Wrote {alt} instead. Close Excel + re-run OR rename {alt.name} -> {OUT.name}.")
print(f"Sheets: {wb.sheetnames}")
