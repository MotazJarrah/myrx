"""
One-shot: add the Stripe pause/resume orchestration TODO to:
  1. docs/launch_checklist.xlsx  (Phase 1 - T-30 days, Billing category)
  2. docs/reminders.xlsx          (Event-Triggered Actions)

Locked May 28 2026 — deletion-grace-period flow ships without Stripe orchestration.
The actual Stripe API integration (pause subscription on grace start, resume +
charge missed cycles on reactivation) is deferred to a follow-up build but
must NOT be forgotten before live paying coaches start using the flow.

Run once. Idempotent — checks for existing rows by Item / Trigger text before
appending so re-running won't duplicate.
"""
import openpyxl
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
CHECKLIST = REPO / 'docs' / 'launch_checklist.xlsx'
REMINDERS = REPO / 'docs' / 'reminders.xlsx'

STRIPE_ITEM = 'Build Stripe pause/resume orchestration for account-deletion grace period'
STRIPE_WHY = (
    "When a coach with an active Stripe subscription clicks 'Delete my account', the 30-day "
    "grace period should PAUSE Stripe billing (no renewals charged during grace). On reactivation, "
    "if a billing cycle elapsed during grace, charge the user once for the missed cycle then "
    "resume on the original cadence. On grace expiry (no reactivation), cancel the Stripe "
    "subscription cleanly. Without this, paying coaches who request deletion either keep getting "
    "charged during their grace (bad) or get free coverage on reactivation (also bad). "
    "Requires: edge function for pause/resume Stripe API calls + webhook handler updates to "
    "write billing:* events into the activity_events table. "
    "DB lifecycle (profiles.scheduled_for_deletion_at + RPCs + cron) is already live."
)
STRIPE_PRIORITY = 'M+C'  # matches existing convention for Billing items
STRIPE_STATUS = 'Pending'

# ── 1. launch_checklist.xlsx ────────────────────────────────────────────────
wb = openpyxl.load_workbook(CHECKLIST)
sheet_name = next((s for s in wb.sheetnames if 'Phase 1' in s), None)
if not sheet_name:
    raise SystemExit(f"Couldn't find Phase 1 sheet in {CHECKLIST}")
ws = wb[sheet_name]

# Idempotency: bail if the item is already in there.
already_present = False
for row in ws.iter_rows(min_row=2, values_only=True):
    if row and len(row) > 1 and row[1] and STRIPE_ITEM in str(row[1]):
        already_present = True
        break

if already_present:
    print(f"[checklist] Already present in '{sheet_name}' — skipping.")
else:
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value='Billing')
    ws.cell(row=next_row, column=2, value=STRIPE_ITEM)
    ws.cell(row=next_row, column=3, value=STRIPE_WHY)
    ws.cell(row=next_row, column=4, value=STRIPE_PRIORITY)
    ws.cell(row=next_row, column=5, value=STRIPE_STATUS)
    wb.save(CHECKLIST)
    print(f"[checklist] Added row {next_row} to '{sheet_name}'.")

# ── 2. reminders.xlsx ───────────────────────────────────────────────────────
wb = openpyxl.load_workbook(REMINDERS)
sheet_name = next((s for s in wb.sheetnames if 'Event-Triggered' in s), None)
if not sheet_name:
    raise SystemExit(f"Couldn't find Event-Triggered sheet in {REMINDERS}")
ws = wb[sheet_name]

TRIGGER = 'First paying coach signs up via Stripe (real subscription, not test mode)'
THEN = (
    'BEFORE that coach can hit the 30-day point and request deletion, ship the Stripe '
    'pause/resume orchestration edge function. The DB-side deletion flow exists already '
    "but does NOT touch Stripe yet — paying coaches will keep getting charged during grace "
    'unless this lands.'
)
REFERENCE = 'docs/launch_checklist.xlsx → Phase 1 → Billing → Stripe pause/resume row'
STATUS = 'Watch'
NOTES = 'Blocking for billing correctness once paying coaches exist. Non-blocking today (no real-money subs yet).'

already_present = False
for row in ws.iter_rows(min_row=2, values_only=True):
    if row and len(row) > 0 and row[0] and 'pause/resume' in str(row[1] or '').lower():
        already_present = True
        break

if already_present:
    print(f"[reminders] Already present in '{sheet_name}' — skipping.")
else:
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=TRIGGER)
    ws.cell(row=next_row, column=2, value=THEN)
    ws.cell(row=next_row, column=3, value=REFERENCE)
    ws.cell(row=next_row, column=4, value=STATUS)
    ws.cell(row=next_row, column=5, value=NOTES)
    wb.save(REMINDERS)
    print(f"[reminders] Added row {next_row} to '{sheet_name}'.")

print("Done.")
