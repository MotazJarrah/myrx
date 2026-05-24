"""
Build docs/plan_wizard_strings.xlsx — every user-facing string in the
self-coached calorie plan wizard, grouped by where it appears + severity
+ trigger condition, with an empty "Your rewrite" column for the user
to fill in and hand back.

This script is the source of truth for what's in the wizard right now.
Re-run it after any string edit in:
  • mobile/src/lib/planPresets.ts
      - paceProfileWarning() / macroProfileWarning()  → pace/macro chips
      - evaluateRealism()                              → reality screen issues
      - consolidatedSuggestion rationale builder       → apply card
  • mobile/src/components/PlanWizardSheet.tsx
      - RealityCheckScreen classification labels       → pill text
      - severity labels                                → red/amber card headers
      - intro line                                     → below pill
      - apply button + keep-note                       → bottom of apply card

Run via:  python scripts/build_plan_wizard_strings_xlsx.py
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Color palette (mirrors mobile/src/theme.ts where practical) ──
WHITE       = "FFFFFFFF"
BLACK       = "FF111827"
ZINC        = "FF374151"   # neutral header band
SLATE_TINT  = "FFE5E7EB"   # row stripe
RED_TINT    = "FFFEE2E2"   # major-severity rows
RED_BORDER  = "FFDC2626"
AMBER_TINT  = "FFFEF3C7"   # caution-severity rows
AMBER_BORDER= "FFD97706"
GREEN_TINT  = "FFD1FAE5"
GREEN_BORDER= "FF059669"
NEUTRAL_TINT= "FFF3F4F6"   # frame / structural rows

THIN = Side(border_style="thin", color="FFD1D5DB")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── String catalog ───────────────────────────────────────────────
# Each entry: (id, section, severity, where, trigger, current_text)
# severity drives the row fill color (red/amber/green/neutral).

ENTRIES = [
    # ────────── 1. FRAME (pill + intro + severity labels) ──────────
    ("H1", "Frame", "green",
     "Header pill — on track",
     "Classification = on_track (zero issues detected)",
     "Plan looks realistic"),

    ("H2", "Frame", "neutral",
     "Header pill — needs work",
     "Classification != on_track (1+ issues, any severity)",
     "Plan requires optimization"),

    ("H3", "Frame", "green",
     "Intro line — on track",
     "Classification = on_track",
     "Your plan is well-matched — here's what to expect."),

    ("H4", "Frame", "neutral",
     "Intro line — needs work",
     "Classification != on_track",
     "Here's how to optimize your plan."),

    ("H5", "Frame", "red",
     "Severity label — major (red card header)",
     "Renders at the top of every red issue card",
     "Conflict with your goal"),

    ("H6", "Frame", "amber",
     "Severity label — caution (amber card header)",
     "Renders at the top of every amber issue card",
     "Could be more efficient"),

    # ────────── 2. RED BOXES — reality screen major issues ──────────
    ("R1", "Red box (reality screen)", "red",
     "Major: hard cut + lean + extreme training",
     "pace=Lose hard AND bodyFat=Lean AND activity=Extremely Active",
     "You're already lean and training extremely hard. Your body has nothing extra to burn — a hard deficit pulls from muscle and recovery, and your sessions degrade within 2 weeks. Drop to Lose Steady; you'll keep what you've built and your training stays sharp through the cut."),

    ("R2", "Red box (reality screen)", "red",
     "Major: hard cut + lean (less extreme activity)",
     "pace=Lose hard AND bodyFat=Lean AND activity < Extremely Active (R1 catches the more specific case first)",
     "When body fat is already low, hard deficits run out of fat to burn and start cannibalizing muscle. Lose Steady protects the lean tissue you've worked for and gets you to your next single-digit BF on a path you can actually sustain."),

    ("R3", "Red box (reality screen)", "red",
     "Major: any gain + high BF",
     "pace=Gain steady OR Gain hard AND bodyFat=High",
     "Adding mass on top of high body fat means most of the gain goes to fat too, which loads joints and dulls insulin sensitivity. Switch to Lose Steady for 2 months — your composition shifts, then a clean bulk lands on better ground."),

    ("R4", "Red box (reality screen)", "red",
     "Major: hard surplus + sedentary/light + not high BF",
     "pace=Gain hard AND activity=Sedentary OR Lightly Active AND bodyFat != High (R3 catches the high-BF case first)",
     "Without a consistent training stimulus, your body has no signal to build muscle, so a hard surplus just stores as fat. Drop to Gain Steady and add 2 lift sessions a week — by next month your body knows what to do with the extra calories."),

    # ────────── 3. AMBER BOXES — reality screen caution issues ──────────
    ("R5", "Amber box (reality screen)", "amber",
     "Caution: Performance preset + sedentary/light",
     "macro=Performance AND activity <= Lightly Active",
     "Performance carbs assume daily glycogen burn from training. Without that work, the extra carbs just store as fat. Start with Balanced; switch when you're training 4-5 days a week and feel a real difference from the extra carbs."),

    ("R6", "Amber box (reality screen)", "amber",
     "Caution: High-Protein preset + high BF",
     "macro=High-Protein AND bodyFat=High",
     "Extra protein at high body fat has nowhere productive to go without a strength stimulus. Start with Balanced and add 2-3 lift sessions a week — at that point High-Protein becomes muscle-building instead of expensive maintenance."),

    ("R7", "Amber box (reality screen)", "amber",
     "Caution: Keto + extreme training",
     "macro=Keto AND activity=Extremely Active",
     "High-intensity training pulls from glycogen, and Keto keeps glycogen low. Most extremely-active athletes get more out of Performance once they're past 3+ sessions a week."),

    ("R8", "Amber box (reality screen)", "amber",
     "Caution: Maintain + high BF",
     "pace=Maintain AND bodyFat=High",
     "Maintenance holds you exactly where you are — including the joint loading and insulin resistance that come with high body fat. Lose Steady drops body fat at ~1% body weight per week, which is sustainable and starts making daily life feel easier within a month."),

    ("R9", "Amber box (reality screen)", "amber",
     "Caution: Gain Steady + sedentary/light + NOT high BF",
     "pace=Gain Steady AND activity=Sedentary OR Lightly Active AND bodyFat != High",
     "Even a gentle surplus needs a training stimulus to become muscle. Without lift sessions, most of the gain just stores as fat. Switch to Maintain until you're consistently training 2-3 days a week — then Gain Steady actually produces composition change."),

    ("R10", "Amber box (reality screen)", "amber",
     "Caution: Maintain + average BF + sedentary",
     "pace=Maintain AND bodyFat=Average AND activity=Sedentary",
     "Maintaining at sedentary average means daily life slowly drifts you toward higher body fat — same calories, less movement, creeping accumulation over years. Lose Steady drops body fat at ~1% body weight per week without changing your routine drastically."),

    ("R11", "Amber box (reality screen)", "amber",
     "Caution: Lose Hard + average BF + extreme training",
     "pace=Lose hard AND bodyFat=Average AND activity=Extremely Active",
     "Training extremely hard demands recovery, and a 25% deficit pulls from that pool — sessions feel harder and recovery takes longer even with average body fat to burn. Lose Steady lets you keep training quality while still trimming consistently."),

    ("R12", "Amber box (reality screen)", "amber",
     "Caution: Gain Hard + average BF + moderate training",
     "pace=Gain hard AND bodyFat=Average AND activity=Moderately Active",
     "At moderate training (3-5 days a week), a +15% surplus partitions about 40% lean / 60% fat — you're adding more fat than muscle. Gain Steady (+10%) keeps the surplus small enough that more of it lands as actual mass at this training level."),

    # ────────── 4. APPLY CARD (reality screen footer) ──────────
    ("A1", "Apply card", "neutral",
     "Apply label — one change",
     "Consolidated suggestion has 1 changed field",
     "Switch to {label1}"),

    ("A2", "Apply card", "neutral",
     "Apply label — two changes",
     "Consolidated suggestion has 2 changed fields",
     "Switch to {label1} + {label2}"),

    ("A3", "Apply card", "neutral",
     "Apply label — three+ changes",
     "Consolidated suggestion has 3 changed fields (max — pace + macro + activity)",
     "Switch to {label1}, {label2} + {label3}"),

    ("A4", "Apply card", "neutral",
     "Rationale line — one change",
     "Consolidated suggestion has 1 changed field",
     "One change brings your plan in line with where you're starting."),

    ("A5", "Apply card", "neutral",
     "Rationale line — two changes",
     "Consolidated suggestion has 2 changed fields",
     "Two changes bring your plan in line with where you're starting."),

    ("A6", "Apply card", "neutral",
     "Rationale line — three+ changes",
     "Consolidated suggestion has 3 changed fields",
     "A few changes bring your plan in line with where you're starting."),

    ("A7", "Apply card", "neutral",
     "Apply button text",
     "Always (when consolidatedSuggestion is non-null)",
     "Apply suggested changes"),

    ("A8", "Apply card", "neutral",
     "Keep-your-choices note (below apply button)",
     "Always (when consolidatedSuggestion is non-null)",
     "Or tap Save below to keep your original choices — your plan, your call."),

    # ────────── 5. PACE SCREEN inline warning chips ──────────
    # These fire on the Pace selection step (before reality) as
    # gentle nudges so the user doesn't pick a clearly-bad option
    # without seeing the consequence. Same biology as R1-R4 but
    # shorter — they're chips inside a row, not full cards.
    ("P1", "Pace step inline chip", "red",
     "Hard cut + lean + extreme — pace warning chip",
     "Selected pace=Lose hard AND bodyFat=Lean AND activity=Extremely Active",
     "You're already lean and training hard. Your body has nothing extra to burn — a hard deficit pulls from muscle and recovery. Drop to Lose Steady; you'll keep what you've built and your training stays sharp."),

    ("P2", "Pace step inline chip", "red",
     "Hard cut + lean — pace warning chip",
     "Selected pace=Lose hard AND bodyFat=Lean AND activity < Extremely Active",
     "When body fat is already low, hard deficits run out of fat to burn and start cannibalizing muscle. Lose Steady protects the lean tissue you've worked for."),

    ("P3", "Pace step inline chip", "red",
     "Any gain + high BF — pace warning chip",
     "Selected pace=Gain steady OR Gain hard AND bodyFat=High",
     "Adding mass on top of high body fat means most of the gain goes to fat too, which loads joints and dulls insulin sensitivity. Switch to Lose Steady for 2 months — your composition shifts, then a clean bulk lands on better ground."),

    ("P4", "Pace step inline chip", "red",
     "Hard surplus + sedentary/light — pace warning chip",
     "Selected pace=Gain hard AND activity=Sedentary OR Lightly Active",
     "Without a training stimulus, your body has no signal to build muscle, so a hard surplus just stores as fat. Start lifting 2-3x/week first — then a steady surplus partitions much better."),

    # ────────── 6. MACRO SCREEN inline warning chips ──────────
    ("M1", "Macro step inline chip", "amber",
     "Performance + sedentary/light — macro warning chip",
     "Selected macro=Performance AND activity <= Lightly Active",
     "Performance carbs assume daily glycogen burn from training. Without that work, extra carbs just store as fat. Balanced fits where you are — switch when you're training 4-5 days a week."),

    ("M2", "Macro step inline chip", "amber",
     "High-Protein + high BF — macro warning chip",
     "Selected macro=High-Protein AND bodyFat=High",
     "Extra protein at your body fat has nowhere productive to go without a strength stimulus. Add 2-3 lift sessions a week — at that point High-Protein becomes muscle-building instead of expensive maintenance."),
]


def severity_fill(severity: str) -> PatternFill:
    color = {
        "red":     RED_TINT,
        "amber":   AMBER_TINT,
        "green":   GREEN_TINT,
        "neutral": NEUTRAL_TINT,
    }.get(severity, NEUTRAL_TINT)
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def severity_label(severity: str) -> str:
    return {
        "red":     "🔴 RED (major)",
        "amber":   "🟡 AMBER (caution)",
        "green":   "🟢 GREEN (on-track)",
        "neutral": "⚪ NEUTRAL",
    }.get(severity, severity.upper())


def build_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plan Wizard Strings"

    # ── Title row ──
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value="MyRX — Self-coached Plan Wizard Strings")
    title_cell.font = Font(bold=True, size=14, color=WHITE)
    title_cell.fill = PatternFill(start_color=BLACK, end_color=BLACK, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    # ── Subtitle ──
    ws.merge_cells("A2:F2")
    subtitle_cell = ws.cell(row=2, column=1, value=(
        "Fill the 'Your rewrite' column for any string you want to change. "
        "Leave it blank to keep the current text. Hand the file back when done."
    ))
    subtitle_cell.font = Font(italic=True, color=BLACK, size=10)
    subtitle_cell.fill = PatternFill(start_color=SLATE_TINT, end_color=SLATE_TINT, fill_type="solid")
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ws.row_dimensions[2].height = 22

    # ── Column headers (row 3) ──
    headers = ["ID", "Section", "Severity", "Where it appears / When it fires", "Current text", "Your rewrite"]
    header_font = Font(bold=True, color=WHITE, size=11)
    header_fill = PatternFill(start_color=ZINC, end_color=ZINC, fill_type="solid")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = BORDER_ALL
    ws.row_dimensions[3].height = 24

    # ── Data rows ──
    current_section = None
    row_num = 4
    for (entry_id, section, severity, where, trigger, current_text) in ENTRIES:
        # Section divider — render a thin labeled band whenever the section changes
        if section != current_section:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
            band = ws.cell(row=row_num, column=1, value=f"  {section}")
            band.font = Font(bold=True, color=WHITE, size=11)
            band.fill = PatternFill(start_color=ZINC, end_color=ZINC, fill_type="solid")
            band.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row_num].height = 20
            current_section = section
            row_num += 1

        # Combined "where / when" cell — section + trigger stacked on two lines
        where_when = f"{where}\n\nTriggers when: {trigger}"

        row_fill = severity_fill(severity)
        cells = [
            (entry_id,                Alignment(horizontal="center", vertical="top")),
            (section,                 Alignment(horizontal="left",   vertical="top", wrap_text=True, indent=1)),
            (severity_label(severity),Alignment(horizontal="left",   vertical="top", indent=1)),
            (where_when,              Alignment(horizontal="left",   vertical="top", wrap_text=True, indent=1)),
            (current_text,            Alignment(horizontal="left",   vertical="top", wrap_text=True, indent=1)),
            ("",                      Alignment(horizontal="left",   vertical="top", wrap_text=True, indent=1)),
        ]
        for col_idx, (val, align) in enumerate(cells, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill = row_fill
            cell.alignment = align
            cell.border = BORDER_ALL
            cell.font = Font(size=10, color=BLACK)
        # Auto-set row height roughly proportional to longest text length.
        text_len = max(len(where_when), len(current_text))
        # ~85 chars per line at our column width; add 28 px per wrapped line.
        lines = max(1, (text_len // 60) + 1)
        ws.row_dimensions[row_num].height = max(28, lines * 16)
        row_num += 1

    # ── Column widths ──
    ws.column_dimensions["A"].width = 8     # ID
    ws.column_dimensions["B"].width = 24    # Section
    ws.column_dimensions["C"].width = 22    # Severity
    ws.column_dimensions["D"].width = 42    # Where / trigger
    ws.column_dimensions["E"].width = 70    # Current text
    ws.column_dimensions["F"].width = 70    # Your rewrite

    # Freeze title + header rows so they stay visible as you scroll.
    ws.freeze_panes = "A4"

    return wb


def main() -> None:
    out_path = Path("C:/Users/motaz/OneDrive/Desktop/MyRX/docs/plan_wizard_strings.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    # On Windows, an Excel-open XLSX is exclusively locked. Try the
    # canonical filename first; if it fails with a permission error,
    # write a date-suffixed variant so the user gets fresh content
    # without having to close Excel.
    try:
        wb.save(out_path)
        target = out_path
    except PermissionError:
        from datetime import datetime
        suffix = datetime.now().strftime("%Y%m%d-%H%M")
        target = out_path.with_name(f"plan_wizard_strings_{suffix}.xlsx")
        wb.save(target)
        print(f"  (canonical file was open in Excel — wrote versioned copy instead)")
    print(f"Wrote {target.resolve()}")
    print(f"  rows: {len(ENTRIES)} string entries")


if __name__ == "__main__":
    main()
