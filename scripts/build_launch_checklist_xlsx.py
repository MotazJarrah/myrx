"""
Build docs/launch_checklist.xlsx — complete go-live checklist for MyRX.

Six phase sheets, each with categorized items:
  1. T-30 days  — Foundation (legal, infra, monitoring, app store enrollment)
  2. T-7  days  — Final QA (one last pass before submitting)
  3. T-0       — Launch day (deploy sequence + announce)
  4. T+1  day   — Day-after monitoring
  5. T+7  days  — Week-one retrospective
  6. T+30 days  — First-month cost + metrics review

Each item: Category | Item | Why it matters | Owner | Done? | Notes

Run from repo root:
    python scripts/build_launch_checklist_xlsx.py
"""
from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "launch_checklist.xlsx"

# ─────────── styles ────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="F8FAFC", size=11)
CATEGORY_FILLS = {
    "Legal":      PatternFill("solid", fgColor="E0E7FF"),  # indigo
    "Tech":       PatternFill("solid", fgColor="DBEAFE"),  # blue
    "App Store":  PatternFill("solid", fgColor="F3E8FF"),  # purple
    "Payments":   PatternFill("solid", fgColor="DCFCE7"),  # green
    "Email":      PatternFill("solid", fgColor="FEF3C7"),  # amber
    "Comms":      PatternFill("solid", fgColor="FFEDD5"),  # orange
    "Marketing":  PatternFill("solid", fgColor="FCE7F3"),  # pink
    "Monitoring": PatternFill("solid", fgColor="CFFAFE"),  # cyan
    "Support":    PatternFill("solid", fgColor="FFE4E6"),  # rose
    "Deploy":     PatternFill("solid", fgColor="E5E7EB"),  # slate
    "Announce":   PatternFill("solid", fgColor="FEF9C3"),  # yellow
    "Retro":      PatternFill("solid", fgColor="F5F5F4"),  # stone
    "Finance":    PatternFill("solid", fgColor="D1FAE5"),  # emerald
}
DONE_FILLS = {
    "":          PatternFill("solid", fgColor="FFFFFF"),
    "Pending":   PatternFill("solid", fgColor="FEF3C7"),
    "In Progress": PatternFill("solid", fgColor="DBEAFE"),
    "Done":      PatternFill("solid", fgColor="DCFCE7"),
    "Skip":      PatternFill("solid", fgColor="F5F5F4"),
    "Blocked":   PatternFill("solid", fgColor="FEE2E2"),
}
THIN = Side(border_style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header_row(ws, row_idx, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row_idx].height = 28

def add_phase_sheet(wb, name, subtitle, items):
    ws = wb.create_sheet(name)
    ws["A1"] = name
    ws["A1"].font = Font(bold=True, size=14, color="1F2937")
    ws.merge_cells("A1:F1")
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=10, italic=True, color="475569")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 36

    cols = ["Category", "Item", "Why it matters", "Owner", "Status", "Notes"]
    ws.append([])  # spacer row 3
    ws.append(cols)
    style_header_row(ws, 4, len(cols))

    for i, (cat, item, why, owner, status, notes) in enumerate(items):
        excel_row = 5 + i
        ws.cell(row=excel_row, column=1, value=cat)
        ws.cell(row=excel_row, column=2, value=item)
        ws.cell(row=excel_row, column=3, value=why)
        ws.cell(row=excel_row, column=4, value=owner)
        ws.cell(row=excel_row, column=5, value=status)
        ws.cell(row=excel_row, column=6, value=notes)
        for c in range(1, 7):
            cell = ws.cell(row=excel_row, column=c)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = BORDER
            cell.font = Font(size=10)
        if cat in CATEGORY_FILLS:
            ws.cell(row=excel_row, column=1).fill = CATEGORY_FILLS[cat]
            ws.cell(row=excel_row, column=1).font = Font(size=10, bold=True)
        if status in DONE_FILLS:
            ws.cell(row=excel_row, column=5).fill = DONE_FILLS[status]
            ws.cell(row=excel_row, column=5).font = Font(size=10, bold=True)

    widths = {1: 13, 2: 44, 3: 44, 4: 12, 5: 14, 6: 38}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A5"

# ═══════════════ Workbook ═════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)

# ─────────── Overview ──────────────────────────────────────────────────────────
ov = wb.create_sheet("Overview")
ov["A1"] = "MyRX Launch Checklist — Overview"
ov["A1"].font = Font(bold=True, size=16, color="1F2937")
ov.merge_cells("A1:C1")

intro = (
    "Six-phase checklist for going live with the coach platform v1 + B2C tiers. "
    "Each phase has its own sheet. Work through phases in order — Phase N+1 assumes "
    "Phase N is mostly green.\n\n"
    "Owner column: 'M' = Motaz (user), 'C' = Claude (assistant), 'L' = Lawyer (if engaged). "
    "Status: blank → Pending → In Progress → Done. Or Skip (intentional pass) / Blocked (dependency)."
)
ov["A3"] = intro
ov["A3"].alignment = Alignment(wrap_text=True, vertical="top")
ov["A3"].font = Font(size=10)
ov.merge_cells("A3:C3")
ov.row_dimensions[3].height = 90

phases = [
    ("Phase 1 — T-30 days",  "Foundation (legal sign-off, infra hardening, app store enrollment, payment go-live setup)"),
    ("Phase 2 — T-7 days",   "Final QA pass + mobile submissions (allow 1-2 weeks for App Store, 2-3 days for Play)"),
    ("Phase 3 — Launch day", "Deploy sequence + announce (1-hr active monitoring)"),
    ("Phase 4 — T+1 day",    "Day-after metrics + support inbox triage"),
    ("Phase 5 — T+7 days",   "Week-one retrospective + bug triage from first reports"),
    ("Phase 6 — T+30 days",  "First-month cost review + metrics + vendor plan adjustments"),
]
ov.append([])
ov.append(["Phase", "Sheet name", "Focus"])
style_header_row(ov, 5, 3)
for i, (phase, focus) in enumerate(phases):
    ov.cell(row=6+i, column=1, value=phase)
    ov.cell(row=6+i, column=2, value=phase)
    ov.cell(row=6+i, column=3, value=focus)
    for c in range(1, 4):
        ov.cell(row=6+i, column=c).font = Font(size=10)
        ov.cell(row=6+i, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        ov.cell(row=6+i, column=c).border = BORDER

ov.column_dimensions["A"].width = 24
ov.column_dimensions["B"].width = 24
ov.column_dimensions["C"].width = 70

# ─────────── Phase 1 — T-30 days (Foundation) ─────────────────────────────────
phase1 = [
    # Legal
    ("Legal", "All 9 legal docs reviewed for accuracy", "Stale prose = liability. Esp. Privacy + Refund.", "M+L", "Pending",
     "Files in docs/legal/*.docx. Self-review at minimum; lawyer review if scope changes (B2C launch).") ,
    ("Legal", "GDPR data export endpoint works for any user",
     "Required if any EU user. Privacy Policy promises it.",
     "C", "Pending",
     "Test by creating a throwaway account → request data export → verify all fields exported."),
    ("Legal", "Data deletion (hard delete) flow tested",
     "Right to be forgotten — GDPR Article 17.",
     "C", "Pending",
     "Already implemented (task #138). Re-test end-to-end before launch."),
    ("Legal", "Cookie banner verified on every marketing page",
     "Cookie Policy promises consent before non-essential cookies fire.",
     "M", "Pending",
     "Inspect each landing page: cookie banner appears, declining works, analytics doesn't fire."),
    ("Legal", "DPA download link works from public site",
     "EU coaches need to download + countersign for GDPR compliance.",
     "M", "Pending",
     "Currently the DPA is at /dpa as a webpage. Add a download-as-PDF link if a coach asks for a signable version."),
    ("Legal", "Coach Agreement checkbox enforced at signup",
     "Without it, coaches haven't legally accepted ToS — no liability shield.",
     "C", "Done",
     "Already shipped (task #171). Just confirm it still works after any signup-flow changes."),

    # Tech
    ("Tech", "Cloudflare Pages production stable for 7+ days no incident",
     "Baseline reliability check before promoting to launch.",
     "C", "Pending",
     "Watch the production deploy at myrxfit.com. Track build success rate, any 5xx spikes."),
    ("Tech", "Supabase production tier capacity check",
     "Free tier: 500MB DB + 1GB storage + 2GB bandwidth. Likely fine for launch but check.",
     "M+C", "Pending",
     "Dashboard → Project Settings → Usage. If close to limits, upgrade to Pro ($25/mo)."),
    ("Tech", "Backup strategy verified",
     "Need point-in-time recovery enabled OR weekly DB dumps.",
     "C", "Pending",
     "Supabase Pro includes PITR. Free tier: no built-in backups → need pg_dump cron."),
    ("Tech", "Mobile dev-client APK rebuilt against latest source",
     "Catch any native-module breakage before App Store submission.",
     "C", "Pending",
     "From /mobile: npx expo run:android (and equivalent run:ios when Apple Dev approved)."),

    # App Store
    ("App Store", "Apple Developer Program enrolled + verified",
     "Without it, can't submit to App Store.",
     "M", "Pending",
     "$99/yr. LLC verification can take 1-4 weeks."),
    ("App Store", "Google Play Developer Account enrolled",
     "Without it, can't submit to Play Store.",
     "M", "Pending",
     "$25 one-time."),
    # ── Download App page (the ONLY athlete web surface) ─────────────────────
    ("App Store", "Replace web /app placeholder with full download landing page",
     "Right now /app is a holding page (text only — 'MyRX is mobile-first … coming soon'). Once iOS + Android apps ship, swap to the full launch page.",
     "M+C", "Pending",
     "File: web/src/pages/DownloadAppPlaceholder.jsx. Component is intentionally simple right now; everything below is what gets added at app-launch time."),
    ("App Store", "/app page: App Store badge + click-through",
     "Apple's official 'Download on the App Store' SVG badge. Click → opens App Store page for MyRX.",
     "M+C", "Pending",
     "Apple brand guidelines: https://developer.apple.com/app-store/marketing/guidelines/ — use the official asset, don't redraw."),
    ("App Store", "/app page: Play Store badge + click-through",
     "Google's official 'Get it on Google Play' SVG badge. Click → opens Play Store page for MyRX.",
     "M+C", "Pending",
     "Google brand guidelines: https://play.google.com/intl/en_us/badges/ — use the official asset."),
    ("App Store", "/app page: QR code that deep-links to the install URL",
     "QR code visible to users on desktop. Scan with phone → opens whichever store matches their OS (App Store for iOS, Play Store for Android). Use a service like qrserver.com or generate at build time.",
     "M+C", "Pending",
     "Should be a smart link that detects OS — or two QR codes side by side (iOS / Android)."),
    ("App Store", "/app page: hero copy and screenshots",
     "Replace the 'coming soon' card with real product framing — 1 hero screenshot of the app + 3 supporting screenshots showing strength / cardio / coach view.",
     "M", "Pending",
     "Designer pass or simple grid; Apple/Google store screenshots can be reused."),
    ("App Store", "/app page: invite-token deep link (when arriving via /coach/accept-invite?token=...)",
     "If the user arrived because they clicked an email invite link from desktop, the QR code / smart link must include the invite token so the mobile app's signup flow can pick it up. Universal Link / App Link on the mobile side handles this — page just has to pass the token through to the install URL.",
     "M+C", "Pending",
     "Currently /coach/accept-invite captures the token in URL. /app needs to read URL params and append to install link."),
    ("App Store", "/app page: support email (team@myrxfit.com) visible",
     "If a user has trouble installing / signing in, give them a way to reach support.",
     "M", "Pending",
     "Footer link is fine; don't make it loud."),
    ("App Store", "/app page: 'sign in here' link to /for-coaches for misrouted coaches/admins",
     "Already present in the placeholder — confirm it stays in the full launch version. Coach/admin who somehow lands on /app needs an obvious way back to the coach sign-in.",
     "C", "Done",
     "Already in the placeholder. Just preserve when redesigning."),
    ("App Store", "/app page: legal links footer (Privacy, ToS)",
     "Footer or bottom-of-page links. Required for App Store + Google Play app-marketing-page review.",
     "M+C", "Pending",
     "Standard legal-footer pattern; mirror what /for-coaches has."),
    ("App Store", "/app page: hreflang / SEO meta if marketing internationally",
     "If the app launches in non-English markets, the page needs translated copy + hreflang tags. Defer if launch is English-only initially.",
     "M", "Pending",
     "Skip for English-only launch."),
    ("App Store", "App icons (all required sizes for iOS + Android)",
     "Both stores reject submissions with missing/wrong-size icons.",
     "M+C", "Pending",
     "iOS: 1024x1024 marketing + 180/167/152/120 in-app. Android: 512x512 store + 192 launcher."),
    ("App Store", "Screenshots for store listings (5 minimum per platform)",
     "Conversion lever — best screenshots have the value prop in the first frame.",
     "M", "Pending",
     "iOS: 6.7\" + 6.5\" + 5.5\" sizes. Android: phone + 7\" tablet + 10\" tablet."),
    ("App Store", "Store listing text written (description, keywords, what's new)",
     "SEO inside the store. Determines who finds you.",
     "M", "Pending",
     "Keywords matter most on Apple. Description matters most on Google."),
    ("App Store", "Production keystore SHA256 added to assetlinks.json",
     "Magic-link emails open in app (not browser) on Android. Currently only debug key registered.",
     "C", "Pending",
     "web/public/.well-known/assetlinks.json — add prod SHA256 fingerprint, redeploy web."),
    ("App Store", "TestFlight internal testing — 1 full week with 3+ testers",
     "Catches bugs that don't surface on a single device.",
     "M+C", "Pending",
     "Recruit 2-3 willing friends. Pre-launch is the safest QA window."),
    ("App Store", "Google Play Internal Testing — 1 full week with 3+ testers",
     "Same as above for Android.",
     "M+C", "Pending",
     "Internal testing track is fastest; can iterate same-day on Play."),

    # Payments
    ("Payments", "Stripe Live mode activated",
     "Test keys can't accept real money. Live keys gated on business verification.",
     "M", "Pending",
     "Dashboard → Activate Account. Provide: EIN, bank account for payouts, ID."),
    ("Payments", "Stripe Live webhook endpoint configured",
     "Without this, subscription events (cancel, renew, refund) don't propagate to your DB.",
     "C", "Pending",
     "Stripe → Webhooks → Add endpoint → /functions/v1/stripe-webhook?source=coach_subs. Copy signing secret to STRIPE_WEBHOOK_SECRET_LIVE in Supabase Edge Function secrets."),
    ("Payments", "Test full purchase flow in Stripe Live mode",
     "Use real card with $0.50 charge → confirm DB updates + email sent + refund flow works.",
     "C", "Pending",
     "Refund the test charge afterwards. Document any flow gaps."),
    ("Payments", "Refund-policy response SLA committed",
     "billing@myrxfit.com needs to respond to refund requests within X hours per Refund Policy.",
     "M", "Pending",
     "Document the SLA internally. Even '24 hours' is fine if you stick to it."),

    # Email
    ("Email", "Mail-tester.com score ≥ 9/10",
     "Comprehensive deliverability check before launch.",
     "C", "Pending",
     "Send test from each outbound path to the temp address mail-tester gives. Aim for 10/10."),
    ("Email", "DMARC at p=reject (final state)",
     "Catches end of the 4-week ramp. Sequenced in docs/reminders.xlsx.",
     "C", "Pending",
     "Schedule: pct=100 around Jun 10, reject around Jun 24."),
    ("Email", "Transactional templates reviewed for tone + branding",
     "Coach invite, OTP, password reset all currently functional but generic.",
     "M+C", "Pending",
     "If a designer is engaged, get them to skin the templates."),

    # Monitoring
    ("Monitoring", "Cloudflare alert on 5xx spike",
     "Catches regressions before users report them.",
     "C", "Pending",
     "Cloudflare → Notifications → Add → HTTP 5xx Rate threshold."),
    ("Monitoring", "Supabase log alerts on auth errors",
     "Spike in failed signups = something broken in OTP/email chain.",
     "C", "Pending",
     "Supabase doesn't have native alerts; pipe logs to a third-party (Sentry / BetterStack) OR check daily."),
    ("Monitoring", "Stripe webhook failure alert",
     "Failed webhooks = silent subscription-state drift between Stripe + your DB.",
     "C", "Pending",
     "Stripe → Webhooks → endpoint → Alert on failure threshold."),
    ("Monitoring", "Frontend error tracking enabled (Sentry or similar)",
     "Without it, JS errors in production are invisible.",
     "M+C", "Pending",
     "Sentry free tier covers small launches. Wire into web + mobile."),

    # Support
    ("Support", "support@myrxfit.com tested + auto-reply set up",
     "First impression for users who hit issues.",
     "M", "Pending",
     "Set auto-reply in Zoho: 'Got your message, replying within X hours.'"),
    ("Support", "Customer support response process documented",
     "Internal doc so you (or future hires) know what to do when X email lands.",
     "M", "Pending",
     "Simple wiki page is fine. Categorize: billing / bug / feature request / abuse."),

    # Marketing
    ("Marketing", "Landing page CTAs work + analytics fires",
     "Conversion measurement is impossible without working analytics.",
     "M+C", "Pending",
     "Test from incognito: click each CTA, verify the funnel event fires (if analytics added)."),
    ("Marketing", "Pricing page accurate + matches Stripe products",
     "Mismatched prices = chargeback risk + bad reviews.",
     "M", "Pending",
     "Cross-check web/src/pages/Pricing.jsx + Stripe Product IDs in .env.example."),
    ("Marketing", "SEO basics: title tags, meta descriptions, OG images",
     "Determines what shows up in Google + social shares.",
     "M+C", "Pending",
     "Check every public page has unique <title> and <meta description>."),
    ("Marketing", "Domain transferred to Cloudflare Registrar (if Jul 6 has passed)",
     "Saves ~$10/yr + simpler vendor relationship.",
     "M+C", "Pending",
     "Triggered by docs/reminders.xlsx Jul 6 deadline. Optional — can defer."),
]
add_phase_sheet(wb, "Phase 1 — T-30 days",
                "Foundation work. Most items are blocking — incomplete legal/payments/app-store items will delay launch.",
                phase1)

# ─────────── Phase 2 — T-7 days (Final QA) ─────────────────────────────────────
phase2 = [
    ("Tech", "Full regression test on every authed surface",
     "Catches anything broken since Phase 1 lock.",
     "M+C", "Pending",
     "Walk through: signup → onboarding → log strength → log cardio → log food → settings → delete account."),
    ("Tech", "Synthetic load test (3-5 concurrent users)",
     "Cheap sanity check that Supabase doesn't choke on the launch spike.",
     "C", "Pending",
     "Even just opening 5 incognito windows + doing things simultaneously catches obvious bottlenecks."),
    ("Email", "Send each of 5 templates: signup, password reset, coach invite, OTP, magic link",
     "Confirm every transactional path works AND lands in inbox (not spam).",
     "M+C", "Pending",
     "Use a fresh Gmail + Hotmail + Yahoo account for each. Check all 3 inboxes for each template."),
    ("Email", "DMARC final-state confirmed (p=reject)",
     "Locks down spoofing risk before launch traffic.",
     "C", "Pending",
     "Final step in the ramp. See docs/reminders.xlsx → Jun 24 deadline."),
    ("Payments", "Stripe live: full purchase + cancel + refund flow tested",
     "Production payment flow has its own edge cases (3D Secure, declined cards).",
     "M+C", "Pending",
     "Test with a real card (small charge → refund afterwards). NOT test cards on live mode."),
    ("App Store", "iOS App Store submission lodged",
     "Apple review takes 1-2 weeks typically. Submit at T-7 to land before T-0.",
     "M+C", "Pending",
     "Xcode Archive → Distribute → App Store Connect → Submit for Review."),
    ("App Store", "Google Play submission lodged",
     "Play review is faster (~2-3 days). Submit at T-7 = comfortable buffer.",
     "M+C", "Pending",
     "Play Console → Production track → upload AAB → roll out."),
    ("Monitoring", "All alert channels tested (Slack/email/SMS)",
     "Alerts only useful if they actually reach you.",
     "M", "Pending",
     "Trigger a fake alert from each system → confirm notification arrives."),
    ("Comms", "Launch announcement copy drafted (email + social)",
     "Drafted now so you're not writing it under launch-day pressure.",
     "M", "Pending",
     "Have 2-3 variants ready. Soft-launch tone vs. excited tone."),
    ("Comms", "FAQ page or help center seeded with top expected questions",
     "Deflects support load. Top FAQs: pricing, how to cancel, what data you store, refund process.",
     "M", "Pending",
     "Doesn't need to be elaborate. 10-15 Q&As is plenty for launch."),
]
add_phase_sheet(wb, "Phase 2 — T-7 days",
                "Final QA + lock-down. After this phase, no new features until post-launch.",
                phase2)

# ─────────── Phase 3 — Launch day (T-0) ────────────────────────────────────────
phase3 = [
    ("Deploy", "Switch STRIPE_MODE from test → live in production env vars",
     "Real payments start flowing. Test mode payments fail silently in production.",
     "C", "Pending",
     "Stripe edge function reads STRIPE_MODE. Confirmed live via a $0.50 test charge."),
    ("Deploy", "Cloudflare Pages production on latest source commit",
     "Run wrangler pages deploy web/dist before announce email goes out.",
     "C", "Pending",
     "Verify with: curl -s https://myrxfit.com/ | grep -oE 'index-[^\"]+\\.js' vs local dist/."),
    ("App Store", "iOS app release approved + published (manual or auto-release)",
     "Push the button to make it discoverable in App Store search.",
     "M", "Pending",
     "Apple sometimes auto-releases on approval, sometimes waits for manual push."),
    ("App Store", "Google Play release rolled out to 100% production track",
     "Same for Android.",
     "M", "Pending",
     "Play Console → Production → Manage release → publish."),
    ("Announce", "Launch email sent to mailing list (if exists)",
     "Largest traffic spike of the day, typically.",
     "M", "Pending",
     "Send from team@myrxfit.com via SendGrid (so it's DKIM-signed)."),
    ("Announce", "Social media posts published (LinkedIn, X, Instagram)",
     "Multi-channel for max reach.",
     "M", "Pending",
     "Pre-scheduled at T-0 morning, or live-posted by user."),
    ("Monitoring", "1-hour active live monitoring window after announce",
     "First-hour traffic spike is where things break.",
     "M+C", "Pending",
     "Watch CF Pages dashboard + Supabase logs + Stripe events simultaneously."),
    ("Support", "Inbox actively monitored (1-hour SLA day-of)",
     "First impression matters. Quick responses on launch day pay back for months.",
     "M", "Pending",
     "Plan for: ~5-20 emails depending on launch reach."),
]
add_phase_sheet(wb, "Phase 3 — Launch day",
                "Go-live execution. Sequence matters: payments live BEFORE announce email.",
                phase3)

# ─────────── Phase 4 — T+1 day ─────────────────────────────────────────────────
phase4 = [
    ("Monitoring", "Check error rate vs baseline",
     "Any 5xx spike, JS errors, failed signups need same-day fix.",
     "C", "Pending",
     "Compare to Phase 1 baseline."),
    ("Monitoring", "Check signup conversion rate",
     "If conversion < 30%, something's broken in the funnel.",
     "M+C", "Pending",
     "Analytics → funnel report → drop-off points."),
    ("Payments", "Check payment success rate",
     "Stripe Dashboard → Payments → failed declined %.",
     "M", "Pending",
     "Some decline rate is normal. >5% suggests 3D Secure issue or pricing-page UX problem."),
    ("Email", "Check DMARC aggregate report inbox",
     "First post-launch report shows whether real-world spoof attempts exist + whether legit mail still passes.",
     "C", "Pending",
     "dmarc_rua@onsecureserver.net → forwarded summaries."),
    ("Support", "Respond to every email in inbox",
     "Day-1 response time signals how serious you are. Aim for sub-12-hour.",
     "M", "Pending",
     "Categorize as you respond: bug / billing / feature / abuse."),
]
add_phase_sheet(wb, "Phase 4 — T+1 day",
                "Day-after triage. Catch any silent regressions before they snowball.",
                phase4)

# ─────────── Phase 5 — T+7 days ────────────────────────────────────────────────
phase5 = [
    ("Retro", "What worked / what didn't — written retro",
     "Lessons compound. Capture them while fresh.",
     "M+C", "Pending",
     "30 min, free-form. Save to docs/retros/T+7-launch-retro.md or similar."),
    ("Tech", "Triage bugs reported in first week",
     "Triage matters more than fixing everything. Categorize by severity.",
     "M+C", "Pending",
     "Critical = data loss / payment broken. High = blocks signup. Medium = minor flow issue. Low = cosmetic."),
    ("Marketing", "Adjust marketing based on actual traffic data",
     "What channels delivered? What didn't? Rebalance spend.",
     "M", "Pending",
     "Analytics → acquisition → channel report."),
    ("Comms", "Personal thank-you email to first 10-20 signups",
     "Highest-leverage customer-success move on a small launch.",
     "M", "Pending",
     "Manual send from team@myrxfit.com. Pure relationship-building."),
]
add_phase_sheet(wb, "Phase 5 — T+7 days",
                "Week-one stabilization + customer outreach.",
                phase5)

# ─────────── Phase 6 — T+30 days ───────────────────────────────────────────────
phase6 = [
    ("Finance", "Vendor cost review",
     "First-month bills land — surprises here can blow runway.",
     "M+C", "Pending",
     "Supabase, Cloudflare egress, SendGrid, Twilio, Stripe processing. Compare to forecast."),
    ("Tech", "Vendor capacity check — any nearing limits?",
     "Supabase free tier limits hit silently. Cloudflare has bandwidth allowances.",
     "C", "Pending",
     "If any vendor at >70% capacity, plan upgrade for month 2."),
    ("Finance", "First quarterly metrics review",
     "MRR, ARPU, churn, LTV, CAC. Even rough numbers are useful.",
     "M+C", "Pending",
     "Don't over-engineer the dashboard. Spreadsheet is fine for the first review."),
    ("Marketing", "Iterate on signup funnel based on 30 days of data",
     "Whatever the lowest-converting step is gets the next month's attention.",
     "M+C", "Pending",
     "Common offenders: pricing page, onboarding length, email confirmation friction."),
    ("Comms", "Customer feedback loop — survey or interviews",
     "30 days of usage is enough for users to have opinions.",
     "M", "Pending",
     "5-question survey OR 5 30-min user interviews. Both valuable; interviews give more depth."),
    ("Legal", "Quarterly legal docs review (per docs/reminders.xlsx)",
     "Catches stale prices, vendor names, contact addresses.",
     "M", "Pending",
     "Walk through docs/legal/*.docx for accuracy."),
]
add_phase_sheet(wb, "Phase 6 — T+30 days",
                "First-month review. Set the rhythm for monthly-cycle ops going forward.",
                phase6)

# ─────────── Save ──────────────────────────────────────────────────────────────
OUT.parent.mkdir(parents=True, exist_ok=True)
try:
    wb.save(OUT)
    print(f"Wrote {OUT}")
except PermissionError:
    alt = OUT.with_name("launch_checklist_v2.xlsx")
    wb.save(alt)
    print(f"[WARN] {OUT.name} is open (Excel locked it).")
    print(f"       Wrote {alt} instead. Close Excel + re-run OR rename {alt.name} -> {OUT.name}.")
print(f"Sheets: {wb.sheetnames}")
