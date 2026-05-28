"""
Build docs/testing/coach_invites.xlsx — behavioral E2E test scenarios for the
coach invite flow.

Single sheet of 34 scenarios that exercise the BEHAVIORAL DIFFERENCES between
the three MyRX account roles — Coach, Athlete, Admin — interacting through
the invite flow. UI snapshots / form-field-visibility checks are intentionally
NOT included; those are not test scenarios.

Each row:
  ID | Scenario | Preconditions | Steps | Expected | Failure modes | Priority

Run from repo root:
    python scripts/build_test_scenarios_xlsx.py
"""
from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "testing" / "coach_invites.xlsx"

# --------- styles -------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="F8FAFC", size=11)

PRIORITY_FILLS = {
    "P0": PatternFill("solid", fgColor="FEE2E2"),  # red-100
    "P1": PatternFill("solid", fgColor="FFEDD5"),  # orange-100
    "P2": PatternFill("solid", fgColor="FEF3C7"),  # amber/yellow-100
}
PRIORITY_FONTS = {
    "P0": Font(size=10, bold=True, color="991B1B"),  # red-800
    "P1": Font(size=10, bold=True, color="9A3412"),  # orange-800
    "P2": Font(size=10, bold=True, color="854D0E"),  # yellow-800
}

THIN = Side(border_style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = ["ID", "Scenario name", "Preconditions", "Steps", "Expected result",
        "Failure modes to watch for", "Priority"]

# Column widths (1-indexed). Steps + Expected are the wide ones.
WIDTHS = {1: 8, 2: 36, 3: 38, 4: 58, 5: 46, 6: 38, 7: 10}


def style_header_row(ws, row_idx, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row_idx].height = 28


def add_scenario_sheet(wb, sheet_name, subtitle, rows):
    """rows: list of tuples (id, scenario, preconds, steps, expected, failures, priority)"""
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = sheet_name
    ws["A1"].font = Font(bold=True, size=14, color="1F2937")
    ws.merge_cells("A1:G1")
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=10, italic=True, color="475569")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 56

    ws.append([])  # spacer row 3
    ws.append(COLS)  # row 4
    style_header_row(ws, 4, len(COLS))

    for i, row in enumerate(rows):
        excel_row = 5 + i
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=excel_row, column=c, value=val)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = BORDER
            cell.font = Font(size=10)
        # Priority cell colouring
        prio = row[6]
        if prio in PRIORITY_FILLS:
            pc = ws.cell(row=excel_row, column=7)
            pc.fill = PRIORITY_FILLS[prio]
            pc.font = PRIORITY_FONTS[prio]
            pc.alignment = Alignment(horizontal="center", vertical="center")
        # Estimate row height from steps length (rough)
        steps_lines = str(row[3]).count("\n") + 1
        expected_lines = str(row[4]).count("\n") + 1
        failure_lines = str(row[5]).count("\n") + 1
        max_lines = max(steps_lines, expected_lines, failure_lines, 3)
        ws.row_dimensions[excel_row].height = max(40, min(max_lines * 16, 260))

    # Column widths
    for col_idx, w in WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Freeze header + ID column
    ws.freeze_panes = "B5"

    # Auto-filter on header row
    last_col = get_column_letter(len(COLS))
    last_row = 4 + len(rows)
    ws.auto_filter.ref = f"A4:{last_col}{last_row}"


# ============== Workbook ======================================================
wb = Workbook()
wb.remove(wb.active)

# --------- The 34 behavioral scenarios ----------------------------------------
# Roles in the system:
#   ATHLETE — mobile-only. May be free, self-coached, or attached to a coach.
#   COACH   — web-only. profiles.is_coach=true. Sends invites.
#   ADMIN   — web-only. profiles.is_superuser=true. Owns the platform.

scenarios = [
    # ===== SENDING: coach x invitee account-type matrix (P0) =================
    ("T-001",
     "Coach invites brand-new email (no existing profile)",
     "Coach Sarah is signed in on web with an active subscription. "
     "newuser@example.com has never created a MyRX account and has no row in profiles or coach_invites.",
     "1. Sarah opens her Pending Invites panel.\n"
     "2. Sarah types newuser@example.com.\n"
     "3. Sarah clicks Send.",
     "Invite row created (status=pending, expires_at=now+14d). "
     "SendGrid sends the branded email to the invitee. "
     "Sarah's Pending Invites list shows the new row immediately.",
     "- No row inserted (RLS / constraint issue).\n"
     "- Row inserted but email never sent (silent SendGrid failure).\n"
     "- Expiry not set or set wrong.",
     "P0"),

    ("T-002",
     "Coach invites email of an existing FREE athlete (no coach attached, never opened app)",
     "freeuser@example.com has a profiles row created from a prior signup attempt: "
     "is_coach=false, is_superuser=false, coach_id IS NULL. Athlete has never opened the mobile app.",
     "1. Sarah opens Send Invite.\n"
     "2. Sarah types freeuser@example.com.\n"
     "3. Sarah clicks Send.",
     "Invite goes out (existing profile does NOT block sending). Invite row created with status=pending. "
     "Athlete will see the invite via the banner on their next app launch (see T-031).",
     "- Backend rejects with 'email already in use' (wrong assumption).\n"
     "- Invite sent but never surfaced inside the app (banner watcher broken).",
     "P0"),

    ("T-003",
     "Coach invites email of a SELF-COACHED athlete with training history",
     "selfcoached@example.com is signed up on mobile, has is_self_coached=true, "
     "coach_id IS NULL, and has logged dozens of strength + cardio + bodyweight entries.",
     "1. Sarah opens Send Invite.\n"
     "2. Sarah types selfcoached@example.com.\n"
     "3. Sarah clicks Send.",
     "Invite created and email sent. Backend does not reject just because user is self-coached. "
     "Acceptance flow later flips is_self_coached to false and preserves the calorie plan (see T-027).",
     "- Send rejected with 'user already coached' (misreading is_self_coached as 'has a coach').\n"
     "- Invite sent but app banner never appears.",
     "P0"),

    ("T-004",
     "Coach invites email of an athlete currently on ANOTHER coach's roster",
     "rosterswap@example.com has coach_id = Coach Bob's id. Coach Sarah wants to poach.",
     "1. Sarah opens Send Invite.\n"
     "2. Sarah types rosterswap@example.com.\n"
     "3. Sarah clicks Send.",
     "Invite goes out. Backend does NOT auto-swap the athlete to Sarah on send — it only creates a pending "
     "invite. The swap-confirmation happens at accept-time inside the athlete's mobile app (see T-017).",
     "- Send rejected as 'athlete already coached' (blocks legitimate market behavior).\n"
     "- Backend silently auto-swaps coach_id on send (catastrophic data leak to Sarah without consent).\n"
     "- Bob receives any notification about the attempted invite.",
     "P0"),

    ("T-005",
     "Coach invites email already on THEIR OWN roster",
     "client@example.com has coach_id = Sarah's id. Sarah accidentally tries to invite them.",
     "1. Sarah opens Send Invite.\n"
     "2. Sarah types client@example.com (her own client).\n"
     "3. Sarah clicks Send.",
     "Backend returns 400 with code already_on_roster. UI shows friendly message: "
     "'That client is already on your roster.' No duplicate invite row, no email.",
     "- Duplicate invite created.\n"
     "- Spammy email re-sent to existing client.\n"
     "- Error message is generic 'something went wrong'.",
     "P0"),

    ("T-006",
     "Coach invites email belonging to ANOTHER COACH account",
     "othercoach@example.com has is_coach=true. Sarah tries to invite them as a client.",
     "1. Sarah opens Send Invite.\n"
     "2. Sarah types othercoach@example.com.\n"
     "3. Sarah clicks Send.",
     "Backend returns 400 with code cant_invite_coach. UI shows friendly message: "
     "'That email belongs to a coach. Coaches can't be invited as clients.' No invite row, no email.",
     "- Invite created -> the other coach gets a confusing email asking them to be a client.\n"
     "- If the other coach somehow accepted, role becomes ambiguous (both is_coach and coach_id set).\n"
     "- Error message leaks 'this email is a coach' (mild privacy issue but acceptable here).",
     "P0"),

    ("T-007",
     "Coach invites email belonging to an ADMIN account",
     "admin@myrxfit.com has is_superuser=true. Sarah tries to invite them.",
     "1. Sarah opens Send Invite.\n"
     "2. Sarah types admin@myrxfit.com.\n"
     "3. Sarah clicks Send.",
     "Backend returns 400 with code cant_invite_admin. UI shows friendly message: "
     "'Admin accounts can't be coached.' No invite row, no email.",
     "- Invite created and somehow accepted -> admin loses is_superuser or gains coach_id.\n"
     "- Admin role silently flipped.\n"
     "- Generic 500 instead of friendly 400.",
     "P0"),

    ("T-008",
     "Coach invites email of a DEACTIVATED account",
     "deactivated@example.com had a profile but the account is now disabled "
     "(e.g. deleted_at IS NOT NULL, or status=deactivated).",
     "1. Sarah opens Send Invite.\n"
     "2. Sarah types deactivated@example.com.\n"
     "3. Sarah clicks Send.",
     "Backend returns 400 with code account_deactivated. UI shows friendly message: "
     "'That account is deactivated and can't receive invites.' No invite row, no email.",
     "- Invite created and sent into a black hole (account can never accept).\n"
     "- Backend treats deactivated as free and lets it through.\n"
     "- Reactivation later happens with a stale coach_id from the silent acceptance.",
     "P0"),

    ("T-009",
     "Coach revokes a pending invite, then sends a fresh invite to the same email",
     "Sarah previously sent invite to retry@example.com, then revoked it. Old row status=revoked. "
     "No other pending invite exists for that email from Sarah.",
     "1. Sarah opens Send Invite.\n"
     "2. Sarah types retry@example.com (same email as the revoked one).\n"
     "3. Sarah clicks Send.",
     "New pending row created. Email sent normally. Old revoked row preserved for audit but is NOT used "
     "for the 'already pending' check. Revoked status does NOT block a fresh resend.",
     "- Backend rejects because of the historical revoked row (coach can never re-invite).\n"
     "- Backend resurrects/updates the revoked row instead of creating a new one (audit trail lost).\n"
     "- New email contains the OLD revoked token (link will fail on tap).",
     "P0"),

    ("T-010",
     "Coach sends TWO invites to same email within 14 days (no revoke between)",
     "Sarah sent an invite to dup@example.com 2 days ago. Row is still status=pending, not expired, not revoked.",
     "1. Sarah opens Send Invite again.\n"
     "2. Sarah types dup@example.com.\n"
     "3. Sarah clicks Send.",
     "Backend returns 400 with code invite_already_pending. UI shows friendly message: "
     "'You already have a pending invite to this address — use Resend to re-email.' "
     "No duplicate row created. Existing token unchanged.",
     "- Duplicate pending row created (two valid tokens for same target -> race on accept).\n"
     "- Existing token regenerated (original email's link silently breaks).\n"
     "- No friendly message; raw 409 / 500.",
     "P0"),

    ("T-011",
     "Non-coach user calls the send-invite API directly",
     "Either an ATHLETE (is_coach=false, is_superuser=false) or an ADMIN (is_superuser=true but is_coach=false) "
     "obtains a JWT and tries to call the send-invite edge function manually.",
     "1. Capture an athlete or admin JWT from a real session.\n"
     "2. POST to the send-invite edge function with that JWT + any payload.\n"
     "3. Observe response.",
     "Edge function returns 403 with code coach_required. No invite row created. "
     "Athletes and admins both rejected by the same gate — only is_coach=true callers are allowed.",
     "- Function returns 200 and creates invite (CRITICAL privilege escalation).\n"
     "- Function returns 200 for admin but 403 for athlete (admin shouldn't be able to send invites either).\n"
     "- Function returns 500 (still safe but ugly).",
     "P0"),

    # ===== SENDING: multi-coach interactions (P1) ==========================
    ("T-012",
     "Two different coaches send invites to the SAME email simultaneously",
     "Coach Sarah and Coach Mike are both active. multi@example.com has no profile and no pending invites.",
     "1. Sarah sends an invite to multi@example.com.\n"
     "2. At nearly the same time, Mike sends an invite to multi@example.com.\n"
     "3. Both sends complete.",
     "BOTH invites succeed. Two rows in coach_invites, one per coach, both status=pending. "
     "No cross-coach dedupe — competitive invites are allowed by design. "
     "Acceptance flow will let the invitee pick one (see T-025).",
     "- Mike's send rejected because Sarah got there first (would block competitive market).\n"
     "- Only one row created (whose data won? unspecified behavior).\n"
     "- Race condition leaves a row in a half-written state.",
     "P1"),

    # ===== ACCEPTING: invitee x current-state matrix (P0) ==================
    ("T-013",
     "Brand-new email -> link -> store -> install -> signup with SAME email -> auto-match",
     "Invite sent to neweremail@example.com. That email has no profile yet. "
     "Invitee is on a fresh device with no MyRX app installed.",
     "1. Invitee opens the email and taps the invite link.\n"
     "2. Smart link routes to App Store (iOS) or Play Store (Android).\n"
     "3. Invitee installs MyRX and opens it for the first time.\n"
     "4. Invitee signs up using THE SAME email (neweremail@example.com) via email OTP.\n"
     "5. Invitee completes onboarding.",
     "After auth succeeds, the email-match watcher detects the pending invite. App shows confirm dialog: "
     "'Coach Sarah invited you — accept?' Invitee taps Accept. coach_id is set, banner dismissed, "
     "athlete is now on Sarah's roster.",
     "- Email-match watcher never fires (invite stays patient but invisible).\n"
     "- Confirm dialog appears but Accept does nothing (mutation handler missing).\n"
     "- coach_id set silently with no user confirmation (consent issue).",
     "P0"),

    ("T-014",
     "Brand-new email -> link -> install -> signup with DIFFERENT email -> manual code fallback",
     "Invite sent to original@example.com but invitee signs up with another@example.com on the new install.",
     "1. Invitee taps invite link, gets routed to store, installs app.\n"
     "2. Invitee signs up with another@example.com (NOT the email the invite was sent to).\n"
     "3. Invitee completes onboarding — no banner appears (emails don't match).\n"
     "4. Invitee opens Settings -> 'Have an invite code?'\n"
     "5. Invitee pastes the invite link or token from the original email.",
     "Auto-match did not fire (correctly — emails differ). Fallback flow accepts the pasted token, "
     "verifies it server-side, attaches coach_id to the currently-signed-in user.",
     "- 'Have an invite code?' option missing (invitee permanently stuck).\n"
     "- Fallback flow rejects token because email-on-account doesn't match email-on-invite (too strict for this case).\n"
     "- Fallback flow silently accepts but logs the wrong email to audit trail.",
     "P0"),

    ("T-015",
     "Existing signed-in FREE athlete with history receives an invite -> banner on next launch",
     "Athlete is signed in on mobile, has is_self_coached=false, coach_id IS NULL, and has months of "
     "logged strength PRs, weight log entries, calorie logs, and mobility ROM data. "
     "Coach Sarah sends an invite to the athlete's email while the app is in the background.",
     "1. Background the mobile app (or quit it).\n"
     "2. Wait long enough for the email-match debounce window to elapse (per spec).\n"
     "3. Re-launch the app.\n"
     "4. Banner appears: 'Coach Sarah invited you — accept?'\n"
     "5. Athlete taps Accept.",
     "coach_id set to Sarah. is_self_coached unchanged (was already false). "
     "ALL existing training data (PRs, weight log, calories, mobility) is preserved byte-for-byte — "
     "nothing wiped, nothing reset. Coach can immediately see everything (see T-026).",
     "- Banner never appears (watcher not running on launch).\n"
     "- Accepting wipes history (CATASTROPHIC).\n"
     "- coach_id set but athlete's data not visible to coach (RLS missing the coach branch).",
     "P0"),

    ("T-016",
     "Existing signed-in SELF-COACHED athlete with calorie plan accepts invite",
     "Athlete has is_self_coached=true, coach_id IS NULL, and an active calorie_plans row "
     "(target kcal, macros, goal weight). Receives invite from Sarah.",
     "1. Athlete opens app, sees banner from Sarah.\n"
     "2. Athlete taps Accept.",
     "coach_id set to Sarah. is_self_coached flips to false. "
     "The calorie_plans row IS PRESERVED byte-for-byte — no delete, no reset to defaults. "
     "Coach can later edit it via the admin portal, but the athlete's existing plan is the starting point.",
     "- calorie_plans row deleted (athlete loses their tailored plan).\n"
     "- Plan reset to system defaults.\n"
     "- is_self_coached stays true (flag drift; admin filters may exclude this athlete).",
     "P0"),

    ("T-017",
     "Existing athlete on Coach Bob's roster accepts fresh invite from Coach Sarah (SWAP)",
     "Athlete has coach_id = Bob's id. Sarah sends an invite to that athlete's email.",
     "1. Athlete opens app, sees banner: 'Coach Sarah invited you. You're currently with Coach Bob.'\n"
     "2. Athlete taps Accept.\n"
     "3. Confirmation dialog: 'Switch from Coach Bob to Coach Sarah?'\n"
     "4. Athlete confirms.",
     "coach_id atomically updates from Bob's id to Sarah's id. "
     "Bob immediately loses RLS access to athlete's data. Sarah immediately gains RLS access to ALL "
     "historical data (efforts, food_logs, bodyweight, mobility, chat). "
     "Athlete's training history is preserved unchanged. (See T-028 and T-029 for the verification side.)",
     "- coach_id updated but Bob's queries still return the athlete's data (RLS not enforced).\n"
     "- Athlete data lost on swap (CATASTROPHIC).\n"
     "- Swap happens without the second confirmation dialog (consent issue).\n"
     "- Sarah only sees post-swap data, not pre-swap history.",
     "P0"),

    ("T-018",
     "Athlete already on Coach Sarah's roster taps another Sarah invite (no-op)",
     "Athlete's coach_id is already Sarah's id. Somehow Sarah sent another invite (or the athlete tapped "
     "an old email) and the link/banner reaches the athlete.",
     "1. Athlete taps the duplicate invite (banner or email link).\n"
     "2. App opens the accept-invite handler.",
     "App silently no-ops — accept handler detects coach_id already matches sender and does nothing. "
     "Athlete may see a soft 'You're already on Coach Sarah's roster' confirmation, but no DB mutation, "
     "no duplicate, no error, no spammy banner.",
     "- Re-accept fires a mutation that double-inserts or breaks coach_id constraint.\n"
     "- Athlete shown a scary error.\n"
     "- App crashes on the no-op path.",
     "P1"),

    ("T-019",
     "COACH taps an invite link while signed in as themselves on web",
     "Coach Sarah is signed in on web. She taps her own invite link (e.g. for QA) "
     "OR another coach's invite link sent to her own email.",
     "1. From mail client on the same browser session, tap the invite link.\n"
     "2. Web router opens /coach/accept-invite?token=...",
     "Web shows rejection screen: 'Coaches can't accept invites — sign in as a client account.' "
     "No DB mutation. Sarah's profile is untouched (still is_coach=true, no coach_id added).",
     "- Sarah's profile gets coach_id = some other coach's id (or her own id — self-reference!).\n"
     "- App tries to treat Sarah as a client and crashes.\n"
     "- is_coach silently flipped to false.",
     "P0"),

    ("T-020",
     "ADMIN taps an invite link while signed in on web",
     "Admin (is_superuser=true) is signed in on web. They tap a coach invite link sent to their email.",
     "1. From mail client, tap the link.\n"
     "2. Web router opens /coach/accept-invite?token=...",
     "Web shows rejection screen: 'Admin accounts can't accept coach invites.' "
     "No DB mutation. Admin remains is_superuser=true with no coach_id.",
     "- Admin gets coach_id set (now appears as a client in coach's roster, role confusion).\n"
     "- is_superuser silently flipped to false (admin loses access).\n"
     "- Generic error instead of friendly rejection.",
     "P0"),

    ("T-021",
     "Athlete A signed in but taps invite sent to athlete B (different email)",
     "Athlete A is signed in on mobile with email a@example.com. Invite was sent to b@example.com (different email). "
     "Athlete A somehow received the email (forwarded, shared inbox, QA).",
     "1. Athlete A taps the invite link from inside their mail client.\n"
     "2. App opens accept-invite handler.",
     "App shows: 'This invite was sent to b@example.com. Sign out and sign in as that user to accept.' "
     "No mutation. Athlete A's coach_id unchanged.",
     "- Invite silently attaches to Athlete A despite email mismatch (CRITICAL — wrong person becomes coached).\n"
     "- App shows generic 'invalid token' (invitee can't tell what happened).\n"
     "- App suggests Athlete A change their email to b@example.com (privilege confusion).",
     "P0"),

    ("T-022",
     "Invitee taps an EXPIRED invite link (> 14 days)",
     "coach_invites row exists with expires_at < now. Was status=pending, now effectively expired.",
     "1. Invitee opens the old invite email (15+ days old).\n"
     "2. Taps the link.",
     "App / web shows friendly message: 'Your invite from Coach Sarah expired. Ask them to resend.' "
     "Backend marks invite as expired if not already. No coach_id mutation.",
     "- Expired token still accepted (server-side expiration not enforced).\n"
     "- Generic 'invalid token' error (user confused about whether the link was fake or expired).\n"
     "- App crashes on expired token path.",
     "P0"),

    ("T-023",
     "Invitee taps an invite link the coach REVOKED",
     "coach_invites row has status=revoked. Invitee taps the link from the original email.",
     "1. From mail client, tap link in the revoked invite's email.",
     "App / web shows friendly message: 'This invite is no longer valid.' Token verification "
     "rejects revoked status. No mutation.",
     "- Revoked token still accepted (revoke is cosmetic-only).\n"
     "- Message says 'expired' instead of 'no longer valid' (misleading).\n"
     "- Coach gets no audit trail of the rejected acceptance attempt.",
     "P0"),

    ("T-024",
     "Invitee taps an invite link they ALREADY accepted",
     "Invite was successfully accepted earlier (status=accepted). Same invitee taps the same link again "
     "from the original email.",
     "1. Tap link in the already-accepted invite email.",
     "App / web shows friendly message: 'This invite has already been used.' "
     "No second mutation. Athlete's existing coach_id unchanged.",
     "- Re-accept silently re-fires the mutation (audit log clutter).\n"
     "- Re-accept re-sends welcome notifications to coach.\n"
     "- coach_id gets set again to same value (no harm but a write nonetheless).",
     "P1"),

    ("T-025",
     "Invitee has multiple pending invites from different coaches",
     "Coach Sarah AND Coach Mike BOTH have pending invites to multi@example.com (from T-012).",
     "1. Invitee installs app and signs in with multi@example.com.\n"
     "2. App email-match watcher fires.",
     "App surfaces ALL pending invites: 'You have invites from Coach Sarah and Coach Mike. Pick one or decline all.' "
     "Invitee taps one -> that invite marked accepted, coach_id set; the OTHER invite auto-marked declined "
     "(not orphaned, not still pending).",
     "- Only the most recent invite shown (invitee may miss the older one).\n"
     "- Both shown but both accepted simultaneously (coach_id conflict).\n"
     "- Neither shown (list rendering broken when count > 1).\n"
     "- The unselected invite remains pending forever (orphan).",
     "P1"),

    # ===== DATA PERSISTENCE: the conversion value prop (P0) ================
    ("T-026",
     "Free athlete with 3 months of history accepts invite -> coach immediately sees everything",
     "Athlete has 3 months of logged strength efforts, cardio sessions, weight log, food logs, mobility ROM. "
     "Accepts invite from Coach Sarah (see T-015).",
     "1. Confirm athlete has rich historical data BEFORE accept (note approximate counts).\n"
     "2. Athlete accepts invite.\n"
     "3. Sarah opens her coach portal -> selects this athlete -> navigates each tab "
     "(Profile, Efforts, Bodyweight, Calories, Mobility).",
     "Sarah sees ALL historical data — every effort, every weight entry, every food log, every ROM record. "
     "Charts and PR badges populate from data that pre-dates the acceptance.",
     "- Coach sees empty list (RLS coach-role policy missing on one or more tables).\n"
     "- Coach sees only post-accept data (RLS filters by created_at vs coach_id with wrong logic).\n"
     "- Partial visibility (e.g. efforts visible but food_logs empty -> one policy missing the coach branch).",
     "P0"),

    ("T-027",
     "Self-coached athlete's calorie plan is intact after accept",
     "Athlete has an active calorie_plans row. Accepts invite (see T-016). "
     "Capture the plan values (target kcal, macros, goal weight) BEFORE accept.",
     "1. Note the athlete's calorie_plans values before accept (write down kcal, protein, carbs, fat, goal weight).\n"
     "2. Athlete accepts invite.\n"
     "3. Re-query calorie_plans for that user.\n"
     "4. Coach navigates to athlete's Calories tab in coach portal.",
     "Row is unchanged. is_self_coached flipped to false but plan parameters are byte-for-byte identical. "
     "Coach sees the plan and can edit it from this point forward.",
     "- Plan deleted on accept (CATASTROPHIC user-facing data loss).\n"
     "- Plan reset to system defaults.\n"
     "- is_self_coached still true.\n"
     "- Coach can't see the plan (RLS gap).",
     "P0"),

    ("T-028",
     "After swap: OLD coach loses ALL access to athlete's data (RLS enforcement)",
     "Athlete was on Coach Bob's roster. Swapped to Coach Sarah (T-017).",
     "1. As Coach Bob (signed in separately on web): refresh his client roster page.\n"
     "2. Query the athlete's efforts, food_logs, bodyweight, mobility via the admin RPCs.\n"
     "3. Try to open the athlete's profile via a direct URL in Bob's session.",
     "Athlete no longer appears in Bob's roster. All Bob's queries against that athlete's data return zero rows "
     "(RLS filters by coach_id = auth.uid()). Direct URL access returns 'not found' or empty data.",
     "- Bob still sees the athlete in his roster (cache or refresh issue).\n"
     "- Bob's queries still return data (RLS uses created_at or first-coach-id instead of current coach_id).\n"
     "- Bob can still send messages to the athlete.",
     "P0"),

    ("T-029",
     "After swap: NEW coach sees ALL historical data immediately in the same session",
     "Same scenario as T-028, from Sarah's side. Sarah just received the swapped athlete from Bob.",
     "1. As Coach Sarah: refresh her client roster.\n"
     "2. Click into the swapped athlete.\n"
     "3. Verify Profile, Efforts (including pre-swap PRs), Bodyweight (including pre-swap weigh-ins), "
     "Calories (including pre-swap food logs), Mobility (including pre-swap ROM).",
     "Athlete appears in Sarah's roster. ALL historical data visible immediately — pre-swap, pre-accept, "
     "pre-anything. No data gap, no filter by acceptance date. No need for Sarah to sign out and back in.",
     "- Athlete missing from Sarah's roster (trigger on coach_id update missing).\n"
     "- Sarah sees only data from acceptance date forward (wrong RLS filter).\n"
     "- Sarah has to sign out / back in for the new client to appear.",
     "P0"),

    ("T-030",
     "After swap: athlete and new coach can exchange chat messages",
     "Athlete has just been swapped from Bob to Sarah (T-017). chat_enabled defaults to true on attach.",
     "1. Athlete opens chat panel in mobile app — should see Sarah, NOT Bob.\n"
     "2. Athlete sends a message to Sarah.\n"
     "3. Sarah opens her Messages tab in coach portal.\n"
     "4. Sarah sees the message and sends a reply.\n"
     "5. Athlete receives Sarah's reply in realtime.",
     "chat_enabled is true. Athlete's chat panel shows Sarah as their coach. "
     "Two-way messaging works in realtime. Sarah's previous-coach (Bob) does NOT appear in the athlete's "
     "chat list. Bob does NOT receive any of the new messages.",
     "- chat_enabled stays false after swap (athlete can't message new coach).\n"
     "- Athlete's chat still shows Bob (cache).\n"
     "- Bob can still read or send messages in the conversation (RLS gap on messages table).\n"
     "- Realtime channel doesn't reconnect to Sarah.",
     "P0"),

    # ===== PATIENT-INVITE DETECTION: 'always discoverable' (P1) ============
    ("T-031",
     "Coach invites X. X never taps the email. X opens the app next day -> banner appears",
     "X is signed in on mobile but didn't tap the invite email. Coach Sarah sent it yesterday.",
     "1. Coach sends invite at time T.\n"
     "2. Invitee does NOT interact with the email.\n"
     "3. Invitee opens the mobile app on the next session (after the debounce window).\n"
     "4. Background invite-detection watcher runs.",
     "Banner appears on the dashboard: 'Coach Sarah invited you — accept?' "
     "Invitee can accept without ever opening the email. This is the 'patient invite' guarantee — "
     "the invite is discoverable by any path the invitee eventually takes.",
     "- Banner never appears (watcher not registered on launch).\n"
     "- Banner appears too often (debounce window broken, fires every render).\n"
     "- Watcher fires but query misses the invite (email-match query bug).",
     "P1"),

    ("T-032",
     "Invitee signs OUT then back IN with the same email -> banner re-appears",
     "Invitee accepted nothing yet, signed out for some reason, signs back in.",
     "1. While signed in with banner visible, sign out of the app.\n"
     "2. Sign back in with the same email (the email the invite was sent to).\n"
     "3. Reach the dashboard.",
     "Banner re-appears. Sign-out / sign-in does not 'consume' or hide the pending invite. "
     "Invite stays patient until accepted, declined, expired, or revoked.",
     "- Banner permanently gone after sign-out (some local 'dismissed' flag was misused).\n"
     "- Banner reappears even after explicit Decline (decline didn't update DB).\n"
     "- Sign-in fails because of the pending invite (shouldn't block auth).",
     "P1"),

    ("T-033",
     "Coach revokes a pending invite before invitee sees the banner -> banner removed on next launch",
     "Coach sent an invite. Before invitee's next app launch, coach revokes it.",
     "1. Coach sends invite at time T.\n"
     "2. Invitee does NOT open the app yet.\n"
     "3. Coach revokes the invite at time T+1h.\n"
     "4. Invitee opens the app for the first time after T+1h.",
     "Watcher runs on launch. No banner appears (revoked invites are filtered out of the patient-detection query). "
     "The invitee has zero indication they were ever invited.",
     "- Banner appears anyway (revoked filter missing from watcher query).\n"
     "- Banner appears once then disappears mid-render (race condition).\n"
     "- Watcher errors silently and logs nothing.",
     "P1"),

    ("T-034",
     "Invite expires before invitee opens app -> no banner; coach sees 'expired' in their list",
     "Coach sent invite. 15+ days pass with no invitee interaction.",
     "1. Coach sends invite at time T.\n"
     "2. 15 days pass. Invitee never opens app.\n"
     "3. Invitee opens the app for the first time after T+15d.\n"
     "4. Coach opens her Pending Invites list.",
     "On invitee's side: watcher runs, no banner (expired invites filtered out of patient-detection). "
     "On coach's side: invite row shows status=expired (either lazily marked on read, or by a background job). "
     "Coach can choose to resend if she wants.",
     "- Banner appears anyway (expiry filter missing in watcher).\n"
     "- Coach's list still shows status=pending (expiry not enforced on read).\n"
     "- Lazy expiration causes a write on every read (load spike).",
     "P1"),
]

# --------- Build the single sheet ---------------------------------------------
SUBTITLE = (
    "34 behavioral scenarios covering the three account roles (Coach / Athlete / Admin) and how "
    "they interact through the invite flow. Each scenario exercises a real cross-role decision: "
    "who can invite whom, what happens when an existing user is invited, how coach-swaps preserve "
    "data, and how patient-invite detection keeps an invite discoverable. UI snapshot tests are "
    "intentionally excluded — they are not behavioral checks."
)

add_scenario_sheet(wb, "Coach Invites", SUBTITLE, scenarios)

# --------- Save ---------------------------------------------------------------
OUT.parent.mkdir(parents=True, exist_ok=True)
try:
    wb.save(OUT)
    print(f"Wrote {OUT}")
except PermissionError:
    alt = OUT.with_name("coach_invites_v2.xlsx")
    wb.save(alt)
    print(f"[WARN] {OUT.name} is open (Excel locked it).")
    print(f"       Wrote {alt} instead. Close Excel + re-run OR rename {alt.name} -> {OUT.name}.")

# Priority breakdown
p_counts = {"P0": 0, "P1": 0, "P2": 0}
for row in scenarios:
    p_counts[row[6]] = p_counts.get(row[6], 0) + 1

print(f"Sheets: {wb.sheetnames}")
print(f"Total scenarios: {len(scenarios)}  (P0: {p_counts['P0']}  P1: {p_counts['P1']}  P2: {p_counts['P2']})")

# --------- Verification: re-open + count rows --------------------------------
print("\nVerification (re-parsing the file):")
wb2 = load_workbook(OUT if OUT.exists() else OUT.with_name("coach_invites_v2.xlsx"))
for name in wb2.sheetnames:
    ws = wb2[name]
    # Data starts at row 5
    data_count = 0
    for row in ws.iter_rows(min_row=5, max_col=1, values_only=True):
        if row[0]:
            data_count += 1
    print(f"  {name}: {data_count} scenarios")
