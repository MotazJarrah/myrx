"""
Build docs/vendors.xlsx — the locked-down vendor inventory for MyRX.

Sheets:
  1. Active Vendors           — every paid/active SaaS we depend on
  2. Domain Transfer Countdown — GoDaddy 60-day ICANN lock formula
  3. Future / Planned         — vendors approved or shortlisted
  4. Deprecated / Removed     — vendors we used but cut
  5. Public Data Sources      — no-account data feeds we ingest

Run from repo root:
    python scripts/build_vendors_xlsx.py
Output: docs/vendors.xlsx (overwrites)
"""
from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "vendors.xlsx"

# ─────────────── Style helpers ───────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F2937")       # slate-800
HEADER_FONT = Font(bold=True, color="F8FAFC", size=11)     # slate-50
SUBHEAD_FILL = PatternFill("solid", fgColor="334155")      # slate-700
SUBHEAD_FONT = Font(bold=True, color="F8FAFC", size=10)

STATUS_FILLS = {
    "Active":      PatternFill("solid", fgColor="DCFCE7"),  # green-100
    "Trial":       PatternFill("solid", fgColor="FEF3C7"),  # amber-100
    "Pending":     PatternFill("solid", fgColor="DBEAFE"),  # blue-100
    "Deprecated":  PatternFill("solid", fgColor="FEE2E2"),  # red-100
    "Planned":     PatternFill("solid", fgColor="E0E7FF"),  # indigo-100
}

THIN = Side(border_style="thin", color="CBD5E1")           # slate-300
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header_row(ws, row_idx: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row_idx].height = 28

def style_data_rows(ws, start_row: int, end_row: int, cols: int, status_col: int | None = None) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = BORDER
            cell.font = Font(size=10)
        if status_col is not None:
            status_val = str(ws.cell(row=r, column=status_col).value or "")
            if status_val in STATUS_FILLS:
                ws.cell(row=r, column=status_col).fill = STATUS_FILLS[status_val]
                ws.cell(row=r, column=status_col).font = Font(size=10, bold=True)
        ws.row_dimensions[r].height = None  # auto

def autosize(ws, widths: dict[int, int]) -> None:
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

# ─────────────── Workbook ────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ═══════════════ Sheet 1: Active Vendors ═════════════════════════════
ws = wb.create_sheet("Active Vendors")

cols = ["Vendor", "Category", "What It Does For Us", "Plan / Tier",
        "Cost", "Login URL", "Account Email", "Important IDs / Secrets",
        "Renewal", "Status", "Notes"]
ws.append(cols)
style_header_row(ws, 1, len(cols))

active_vendors = [
    # Vendor, Category, What it does, Plan, Cost, URL, Email, IDs, Renewal, Status, Notes
    ["GoDaddy", "Domain Registrar",
     "Holds registration of myrxfit.com. Nothing else — DNS is at Cloudflare.",
     "Standard domain registration", "~$20/yr",
     "https://account.godaddy.com", "motaz.jarrah@hotmail.com",
     "Domain: myrxfit.com",
     "TBD — fill in", "Active",
     "Plan: migrate domain to Cloudflare Registrar after 60-day ICANN lock expires. See 'Domain Transfer Countdown' sheet."],

    ["Cloudflare", "DNS + CDN + Edge platform",
     "DNS for myrxfit.com, Pages (web hosting), Workers (food-search worker), D1 (food library DB), R2 (food-source ZIP mirror).",
     "Free tier (Workers Free + Pages Free + R2 free 10GB)", "$0/mo",
     "https://dash.cloudflare.com", "motaz.jarrah@hotmail.com",
     "Account ID: d42e96189bfa3cacb2aaab8231eb0097\nZone ID (myrxfit.com): bc3ca8a0f1756627df277e8dc0ca3602\nProject: myrx (Pages)\nWorker: food-search\nD1 DB: myrx-food-library\nR2 bucket: myrx-food-mirror",
     "Free (renews monthly)", "Active",
     "Free-tier abuses: D1 wrangler CLI bypasses HTTP rate limits — we use this for bulk imports. Workers Free = 100k req/day."],

    ["Supabase", "Backend (Auth + DB + Functions + Storage)",
     "Auth (email/password + magic link + phone OTP), Postgres database, Edge Functions (Deno), Storage (avatars), Realtime (chat/suggestions).",
     "Free tier (up to 500MB DB + 1GB storage + 2GB bandwidth)", "$0/mo",
     "https://supabase.com/dashboard", "motaz.jarrah@hotmail.com",
     "Project ID: xtxzfhoxyyrlxslgzvty\nProject URL: https://xtxzfhoxyyrlxslgzvty.supabase.co\nDB host: aws-0-us-east-1.pooler.supabase.com (pooler)",
     "Free (renews monthly)", "Active",
     "Edge function secrets configured: STRIPE_SECRET_KEY_TEST, TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN, TWILIO_VERIFY_SERVICE_SID, SENDGRID_API_KEY, SENDGRID_FROM."],

    ["Zoho Mail", "Inbox + outbound mail",
     "Receives all team@myrxfit.com mail. Sends outbound when composed from Outlook desktop/mobile or Zoho webmail.",
     "Mail Lite ($1/user/month, billed annually)", "$12/yr",
     "https://mailadmin.zoho.com (admin)\nhttps://mail.zoho.com (webmail)", "team@myrxfit.com",
     "Org: Northern Princess LLC\nDomain: myrxfit.com\nDKIM selector: zmail (pending verification — see task #189)",
     "05/18/27", "Active",
     "Inbound MX: mx.zoho.com / mx2.zoho.com / mx3.zoho.com\nIMAP: imappro.zoho.com:993 SSL\nSMTP: smtppro.zoho.com:465 SSL\nApp Password required (Zoho Security → Application-Specific Passwords)"],

    ["SendGrid (Twilio)", "Transactional email (outbound, app-generated only)",
     "Sends app-generated email ONLY: coach invites, Supabase Auth OTPs (magic link + recovery), future B2C transactional. Human-composed outbound (Outlook desktop + mobile + Zoho webmail) goes via Zoho native SMTP — see Zoho row.",
     "Free Trial (until 07/25/2026, then likely Essentials $20/mo or Pro)", "$0/mo (trial)",
     "https://app.sendgrid.com", "motaz.jarrah@hotmail.com",
     "Sender: team@myrxfit.com (Single Sender verified)\nDomain Auth: myrxfit.com (DKIM selectors s1, s2 + em9938 CNAME)\nLink Branding: url6706.myrxfit.com",
     "07/25/2026 (trial end)", "Trial",
     "Single active API key: 'MyRX outbound (coach invites + auth OTPs)' — Full Access — used by Supabase edge functions. Stored in Supabase function secret SENDGRID_API_KEY."],

    ["Twilio", "SMS Verify (phone OTPs)",
     "Sends phone OTP codes for signup verification + phone-change flow. Used via Twilio Verify API (not Programmable Messaging — no A2P 10DLC required).",
     "Pay-as-you-go (Twilio Trial converted to paid)", "~$0.05 per SMS sent",
     "https://console.twilio.com", "motaz.jarrah@hotmail.com",
     "Account SID: stored in Supabase Edge Function secrets (TWILIO_ACCOUNT_SID)\nVerify Service SID: stored as TWILIO_VERIFY_SERVICE_SID",
     "n/a (pay-as-you-go)", "Active",
     "Used by edge functions send-phone-otp + verify-phone-otp. Trial had verified-caller-ID restriction; upgrading to paid lifted that."],

    ["Stripe", "Payments",
     "Coach subscription billing (Starter $19/mo, Pro $39/mo, Elite $99/mo). Future B2C lifetime tiers (CoreRX $39, FullRX $59).",
     "Standard processing (2.9% + 30¢)", "Per-transaction fees only",
     "https://dashboard.stripe.com", "motaz.jarrah@hotmail.com",
     "Currently TEST mode. Live keys not yet pulled (deferred to Phase 7).\nProduct IDs in .env.example: COACH_STARTER, COACH_PRO, COACH_ELITE, CORERX, FULLRX",
     "n/a (no subscription)", "Active",
     "Edge function: stripe-webhook. Source query param distinguishes b2c vs coach_subs. Webhook signing secret in Supabase function secrets."],

    ["Anthropic Claude", "AI dev tool",
     "This conversation. Used as pair-programmer + ops automation for MyRX dev work.",
     "Claude Code (Pro / Max plan)", "TBD",
     "https://claude.ai", "motaz.jarrah@hotmail.com",
     "n/a (interactive sessions)",
     "Monthly", "Active",
     "Not used inside the MyRX app itself — purely a dev productivity tool. No customer data flows through Claude unless explicitly pasted in chat."],

    ["GitHub", "Source control + CI",
     "Hosts the MotazJarrah/myrx repo. GitHub Actions runs the food-library sync orchestrator on cron + manual triggers.",
     "Free (public/private repo on free tier)", "$0/mo",
     "https://github.com/MotazJarrah/myrx", "motaz.jarrah@hotmail.com",
     "Repo: MotazJarrah/myrx (default branch: main)\nGHA secrets: CLOUDFLARE_API_TOKEN, USDA_API_KEY",
     "Free (renews monthly)", "Active",
     "Cloudflare Pages does NOT auto-deploy from GitHub on this project. Deploys go via `wrangler pages deploy web/dist`. git push is source-of-truth only."],

    ["Samsung Developer Program", "Wearable SDK access",
     "Samsung Health Data SDK v1.1.0 — reads HR, Steps, Workouts from Galaxy Watch via Samsung Health phone app.",
     "Free developer program (approved May 20 2026)", "$0",
     "https://developer.samsung.com", "motaz.jarrah@hotmail.com",
     "App: MyRX (com.myrx.app)\nSDK: com.samsung.android.sdk:health-data:1.1.0",
     "n/a", "Active",
     "Production users skip Samsung Health 'Developer Mode for Data Read' that we toggle during dev. Galaxy S25 Ultra is the verified test device."],

    ["Expo / EAS", "Mobile dev tooling",
     "Expo SDK 54 powers the React Native mobile app. EAS handles cloud builds for iOS + Android (when needed).",
     "Free tier (limited concurrent builds)", "$0/mo",
     "https://expo.dev", "motaz.jarrah@hotmail.com",
     "App slug: myrx (matches package com.myrx.app)\nUses dev-client APK locally, NOT Expo Go (Reanimated 4 + new arch incompatible).",
     "Free (renews monthly)", "Active",
     "Local dev runs against physical Galaxy S25 via wireless adb (10.0.0.226:5555). Metro on laptop:8081. See CLAUDE.md → Mobile dev environment."],
]

for row in active_vendors:
    ws.append(row)
style_data_rows(ws, 2, 1 + len(active_vendors), len(cols), status_col=10)
autosize(ws, {1: 22, 2: 22, 3: 38, 4: 28, 5: 14, 6: 36, 7: 26, 8: 40, 9: 16, 10: 13, 11: 50})
ws.freeze_panes = "A2"

# ═══════════════ Sheet 2: Domain Transfer Countdown ═══════════════════
ws2 = wb.create_sheet("Domain Transfer Countdown")

ws2["A1"] = "GoDaddy → Cloudflare Domain Transfer Countdown"
ws2["A1"].font = Font(bold=True, size=14, color="1F2937")
ws2.merge_cells("A1:D1")

ws2["A2"] = "ICANN policy locks any domain to its current registrar for 60 days after registration OR after the most recent transfer. GoDaddy confirmed the exact unlock date in the domain's settings page — hardcoded into B6 below."
ws2["A2"].font = Font(size=10, italic=True, color="475569")
ws2["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws2.merge_cells("A2:D2")
ws2.row_dimensions[2].height = 40

# Inputs / outputs
ws2["A4"] = "Field"
ws2["B4"] = "Value"
style_header_row(ws2, 4, 2)

import datetime as _dt
ws2["A5"] = "Domain"
ws2["B5"] = "myrxfit.com"

ws2["A6"] = "Transfer-eligible date (per GoDaddy)"
ws2["B6"] = _dt.date(2026, 7, 6)
ws2["B6"].number_format = "yyyy-mm-dd"

ws2["A7"] = "Today"
ws2["B7"] = "=TODAY()"
ws2["B7"].number_format = "yyyy-mm-dd"

ws2["A8"] = "Days remaining"
ws2["B8"] = "=MAX(0, B6 - TODAY())"

ws2["A9"] = "Status"
ws2["B9"] = '=IF(B6-TODAY()<=0,"✅ Transferable now — initiate at Cloudflare Registrar","🔒 Locked at GoDaddy for "&(B6-TODAY())&" more days (unlock: "&TEXT(B6,"mmm d, yyyy")&")")'

ws2["A10"] = "GoDaddy Domain Lock"
ws2["B10"] = "ON (must turn OFF before initiating transfer)"

ws2["A11"] = "Email Privacy"
ws2["B11"] = "ON (proxy: domainsbyproxy.com — leave as-is for privacy)"

for r in range(5, 12):
    for c in (1, 2):
        cell = ws2.cell(row=r, column=c)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.font = Font(size=11) if c == 1 else Font(size=11, bold=True)
    ws2.row_dimensions[r].height = 22

ws2["A14"] = "Transfer steps (when unlocked):"
ws2["A14"].font = Font(bold=True, size=11)
steps = [
    "1. GoDaddy → My Products → Domains → myrxfit.com → Settings → DISABLE Domain Lock toggle",
    "2. Still on Settings page → Transfer section → 'Transfer to Another Registrar' link",
    "3. Follow GoDaddy's prompts → Generate Authorization (EPP) Code → emailed to motaz.jarrah@hotmail.com",
    "4. Cloudflare → Domain Registration → Transfer Domains → enter myrxfit.com + auth code from email",
    "5. Cloudflare charges 1 year of registration at-cost (~$9.77, no markup) — saves ~$10/yr vs GoDaddy",
    "6. Approve transfer from the confirmation email Cloudflare/GoDaddy send (window: ~5 days)",
    "7. Total downtime: ZERO — DNS already lives at Cloudflare; only the registrar entity changes",
    "8. After transfer: Cloudflare auto-enables registry lock + WHOIS privacy. Domain Lock toggle in old GoDaddy panel becomes irrelevant.",
]
for i, s in enumerate(steps):
    ws2.cell(row=15 + i, column=1, value=s).font = Font(size=10)
    ws2.cell(row=15 + i, column=1).alignment = Alignment(wrap_text=True)
    ws2.merge_cells(start_row=15 + i, start_column=1, end_row=15 + i, end_column=4)

autosize(ws2, {1: 38, 2: 38, 3: 18, 4: 18})
ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 42

# ═══════════════ Sheet 3: Future / Planned ════════════════════════════
ws3 = wb.create_sheet("Future and Planned")
ws3.append(cols)
style_header_row(ws3, 1, len(cols))

planned = [
    ["Apple Developer Program", "Mobile dist (iOS)",
     "Required to publish MyRX iOS app + use HealthKit entitlement.",
     "Apple Developer Program (individual or LLC)", "$99/yr",
     "https://developer.apple.com", "motaz.jarrah@hotmail.com (TBD)",
     "Pending enrollment. Need Apple ID + LLC verification documents.",
     "Annual", "Pending",
     "Apple HealthKit integration is Phase 2 of wearable work. iOS reflection checklist already in CLAUDE.md."],

    ["Google Play Developer Account", "Mobile dist (Android)",
     "Required to publish MyRX Android app on Play Store.",
     "One-time $25 registration", "$25 one-time",
     "https://play.google.com/console", "motaz.jarrah@hotmail.com (TBD)",
     "Pending enrollment.",
     "n/a (one-time)", "Pending",
     "Android App Links already configured via public/.well-known/assetlinks.json (debug keystore SHA256). Production keystore SHA256 must be added before Play release."],

    ["Strava", "Wearable / activity integration",
     "OAuth2 + REST. Read user workouts (Running, Cycling, etc.) for activity dashboard.",
     "Free developer tier", "$0",
     "https://www.strava.com/settings/api", "TBD",
     "Need to register API app: name, website, callback URL.",
     "n/a", "Planned",
     "Build order #1 per CLAUDE.md (Strava → Fitbit → Apple HealthKit → Samsung SDK → Garmin → Whoop → Polar)."],

    ["Fitbit Web API", "Wearable / activity integration",
     "OAuth2. Read user activity + sleep + HR.",
     "Personal tier (free, instant)", "$0",
     "https://dev.fitbit.com/apps", "TBD",
     "Personal-tier app registration is instant. Production rate limits need approval later.",
     "n/a", "Planned",
     "Build order #2."],

    ["Garmin Health API", "Wearable / activity integration",
     "OAuth1.0a + webhooks. Read user health data from Garmin devices.",
     "Garmin Developer Program (free, requires approval)", "$0",
     "https://developer.garmin.com/gc-developer-program/health-api/", "TBD",
     "Application must be submitted; approval typically 2-4 weeks.",
     "n/a", "Planned",
     "Build order #5. See docs/integrations/developer-program-applications.md (TBD)."],

    ["Whoop API v1", "Wearable / recovery integration",
     "OAuth2 + webhooks. Read recovery, strain, sleep.",
     "Whoop Developer Program (free, requires approval)", "$0",
     "https://developer.whoop.com", "TBD",
     "Application approval ~1-2 weeks.",
     "n/a", "Planned", "Build order #6."],

    ["Polar AccessLink", "Wearable / activity integration",
     "OAuth2. Read user training data from Polar devices.",
     "Polar Business team approval (free)", "$0",
     "https://www.polar.com/accesslink-api/", "TBD",
     "Polar Business team approval ~1-2 weeks.",
     "n/a", "Planned", "Build order #7."],

    ["Apple HealthKit", "Wearable / health platform",
     "iOS native framework — read HR, steps, workouts, body data.",
     "Free (included in Apple Developer Program)", "Included in $99/yr Apple Dev",
     "n/a (native framework)", "n/a",
     "Requires app entitlement: com.apple.developer.healthkit",
     "n/a", "Planned",
     "Build order #3. Mirrors Samsung Health Data SDK architecture but iOS-side."],
]
for row in planned:
    ws3.append(row)
style_data_rows(ws3, 2, 1 + len(planned), len(cols), status_col=10)
autosize(ws3, {1: 28, 2: 22, 3: 38, 4: 28, 5: 16, 6: 36, 7: 24, 8: 40, 9: 12, 10: 12, 11: 50})
ws3.freeze_panes = "A2"

# ═══════════════ Sheet 4: Deprecated / Removed ════════════════════════
ws4 = wb.create_sheet("Deprecated and Removed")
deprecated_cols = ["Vendor", "Was Used For", "Replaced By", "Removed On", "Cleanup Status", "Notes"]
ws4.append(deprecated_cols)
style_header_row(ws4, 1, len(deprecated_cols))

deprecated = [
    ["Resend", "Outbound email relay (Zoho → Resend → AWS SES → recipient).",
     "SendGrid (app email) + Zoho native SMTP (Outlook outbound)",
     "2026-05-27",
     "Gateway deleted from Zoho admin. DNS records removed (resend._domainkey TXT, send.myrxfit.com MX + TXT). Account itself still exists at resend.com — user to close manually.",
     "Account close on user todo (low priority — free tier auto-deactivates after ~60 days dormant)."],

    ["Netlify", "Earlier web hosting (before Cloudflare Pages migration).",
     "Cloudflare Pages (project: myrx)",
     "Pre-Apr 2026",
     "Account deleted entirely. .netlify/ folder removed from repo.",
     "DO NOT re-introduce. Deploy is strictly: wrangler pages deploy web/dist."],

    ["GoDaddy DomainConnect", "Automated DNS provisioning shortcut.",
     "Direct Cloudflare DNS management (manual or API)",
     "2026-05-27",
     "DNS CNAME _domainconnect.myrxfit.com → _domainconnect.gd.domaincontrol.com removed.",
     "Was a dangling DNS leftover from GoDaddy days. We use Cloudflare DNS directly now."],

    ["GoDaddy Commerce / pay.myrxfit.com", "Earlier payment landing page experiment.",
     "Stripe (coach subs + B2C tiers)",
     "2026-05-27",
     "CNAME pay.myrxfit.com → paylinks.commerce.godaddy.com removed (was returning 404).",
     "Page was dead. Cleaned during email-vendor consolidation DNS audit."],
]
for row in deprecated:
    ws4.append(row)
style_data_rows(ws4, 2, 1 + len(deprecated), len(deprecated_cols))
autosize(ws4, {1: 26, 2: 38, 3: 36, 4: 14, 5: 50, 6: 50})
ws4.freeze_panes = "A2"

# ═══════════════ Sheet 5: Public Data Sources ═════════════════════════
ws5 = wb.create_sheet("Public Data Sources")
ds_cols = ["Source", "What We Use It For", "Access Method", "Cost", "API Key Storage", "Refresh Cadence", "Notes"]
ws5.append(ds_cols)
style_header_row(ws5, 1, len(ds_cols))

data_sources = [
    ["USDA FoodData Central", "Branded + foundation + sr_legacy + survey food data (~470k rows).",
     "Manual ZIP download from https://fdc.nal.usda.gov/download-datasets uploaded to R2 mirror, then ingested by GHA orchestrator.",
     "Free (public US gov data, no API key required for download path)",
     "GHA secret: USDA_API_KEY (only used by deprecated direct-scrape path; bulk ingest uses R2 mirror).",
     "Twice a year (April + October/November)",
     "USDA's CDN (fdc-datasets.ars.usda.gov) returns ENOTFOUND from cloud-egress IPs. R2 mirror works around this — admin downloads ZIPs locally, drops into food-library admin upload UI, GHA pulls from R2."],

    ["OpenNutrition", "Branded foods + recipes (~60k rows). Supplements USDA's coverage of international + smaller-brand items.",
     "Manual ZIP download from https://www.opennutrition.app/download, same R2-mirror flow.",
     "Free", "n/a (open dataset)",
     "Less frequent than USDA — manual check periodically",
     "Watermark stored in sync_state.on_last_version."],

    ["OpenFoodFacts", "UPC barcode lookup (mobile food-log scanner).",
     "REST API via Cloudflare Worker proxy at /api/off-search",
     "Free", "n/a (no auth required)",
     "Real-time (per scan)",
     "Worker proxy enforces 8s timeout. Scan UI falls through to manual UPC entry if OFF times out."],
]
for row in data_sources:
    ws5.append(row)
style_data_rows(ws5, 2, 1 + len(data_sources), len(ds_cols))
autosize(ws5, {1: 24, 2: 40, 3: 44, 4: 14, 5: 40, 6: 28, 7: 50})
ws5.freeze_panes = "A2"

# ═══════════════ Save ═════════════════════════════════════════════════
OUT.parent.mkdir(parents=True, exist_ok=True)
try:
    wb.save(OUT)
    print(f"Wrote {OUT}")
except PermissionError:
    # File is open in Excel — write to a sibling and instruct swap
    alt = OUT.with_name("vendors_v2.xlsx")
    wb.save(alt)
    print(f"⚠ {OUT.name} is open (Excel/Numbers/LibreOffice locked it).")
    print(f"   Wrote {alt} instead. Close Excel, then either re-run this script")
    print(f"   or rename {alt.name} → {OUT.name} manually.")
print(f"Sheets: {wb.sheetnames}")
