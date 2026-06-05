"""
Build docs/TASK_PIPELINE.xlsx — the MyRX cross-session task ledger.

WHY THIS EXISTS
We discuss many things across many sessions and often fork mid-task. This
sheet is the single place that remembers: every task (done + pending), a
stable numeric ID to refer back to, where we left off, what's done, what's
still open, the files/commits involved, and when it was last touched.

HOW TO MAINTAIN IT (read this — the .xlsx is GENERATED, not hand-edited)
  1. Edit the TASKS list below (add a row, flip a status, update the
     "left off" / "next" text).
  2. Re-run:  python scripts/build_task_pipeline_xlsx.py
  3. Commit both this script AND docs/TASK_PIPELINE.xlsx.
New tasks get the next free T### id. NEVER reuse or renumber an id — the
whole point is a stable reference ("pick up T021").

Seeded 2026-06-03 from the current CLAUDE.md + the recent working sessions.
It is best-effort for older work (older sessions weren't fully transcribed);
treat it as the living record from here forward.

Run from repo root:
    python scripts/build_task_pipeline_xlsx.py
"""
from __future__ import annotations
import datetime as _dt
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "TASK_PIPELINE.xlsx"

# ─────────── styles ────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="111721")   # MyRX dark
HEADER_FONT = Font(bold=True, color="CAF240", size=11)  # MyRX lime
TITLE_FONT  = Font(bold=True, color="111721", size=18)
SUB_FONT    = Font(color="475569", size=10, italic=True)

STATUS_FILLS = {
    "Done":        PatternFill("solid", fgColor="DCFCE7"),  # green-100
    "In progress": PatternFill("solid", fgColor="DBEAFE"),  # blue-100
    "Pending":     PatternFill("solid", fgColor="FEF3C7"),  # amber-100
    "Deferred":    PatternFill("solid", fgColor="E0E7FF"),  # indigo-100
    "Parked":      PatternFill("solid", fgColor="F1F5F9"),  # slate-100
    "Reverted":    PatternFill("solid", fgColor="FEE2E2"),  # red-100
    "Closed":      PatternFill("solid", fgColor="E5E7EB"),  # gray-200 — dropped / won't do
}
ZEBRA = PatternFill("solid", fgColor="F8FAFC")

THIN = Side(border_style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ─────────── columns ────────────────────────────────────────────────────────────
COLS = [
    ("ID",                 8),
    ("Task",               34),
    ("Area",               16),
    ("Surface",            14),
    ("Status",             13),
    ("Where we left off / what we did", 70),
    ("What's still open / next",        46),
    ("Key files & commits",             40),
    ("Last touched",       14),
]

# ─────────── the ledger ─────────────────────────────────────────────────────────
# (id, task, area, surface, status, left_off, next_open, files_commits, last_touched)
TASKS = [
    # ── Active / recent working sessions (most relevant for "where are we") ──
    ("T001", "Hydration mascot (pixel slime)", "Hydration", "Mobile", "Done",
     "Replaced the progress ring with HydrationPet: a pixeowl pixel slime over an animated day/night PixelScene. Pace-aware mood (compares intake vs where you should be by ~9pm), first-person rotating captions (no emojis), Cup/Bottle quick-adds, drink-hop reaction. Final idle = a hand-authored 16-frame loop using a custom off-sheet 'settle' frame (idle-extra.png).",
     "None — shipped. Hydration PAGE redesign is a separate open item (T016).",
     "mobile/src/components/HydrationPet.tsx, PixelScene.tsx; assets/pet/; commit bb7ade2",
     "2026-06-03"),

    ("T002", "Hydration water target (science-correct)", "Hydration", "Mobile", "Done",
     "Target = 35 mL/kg/day from the client's LATEST LOGGED bodyweight (not profile weight), fallback by gender. Added attribution line (National Academies / Mayo / EFSA). Fixed the earlier kg/lb inconsistency.",
     "None.",
     "mobile/app/(app)/hydration.tsx; commit bb7ade2",
     "2026-06-03"),

    ("T003", "Settings cleanup (fluid unit, invite, distance/swim)", "Settings", "Mobile", "Done",
     "Added a Fluid (oz/mL) unit toggle. Removed the 'paste an invite code' block. Merged Distance + Swim-distance into ONE toggle ('mi · yd' / 'km · m', middle-dot separator). settings.tsx is the real settings file (not profile.tsx).",
     "None.",
     "mobile/app/(app)/settings.tsx; commit bb7ade2",
     "2026-06-03"),

    ("T004", "Heart page corrections", "Heart", "Mobile", "Done",
     "Fixed resting-number inconsistency, the 'no data today' vs 'Latest HR' contradiction, 6-of-7-day padding, and the chart 'Peak 91-108' (now single value). Bucketing switched from UTC to local date; resting fallback = AVG of daily lows. Removed the 3 resting-HR tips entirely.",
     "None.",
     "mobile heart.tsx, HrRangeChart.tsx, RestingHrIndicator.tsx; commit bb7ade2",
     "2026-06-03"),

    ("T005", "Calorie plan live sync (coach edit -> athlete)", "Calories/Plan", "Backend+Mobile", "Done",
     "Coach 'Update plan' didn't refresh the athlete's Calories page in real time. Root cause: calorie_plans + bodyweight weren't in the realtime publication AND mobile had no subscription. Fixed: published both tables (REPLICA IDENTITY FULL) + added a mobile realtime subscription that re-fetches on change. Audited all coach->client flows; this was the only one missing realtime.",
     "None.",
     "mobile calories.tsx; supabase migration 20260603a; commit 88be006",
     "2026-06-03"),

    ("T006", "Goal-reached resets on new phase", "Calories/Plan", "Backend", "Done",
     "Progress bar stuck at 100% after a coach changed the goal. Fixed at the DB layer: trigger reset_goal_reached_on_phase_change clears goal_reached whenever the goal OR starting weight changes (covers every writer). Sticky 100% only persists through weight fluctuations; an explicit save/goal-change re-baselines. One-time data fix cleared the stale prod row.",
     "None.",
     "supabase migration 20260603b; commit 88be006/86b88c3",
     "2026-06-03"),

    ("T007", "Remove redundant 'Reset goal' button", "Calories/Plan", "Web", "Done",
     "After T006 made 'Update plan' clear the reached flag, the separate 'Reset goal' button in MacroPlanEditor was redundant. Removed the button, handler, state, icon import; fixed the help copy that referenced it.",
     "None.",
     "web MacroPlanEditor.jsx",
     "2026-06-03"),

    ("T008", "'Goal reached' banner persistence fix", "Calories/Plan", "Web", "Done",
     "Banner lingered after Update plan because the editor used a stale local plan copy. handleSave now reads the DB-authoritative row back (.select().single()) so the banner clears; copy no longer says 'reset'.",
     "None.",
     "web MacroPlanEditor.jsx",
     "2026-06-03"),

    ("T009", "Calorie timeline unification", "Calories/Plan", "Cross", "Done",
     "One shared timeline calc across signup wizard, Calories page, and coach view (they used to disagree — signup assumed perfect adherence, the page assumed realistic ~20/30 days). Recomp now shows an achievable timeline instead of a static slow band; wrong-direction plans show a mismatch warning; single 'monthsBest' figure everywhere. Decided items: keep 20/30 realistic; recomp shows a real number; one number across all three.",
     "None on the math. (See T014 — adding new dashboard pills is the only calorie-adjacent open item.)",
     "mobile calorieFormulas.ts, planPresets.ts, PlanWizardSheet.tsx, calories.tsx; web calorieFormulas.js, MacroPlanEditor.jsx; commit 86b88c3",
     "2026-06-03"),

    ("T010", "Hide timeline once goal reached", "Calories/Plan", "Mobile", "Done",
     "After goal reached, the Calories page wrongly showed a timeline (below goal) or 'different directions' mismatch (above goal). Now suppressed via the sticky plan.goal_reached flag (works for coached + self-coached).",
     "None.",
     "mobile calories.tsx; commit 632d9c8",
     "2026-06-03"),

    ("T011", "Dashboard BW pill fix", "Dashboard", "Cross", "Done",
     "The weight pill needed two weigh-ins across week windows, so it showed nothing after a single fresh log. Now: change since the last weigh-in, with a current-weight fallback so it shows after ANY log. Mirrored to the admin + coach client dashboards.",
     "None.",
     "mobile dashboard.tsx; web AdminUserDetail.jsx, CoachClientDetail.jsx; commit 86b88c3",
     "2026-06-03"),

    ("T012", "Remove legacy Mobility feature", "Mobility", "Cross", "Done",
     "Mobility/ROM removed everywhere. Mobile: deleted the page + ROMVisualizer, stripped ROM from the dashboard/effortTags/RadialNav/signup. Web admin: deleted AdminMobilityDetail/AdminClientMobility/ROMVisualizer, removed the route + all rom_records usage from AdminOverview/Feed/Dashboard/UserActivity/UserDetail/Navbar/movements.js + marketing copy. RETAINED: rom_records DB table (historical data, no UI) + the unrelated cardio 'Mobility' crawl subtype. Note: first removal agent ran on a stale worktree branch — caught + redone on the real checkout.",
     "Optional follow-ups: T049 (drop rom_records table) and T050 (scrub mobility from legal docs) — both deferred to your call.",
     "deletions + edits across mobile + web; commit 86b88c3",
     "2026-06-03"),

    ("T013", "Radial-nav background recolor", "Navigation", "Mobile", "Closed",
     "Experiment: recolor the radial-nav dome from near-black to dark MyRX-green. Tried hsl(73,70,12) -> hsl(73,45,8) -> hsl(100,45,7); user disliked all. Reverted COLOR_DOME back to colors.background — RadialNav.tsx matches the committed version exactly.",
     "CLOSED 2026-06-03 per user. Green direction rejected + already reverted; not pursuing a radial-nav redesign for now. Reopen if the user picks a new direction.",
     "mobile RadialNav.tsx (reverted)",
     "2026-06-03"),

    ("T014", "Dashboard pills covering all pages", "Dashboard", "Cross", "Done",
     "Gap pages were Sleep + Hydration (Mobility pill dropped — feature removed in T012). Added two chips on ALL three surfaces: 'Nh avg sleep . 7 nights' (indigo, from sleep_sessions.duration_s) + 'N days hit water goal . 7d' (cyan: days in the last 7 where sum(amount_ml x BHI multiplier) >= 35 mL/kg of latest bodyweight). Each chip hides when there's no data, like the existing chips.",
     "DONE 2026-06-03 (awaiting user device test). Surfaces: mobile dashboard.tsx; web AdminUserDetail.jsx + CoachClientDetail.jsx (their SnapshotBadge gained indigo/cyan colors). mobile tsc clean + web build clean. Sleep/Hydration coaching-surface promotions (T043/T044) are separate roadmap items.",
     "(planned) mobile dashboard.tsx; web AdminUserDetail.jsx, CoachClientDetail.jsx",
     "2026-06-03"),

    ("T015", "Mirror finalized admin client-view to coach", "Coach/Admin", "Web", "Parked",
     "Phase 2 of the admin/coach parity work. User re-scoped 2026-06-04: finalize the ADMIN client-detail view first (T078), THEN mirror the finalized result to the coach client view. Audited diffs so far: Admin has 2 extra tabs (Billing, Activity Feed, admin-only by nature) + a richer Calories tab (Food Log / Manual Logs / Macro Plan) vs the coach's inline Macro Plan editor behind 'Manage macros'. Exact mirror scope is TBD until T078 lands.",
     "BLOCKED on T078 (finalize admin client view). Resume after admin is final: decide per-area what mirrors to coach and what stays admin-only.",
     "web admin/AdminUserDetail.jsx + tabs/*; coach/CoachClientDetail.jsx",
     "2026-06-04"),

    ("T016", "Hydration page redesign / update", "Hydration", "Mobile", "Done",
     "Confirmed done by user 2026-06-04: the Hydration page redesign landed via the T052-T062 batch (fluid counting, fast picker, non-intimidating progress, attribution, pond removal, bug fixes).",
     "Shipped via T052-T062.",
     "mobile/app/(app)/hydration.tsx",
     "2026-06-04"),

    ("T017", "Cross-session task pipeline ledger", "Infra/Docs", "Docs", "Done",
     "Built this file — docs/TASK_PIPELINE.xlsx — generated by scripts/build_task_pipeline_xlsx.py. Captures all known tasks (done + pending) with stable IDs + 'where we left off' context. CLAUDE.md points here so every session reads + maintains it.",
     "Keep it updated: edit the TASKS list in the builder script + re-run + commit, every time a task starts/advances/finishes.",
     "scripts/build_task_pipeline_xlsx.py -> docs/TASK_PIPELINE.xlsx; CLAUDE.md pointer",
     "2026-06-03"),

    # ── Completed feature surfaces (prior sessions; from CLAUDE.md locked specs) ──
    ("T018", "Strength: Weighted Standard next-target card", "Strength", "Mobile", "Done",
     "Locked spec: big-weight algorithm (eff curve from best 1RM), adp-zone pill (strength/hypertrophy/endurance) with chevron swipe, rep-max tile row, equipment-specific footers, single coaching cue. Per-zone defaults locked.",
     "Frozen/done. Don't touch finalized surfaces unless a new locked rule applies.",
     "mobile [exercise].tsx (WeightedStandardDetail); CLAUDE.md spec",
     "prior"),

    ("T019", "Strength: Bodyweight consolidated detail", "Strength", "Mobile", "Done",
     "Locked spec: 4 assist tiers (Full RX / Band / Knee / Band+Knee) as a single consolidated page, band-level sub-progression, max-attempt tiles, pill-swipe carousel (Pattern 4), 10-rep graduation rule.",
     "Frozen/done.",
     "mobile [exercise].tsx (BodyweightConsolidatedBlock); CLAUDE.md spec",
     "prior"),

    ("T020", "Strength: Isometric milestones", "Strength", "Mobile", "Done",
     "Locked spec: 12 universal milestones (10-120s), 3 phases (Stability/Durability/Mastery), 3-6-3 tile grid, fmtDuration labels, next-target hero.",
     "Frozen/done.",
     "mobile [exercise].tsx (Isometric); CLAUDE.md spec",
     "prior"),

    ("T021", "Strength: Assisted Machine detail", "Strength", "Mobile", "Done",
     "Locked spec: inverted effective-load math, bodyweight gate (30-day recency), %BW tiles, pin-snapped targets, 'Attempt unassisted' graduation cue. Shared formula dropped Brzycki >10 reps.",
     "Frozen/done.",
     "mobile [exercise].tsx (AssistedMachine); CLAUDE.md spec",
     "prior"),

    ("T022", "Strength: Carry detail (+ Sled Work consolidated)", "Strength", "Mobile", "Done",
     "Locked spec: dual-axis (weight + distance), strongman tiers, per-movement ladders, 3 adp zones (Max Load/Distance Build/Conditioning), unit locks (kg / mi), Sled Work Push|Pull consolidated page.",
     "Frozen/done.",
     "mobile [exercise].tsx (CarryDetail, SledWorkConsolidatedDetail); CLAUDE.md spec",
     "prior"),

    ("T023", "Cardio coaching surface (pace zones)", "Cardio", "Mobile", "Done",
     "Locked spec: Group A endurance activities get Endurance/Threshold/VO2 zones, a live plan-queue generator (polarized 80/20), 45-min session ceiling, amber theme, per-activity prescribed sessions.",
     "Frozen/done.",
     "mobile [activity].tsx (PaceDetail); CLAUDE.md spec",
     "prior"),

    ("T024", "Cardio: Air Bike surface", "Cardio", "Mobile", "Done",
     "Locked spec: cal/min-anchored Sprint/Threshold/Aerobic zones, watts-floor advisory (cal/min x 17.4), gender baseline cold-start, calorie-input log form.",
     "Frozen/done.",
     "mobile [activity].tsx (AirBikeDetail); CLAUDE.md spec",
     "prior"),

    ("T025", "Cardio: Rucking surface", "Cardio", "Mobile", "Done",
     "Locked spec: carry-style (load + distance, not pace) with GoRuck tier ladder, lb + mi unit locks, 3 zones, pack-weight log wheel.",
     "Frozen/done.",
     "mobile [activity].tsx (RuckingDetail); CLAUDE.md spec",
     "prior"),

    ("T026", "Cardio: StairMill surface", "Cardio", "Mobile", "Done",
     "Locked spec: floors-per-minute rate-anchored 3 zones (Allison/Honda/Boreham protocols), floors+time log form.",
     "Frozen/done.",
     "mobile [activity].tsx (StairMillDetail); CLAUDE.md spec",
     "prior"),

    ("T027", "Cardio: Swimming consolidated (4 strokes)", "Cardio", "Mobile", "Done",
     "Locked spec: Free/Back/Breast/Fly variants collapsed into one detail page (stroke-pill carousel), CSS-anchored zones via Riegel proxy, per-100m pace, leaving-interval hero, per-stroke independent fitness.",
     "Frozen/done.",
     "mobile [activity].tsx (SwimmingConsolidatedDetail); CLAUDE.md spec",
     "prior"),

    ("T028", "Cardio: Concept2 ergs (Row/Bike/Ski)", "Cardio", "Mobile", "Done",
     "Locked spec: metric distance always, per-500m split, watts via the Concept2 cubic formula, 4-row hero, shared PaceDetail branching.",
     "Frozen/done.",
     "mobile [activity].tsx (PaceDetail erg branches); CLAUDE.md spec",
     "prior"),

    ("T029", "Cardio catalog cleanup", "Cardio", "Backend+Mobile", "Done",
     "Removed recreational/terrain-confounded activities (walking, hiking, skiing, trail/hill running, MTB, niche machines) across 3 passes; consolidated duplicate variants; moved sled/sandbag carries to strength. Final list = 15 DB rows / 12 visible activities.",
     "Frozen/done.",
     "supabase movements table; mobile [activity].tsx; CLAUDE.md spec",
     "prior"),

    ("T030", "Animation pattern library (1-7)", "Design system", "Mobile", "Done",
     "Documented + locked 7 canonical motion patterns (entrance cascade, ticker number, pulsing chevron, consolidated-page swipe, inline expansion, PhantomWheel inertia, save-button feedback) with exact constants + source locations.",
     "Reference only — reuse patterns; add Pattern 8+ to CLAUDE.md if a genuinely new motion is needed.",
     "CLAUDE.md 'Animation patterns' section",
     "prior"),

    ("T031", "Mobile app port (all athlete surfaces)", "Platform", "Mobile", "Done",
     "Ported every athlete surface to Expo/React Native: dashboard, strength, cardio, bodyweight, calories, heart, hydration, sleep, history(removed), profile/settings, chat/suggestions, full signup journey.",
     "Mobile is the active surface for all new athlete work.",
     "mobile/ tree; CLAUDE.md 'Mobile Mirror'",
     "prior"),

    ("T032", "Web/Mobile role rule (athletes mobile-only)", "Platform", "Cross", "Done",
     "Locked May 27 2026: athletes have ZERO web surfaces; web = coach portal + admin portal only. Athlete web pages archived/removed; post-sign-in routing branches by role; /app download placeholder for athletes on web.",
     "Honor on every routing/signup/feature change. (CLAUDE.md frozen/mobile-mirror language reconciled to this in a later session.)",
     "web App.jsx; CLAUDE.md 'Web / Mobile role rule'",
     "prior"),

    ("T033", "Coach Platform v1 (signup + portal)", "Coach Platform", "Web", "Done",
     "Public coach signup at /coach/* + Stripe Checkout + CoachShell portal (dashboard, clients, invite, messages, briefing, adjustments, profile, client detail). Tier-aware access; coach-attached athletes get full access.",
     "Phase 1/2 shipped. Billing/IAP polish + further coach tooling ongoing.",
     "web pages/coach/*; CLAUDE.md Coach Platform sections",
     "prior"),

    ("T034", "Auth infrastructure", "Auth", "Cross", "Done",
     "Email OTP + magic-link dual model, biometric sign-in (SecureStore), Twilio Verify phone OTP (edge functions), Android App Links, profile-completeness gating, deleted-user detection.",
     "Done. Web OTP zero-tap parked pending a Twilio template approval.",
     "AuthContext, edge functions, supabase email templates",
     "prior"),

    ("T035", "Food library (D1 + search + import)", "Food Library", "Backend", "Done",
     "~470K-row Cloudflare D1 food DB (USDA + OpenNutrition + MYRX), search worker w/ UPC, one-shot bulk_import with in-memory dedup, R2-mirror sync orchestrator (GHA), admin Sync UI.",
     "Open gap: wiring the filter pipeline into the incremental SYNC scripts (bulk import already applies it). See T048.",
     "workers/food-search, scripts/bulk_import, scripts/sync; CLAUDE.md food sections",
     "prior"),

    ("T036", "Health Connect integration (Phase 1)", "Integrations", "Mobile", "Done",
     "Android read-only Health Connect: config plugins (permissions + ViewPermissionUsageActivity + queries visibility), MainActivity delegate registration, manual 'Sync now' (logs last-7-day workouts + HR). Several hard-won Android gotchas documented.",
     "Phase 1 logs to console; mapping HC records -> effort logs is next. Superseded as PRIMARY path by the direct-integration strategy (T040) since Samsung Health doesn't bridge HR/workouts to HC.",
     "mobile plugins/, lib/healthConnect.ts, settings ConnectTab; CLAUDE.md HC section",
     "prior"),

    ("T037", "Calories: food logging", "Calories", "Cross", "Done",
     "food_logs table + FoodLogDrawer (search -> portion picker -> log) + barcode scan + CalorieStrip + per-meal breakdown + admin Food Log tab.",
     "Done. Per-session calorie auto-estimation deferred to the Calories overhaul.",
     "FoodLogDrawer, calories.tsx, food_logs",
     "prior"),

    ("T038", "Chat & suggestions (realtime)", "Messaging", "Cross", "Done",
     "messages table + RLS, realtime chat + suggestions on mobile (ChatSheet/SuggestionSheet) and web (admin/coach), chat_enabled gate, coach-info RPC, v3 unconditional coach-client chat.",
     "Done.",
     "ChatSheet, SuggestionSheet, AdminMessages, CoachMessages",
     "prior"),

    ("T039", "Admin portal", "Coach/Admin", "Web", "Done",
     "Full admin portal: overview, clients roster, weight-goal progress, nutrition grid, activity feed, messages, unified Libraries (movements + foods), exports/archive, per-client detail tabs.",
     "Done. Ongoing per-feature tweaks (e.g. T011 BW mirror, T012 mobility removal).",
     "web pages/admin/*",
     "prior"),

    # ── Pending / roadmap ──
    ("T040", "Direct per-platform integrations", "Integrations", "Cross", "Closed",
     "Strategy locked May 18 2026 after Samsung Health didn't bridge HR/workouts to Health Connect: build dedicated integrations per platform. Build order: Strava -> Fitbit -> Apple HealthKit -> Samsung SDK -> Garmin -> Whoop -> Polar. Cross-cutting infra: OAuth callback worker, webhook receiver worker, user_integrations table, token-refresh cron, per-platform mappers. HC stays as a fallback.",
     "CLOSED 2026-06-03 per user — not pursuing direct per-platform wearable integrations for now. Health Connect Phase 1 (T036) remains in the app as the Android path. Reopen if revisited; the build order + infra notes above are preserved.",
     "(planned) workers/oauth, workers/webhooks, mobile/src/lib/integrations/; CLAUDE.md integrations spec",
     "prior"),

    ("T041", "Bodyweight -> coaching surface", "Bodyweight", "Mobile", "Closed",
     "Closed by user 2026-06-04 (same call as T042): the Bodyweight page stays a tracking surface — weight + chart + goal, no separate coaching layer.",
     "Won't do — bodyweight remains a tracking surface by design.",
     "mobile bodyweight.tsx",
     "2026-06-04"),

    ("T042", "Calories -> coaching surface", "Calories", "Mobile", "Closed",
     "Closed by user 2026-06-04: the Calories page intentionally stays a food-logging surface — no coaching layer. The plan wizard + timeline (T009) cover the planning side; there is no separate 'next step' coaching ambition for calories.",
     "Won't do — calories remains a logging surface by design.",
     "mobile calories.tsx",
     "2026-06-04"),

    ("T043", "Sleep -> coaching surface", "Sleep", "Mobile", "Done",
     "Confirmed done by user 2026-06-04: the Sleep page is already built and already does coaching (next-step guidance), not just tracking.",
     "Shipped. Sleep is a coaching surface.",
     "mobile sleep.tsx",
     "2026-06-04"),

    ("T044", "Hydration -> coaching surface", "Hydration", "Mobile", "Done",
     "Confirmed done by user 2026-06-04: the Hydration coaching surface is already finalized (target + guidance), on top of the recent redesign (T052-T062).",
     "Shipped. Hydration is a coaching surface.",
     "mobile hydration.tsx",
     "2026-06-04"),

    ("T045", "Sleep / Water / Habits tracking surfaces", "Recovery/Habits", "Mobile", "Done",
     "Confirmed done by user 2026-06-04: Sleep and Water (Hydration) tracking surfaces are built. (Sleep/Hydration also do coaching — see T043/T044.)",
     "Shipped.",
     "mobile sleep.tsx, hydration.tsx",
     "2026-06-04"),

    ("T046", "iOS launch checklist", "Platform/iOS", "Mobile", "Deferred",
     "Pre-iOS-launch work: HealthKit entitlements + Info.plist usage strings, IAP/receipt validation, Sign in with Apple, SMS autofill, edge-swipe parity, EAS iOS build profile, etc.",
     "Deferred by user 2026-06-04: only triggered once the app is Android-complete. Revisit after Android ships.",
     "CLAUDE.md iOS checklist; docs/launch_checklist.xlsx",
     "2026-06-04"),

    ("T047", "Finalize Connect page and logic", "Integrations", "Mobile", "Pending",
     "Reframed by user 2026-06-04 (was 'Health Connect Phase 2'): finalize the Connect page (Settings -> Connect) and its sync logic. Covers wiring the integration rows to real connect/sync/disconnect behaviour + the last-sync display. (Background sync / HC write-back from the old Phase-2 scope can fold in here as sub-items.)",
     "Open — finalize the Connect page UI + connect/sync logic.",
     "mobile profile.tsx (ConnectTab), src/lib/healthConnect.ts, src/lib/lastSyncStorage.ts",
     "2026-06-04"),

    ("T048", "Wire filters into food-library SYNC scripts", "Food Library", "Backend", "Done",
     "Audited 2026-06-04: ALREADY DONE. The production sync orchestrator (scripts/sync/run.mjs) reuses the bulk-import loaders (loadUsda/loadOn -> enrichFood + shouldKeepFood = Rules 1-14) and applyDedup (dedup_in_memory.mjs = Rules 15-19), so all 19 rules run on every sync and produce a byte-identical filtered/deduped result to a full rebuild. The 'Sync now' button dispatches sync-food-library.yml -> node scripts/sync/run.mjs (plus monthly cron 0 3 1 * *).",
     "No code change needed. Legacy scripts/d1_migrate/sync_usda.mjs + sync_on.mjs are dead/superseded (no workflow/package.json refs) and have themselves been migrated to enrichFood + getFilterReason anyway. Also fixed the two stale CLAUDE.md bullets that wrongly claimed sync skips the filter pipeline.",
     "scripts/sync/run.mjs, scripts/bulk_import/lib/{usda_loader,on_loader,dedup_in_memory}.mjs, scripts/d1_migrate/lib/filters.mjs",
     "2026-06-04"),

    ("T049", "Drop rom_records table (post-Mobility)", "Mobility", "Backend", "Done",
     "After T012 removed the Mobility UI, the rom_records table was orphaned (no UI reads it). Dropping it is destructive (deletes historical ROM data) so it waited for user OK.",
     "DONE 2026-06-03 per user 'run 8'. Migration 20260603d_drop_rom_records: patched anonymize_account_now() to drop its 'DELETE FROM rom_records' line, then DROP TABLE rom_records CASCADE (RLS policies + activity-log trigger went with it). Verified gone; only trg_log_data_activity keeps a harmless dead 'rom_records' CASE-string branch. delete-user edge-fn source had its rom_records entry removed (per-table try/catch made it non-fatal anyway) — needs a routine `supabase functions deploy delete-user` (non-urgent, admin-only path).",
     "supabase rom_records table; delete-user/admin-user-management functions",
     "2026-06-03"),

    ("T050", "Scrub Mobility from legal docs", "Mobility/Legal", "Web", "Done",
     "Privacy Policy / DPA / Coach Agreement listed 'mobility / range-of-motion' as a collected data category. Left until T012's owner OK'd it.",
     "DONE 2026-06-03 per user 'run 9'. Removed the mobility/ROM wording: PrivacyPolicy (workout-entries list + 'mobility / range-of-motion assessments' + 'ROM records' in the deletion paragraph), DataProcessingAgreement ('mobility ROM'), CoachAgreement ('Mobility' in the training-logs line). Ships with the next web deploy.",
     "web/src/pages/legal/{PrivacyPolicy,DataProcessingAgreement,CoachAgreement}.jsx",
     "2026-06-03"),

    ("T051", "Capture-first workflow rule", "Infra/Docs", "Process", "Done",
     "Locked 2026-06-03 at the user's request: every time the user raises ANY point that needs fixing or discussion — big or small (bug, design tweak, decision, idea, 'look into X') — the FIRST action BEFORE replying is to log it here as a new Pending task with the next T### id, then answer (and tell the user the id). It stays Pending until actually done/decided, then flips to Done/Deferred/Parked/Reverted. Encoded as the CAPTURE-FIRST mandate in CLAUDE.md so it persists across sessions.",
     "Ongoing discipline — applies to every future turn. Nothing to build.",
     "CLAUDE.md (CAPTURE-FIRST bullet); scripts/build_task_pipeline_xlsx.py",
     "2026-06-03"),

    ("T052", "Hydration: count non-water fluids toward goal", "Hydration", "Mobile", "Done",
     "User wants the hydration goal to feel realistic — almost nobody drinks the full plain-water target every day. Asked whether science supports counting non-water fluids. Answer = yes: National Academies (IOM 2004) frames the daily target as TOTAL water from all sources (~80% from beverages of every kind, ~20% from food); Maughan et al. 2016 Beverage Hydration Index shows milk / juice / oral-rehydration hydrate AS WELL AS or better than water, and coffee / tea / soda ~= water (the 'caffeine dehydrates you' idea is a myth at normal intake, Killer 2014); alcohol is the real diuretic exception. Presented a decision on how granular to count fluids. Sub-task of T016 (hydration page redesign).",
     "BUILT 2026-06-03 (option 2, BHI multipliers). DB: migration add_drink_type_to_water_logs (water_logs.drink_type, default 'water', CHECK to the 6 eligible types). App (hydration.tsx): drink registry water/sparkling/coffee/tea/soda=1.0 + milk=1.5; effective hydration = sum(amount_ml x multiplier) drives the pet/water-level, the cups readout, AND the 7-day chart; 35 mL/kg goal unchanged. Sub-parts T053-T056 all Done. tsc clean. Possible polish later: wavy water top, save-confirm flash.",
     "mobile/app/(app)/hydration.tsx",
     "2026-06-03"),

    ("T053", "Hydration: fast drink picker (no dropdowns)", "Hydration", "Mobile", "Done",
     "User: need a clever picker for drink TYPE + SIZE that's fast to tap — explicitly NO dropdowns (a dropdown puts selection effort on the user and won't go well). Proposed two-tap design on one sheet: row of drink-type tiles (icon + tiny label: Water / Sparkling / Coffee / Tea / Diet soda / Milk) + a row of VESSEL-icon size chips (glass / cup / mug / bottle / large bottle mapped to common ml) instead of numbers; optional 'custom' chip opens the existing PhantomWheel for an exact amount.",
     "BUILT 2026-06-03: 6 drink-type tiles (lucide GlassWater/Droplets/Coffee/Leaf/CupSoda/Milk) -> tap reveals a size row (250/350/500 mL or 8/12/16 oz) + a Custom chip that opens the PhantomWheel in a modal. Two taps, no dropdowns. Inline FadeInUp expansion (Pattern 5).",
     "mobile/app/(app)/hydration.tsx; PhantomWheel for custom path",
     "2026-06-03"),

    ("T054", "Hydration: eligibility messaging", "Hydration", "Mobile", "Done",
     "User: make it explicit that only zero-calorie, low-calorie, milk, and non-alcoholic drinks count. Plan: the picker only ever offers eligible drink types (no alcohol, no full-sugar drinks) — that curation is itself the clearest signal — plus a short static info-pill explainer (intent-only per the info-pill rule, no formulas): roughly 'Only no/low-calorie, non-alcoholic drinks count toward hydration — plus milk. Sugary and alcoholic drinks don't.'",
     "BUILT 2026-06-03: the picker only offers eligible types (no alcohol, no full-sugar drinks) + a static note under it: 'Only no- and low-calorie, non-alcoholic drinks count toward hydration — milk included.' Juice / sports drinks excluded as calorie drinks; milk is the sole calorie exception.",
     "mobile/app/(app)/hydration.tsx",
     "2026-06-03"),

    ("T055", "Hydration: log shows what was drunk", "Hydration", "Mobile", "Done",
     "User: the log should show what was drunk, not just a fluid total. Plan: each entry row shows drink icon + type + size + time (e.g. 'Milk · 250 ml · 9:14a'), optionally its hydration contribution.",
     "BUILT 2026-06-03: each log row shows the drink's icon + label + size + time; milk rows carry a small 'x1.5' badge so the BHI bonus is visible.",
     "mobile/app/(app)/hydration.tsx",
     "2026-06-03"),

    ("T056", "Hydration: non-intimidating progress display", "Hydration", "Mobile", "Done",
     "User: raw fluid value (e.g. '2547 ml') is too intimidating as the headline — wants something friendlier. Options proposed: (A) percent + mascot mood, (B) 'cups' metaphor (effective hydration / ~250 ml -> N cups; milk visibly fills 1.5, making the science tangible), (C) rising water level in the existing PixelScene tied to the mascot (pure visual, most on-brand). Recommended C as the headline with a quiet 'cups' readout, raw ml only on tap.",
     "BUILT 2026-06-03: PixelScene gained a fillFrac prop that draws a translucent rising 'pond' (up to ~66% of the scene) from progress; HydrationPet passes the effective fill. Headline readout is a friendly 'X of Y cups' (a cup ~ 250 mL effective); tap it to flip to exact effective-vs-target mL.",
     "mobile/app/(app)/hydration.tsx; HydrationPet.tsx; PixelScene.tsx",
     "2026-06-03"),

    ("T057", "Hydration: attribution styling to match other pages", "Hydration", "Mobile", "Done",
     "User: the hydration attribution should look like the credit line on other detail pages — the small dimmed dot-separated footer (strength 'Epley · Brzycki · Lombardi averaged', cardio 'Riegel · Daniels' · Seiler ...'). FIXED 2026-06-03: hydration's s.attribution now matches the house tinyText style (muted, 11px, lineHeight 16, left-aligned) and was moved to the BOTTOM of the progress card as a footer credit — same placement as strength/cardio (it used to be centered between the readout and the quick-add chips).",
     "None.",
     "mobile/app/(app)/hydration.tsx",
     "2026-06-03"),

    ("T058", "Hydration: remove the rising-water pond", "Hydration", "Mobile", "Done",
     "User doesn't want the 'water flowing' effect in the pixel art. Revert the T056 PixelScene water Rects + the fillFrac prop wiring on HydrationPet/PixelScene. Keep the cups readout (that part stays) — only the in-scene water is removed.",
     "DONE 2026-06-03: removed the water Group from PixelScene + the fillFrac prop on PixelScene & HydrationPet. Scene is back to the original day/night art; the cups readout stays as the progress indicator.",
     "mobile/src/components/PixelScene.tsx; HydrationPet.tsx",
     "2026-06-03"),

    ("T059", "Hydration BUG: fl-oz unit not applied + picker dead", "Hydration", "Mobile", "Done",
     "After switching Settings fluid unit to fl oz, the hydration page stayed in mL AND the picker buttons stopped responding. Regression from the T052-T056 build. Diagnose: (a) does Settings write fluid_unit='oz' and does the profile propagate to the page; (b) why the picker Pressables stopped firing (suspect the FadeInUp Animated.View / Modal / half-loaded bundle).",
     "RESOLVED 2026-06-03: the 'unit stayed mL' part was the ATTRIBUTION line (fixed in T062); the size buttons derive from profile and switch with the unit. The dead picker/delete touches were the page-wide frozen-touch symptom (half-applied Fast Refresh + the now-removed Animated.View expansion). Closed per user 'this is done' — they've been interacting with the live picker through every follow-up tweak, so touches work.",
     "mobile/app/(app)/hydration.tsx; settings.tsx; AuthContext",
     "2026-06-03"),

    ("T060", "Hydration BUG: log entries won't delete", "Hydration", "Mobile", "Done",
     "Deleting today's log rows did nothing. deleteEntry itself was unchanged, so the regression is likely the new log-row layout (lucide Icon intercepting touches inside DeleteAction) or the same page-wide dead-touch issue as T059.",
     "RESOLVED 2026-06-03: DeleteAction + deleteEntry were unchanged & correct (two-tap: trash -> red check). The dead delete was the same page-wide frozen-touch symptom as T059, now cleared. Closed per user 'this is done'.",
     "mobile/app/(app)/hydration.tsx; DeleteAction.tsx",
     "2026-06-03"),

    ("T061", "Hydration: picker REPLACES, not expands", "Hydration", "Mobile", "Done",
     "User rejects the expanding picker. Wanted behaviour: tapping a drink type REPLACES the type-tile row with the size buttons for that drink (with a back / change-drink affordance), instead of revealing an extra row below. Removes the FadeInUp inline expansion entirely.",
     "DONE 2026-06-03: picker is a two-state swap — the 6-type grid is REPLACED by the chosen drink's size buttons; logging a size returns to the grid. No FadeInUp expansion. Refined per user: (a) size-view header is back-only — a ‹ ChevronLeft + the selected drink's icon/label, tappable to return to the grid (removed the redundant 'Change drink' hint); (b) removed the Custom button entirely — sizes are the 3 presets only (250/350/500 mL or 8/12/16 oz), and the PhantomWheel modal + its state/imports/styles came out with it (no dead code).",
     "mobile/app/(app)/hydration.tsx",
     "2026-06-03"),

    ("T062", "Hydration attribution: drop mL + source-names-first", "Hydration", "Mobile", "Done",
     "The attribution's FORMAT didn't match other pages (which lead with source NAMES then a method descriptor: strength 'Epley · Brzycki · Lombardi averaged · % of bodyweight'). First pass dropped the unit entirely; user then said they're FINE with mL showing — just wanted the format fixed. FINAL 2026-06-03: source-names-first, science descriptor LAST — hasWeight 'National Academies · Mayo Clinic · EFSA · 35 mL/kg bodyweight'; no-weight 'National Academies · EFSA · sex-based estimate'. Also the earlier 'unit stayed in mL' (T059) was this line, not the size buttons.",
     "None.",
     "mobile/app/(app)/hydration.tsx",
     "2026-06-03"),

    ("T063", "Dashboard stat pills: uniform size + grid layout", "Dashboard", "Cross", "Done",
     "User wants all stat pills EXACTLY the same size, laid out with '3 on each side' and the last pill centered under them (reads as a 2-column grid: 3 left + 3 right + 7th centered; could also mean 3-per-row — to confirm). Today the 7 pills wrap with text-driven variable widths. Needs fixed-width pills + terse uniform wording so they fit. Mobile dashboard + admin + coach (mirror).",
     "DONE 2026-06-03: built + deployed on all 3 surfaces. Mobile: statChip width 48% + minHeight 44 + radius 12, statsRow justify-center, statChipText flex:1. Web (admin + coach): SnapshotBadge -> flex w-[48%] min-h-[2.75rem] rounded-xl text-center (was an inline-flex whitespace-nowrap pill), container flex-wrap + justify-center. 6 chips form 3 rows of 2 + lone 7th centered; each pill keeps its own window wording.",
     "mobile dashboard.tsx; web AdminUserDetail.jsx, CoachClientDetail.jsx",
     "2026-06-03"),

    ("T064", "Weight pill: show change only, not current weight", "Dashboard", "Cross", "Done",
     "User: the weight pill must always show weight CHANGE, never the current weight. Today it falls back to current weight when there's only one weigh-in. Remove that fallback; show change only and hide the pill when there's nothing to compare.",
     "DONE 2026-06-03: weight chip is change-only on all 3 surfaces — removed the '· latest' current-weight fallback (mobile + admin + coach). Shows the signed change since the last weigh-in (latest − previous); the chip hides entirely when there's only one weigh-in to compare.",
     "mobile dashboard.tsx; web AdminUserDetail.jsx, CoachClientDetail.jsx",
     "2026-06-03"),

    ("T065", "PR pills: 'this month' -> rolling 'last 30 days'", "Dashboard", "Cross", "Done",
     "CORRECTED 2026-06-03 (user): ONLY the PR pills change — 'this month' (calendar) becomes 'last 30 days' (rolling 30-day window); the count logic switches from monthStart to a 30-day cutoff (a PR counts if the best-ever for that exercise/activity was hit within the last 30 days). Every OTHER pill KEEPS its existing window + wording: Food stays 'last 14 days' (still a STREAK, not a count), Lowest HR stays 'last 7 days', Sleep stays '7 nights', Hydration stays '7d', Weight stays 'since last weigh-in'. My earlier 'unify everything to 30d' reading was wrong.",
     "DONE 2026-06-03: PR pills use a rolling 30-day cutoff + read 'last 30 days' on all 3 surfaces (mobile: isThisMonth -> isWithinLast30Days; web admin + coach: monthStartISO -> thirtyAgoISO). Every other pill's window left untouched (Food 14d, HR 7d, Sleep 7 nights, Hydration 7d, Weight since last weigh-in).",
     "mobile dashboard.tsx; web AdminUserDetail.jsx, CoachClientDetail.jsx",
     "2026-06-03"),

    ("T066", "Redeploy delete-user edge function (rom_records removal)", "Backend", "Backend", "Done",
     "Flagged during T049 (rom_records drop): the delete-user edge function's SOURCE had its rom_records table-list entry removed, but the LIVE function still needed a deploy. Tracked as its own row (not buried in the closed T049) after the user's process callout.",
     "DONE 2026-06-03: deployed via MCP deploy_edge_function (CLI had no access token). delete-user is now version 12 (was 11), verify_jwt on; live USER_DATA_TABLES no longer includes rom_records. Source was already committed in 6a58fc3 — this was the runtime push.",
     "supabase/functions/delete-user/index.ts",
     "2026-06-03"),

    ("T067", "Dashboard pills: strip labels + remove emojis", "Dashboard", "Cross", "Done",
     "Screenshot review after T063 shipped: the pills look awful — labels are too verbose ('strength PRs last 30 days', 'days logged in last 14 days', 'days hit water goal · 7d') and wrap to 2-3 lines inside the 48% chips, stranding the number on its own line. User: strip the text 'just enough to be understandable' AND remove ALL emojis (the 🏆 🍴 ❤️ 😴 💧 ⚖️ prefixes). Mobile + admin + coach.",
     "DONE 2026-06-03: removed ALL emoji prefixes + stripped labels to single-line forms on all 3 surfaces — 'N strength PRs · 30d', 'N cardio PR(s) · 30d', 'N food days · 14d', 'N low bpm · 7d', '±N lb change', 'N.Nh sleep · 7d', 'N water days · 7d'. Mobile: chips single-line + centered (minHeight 34, removed flex:1 wrap on the label). Web: SnapshotBadge -> min-h-[2.25rem] + leading-tight, content centered. Built + deployed.",
     "mobile dashboard.tsx; web AdminUserDetail.jsx, CoachClientDetail.jsx",
     "2026-06-03"),

    ("T070", "Dashboard pills: page icons + highlight weight pill", "Dashboard", "Cross", "Done",
     "User (after seeing the emoji-free pills): use each PAGE's lucide icon on its pill instead of no icon — Strength=Dumbbell, Cardio=Activity, Food=Flame, HR=Heart, Weight=Weight/Scale, Sleep=Moon, Hydration=Droplet (confirm exact icons + accents from the nav/RadialNav). Also the weight pill's slate tint blends into the card background — give it a more visible highlight (a distinct colored border/bg, not clashing with the other 6).",
     "MOBILE DONE + reloaded 2026-06-03: each pill now leads with its page's lucide icon (Dumbbell / Activity / Flame / Heart / Weight / Moon / Droplet) in the page accent; weight chip switched from the near-invisible slate to a teal tint (border 0.34 / bg 0.12) so it stands out. WEB (admin + coach) MIRRORED 2026-06-04 — page icons added, built + deployed.",
     "mobile dashboard.tsx; (then) web AdminUserDetail.jsx, CoachClientDetail.jsx",
     "2026-06-03"),

    ("T069", "Dashboard pills: window-based compute + 'no recent data' (don't hide)", "Dashboard", "Cross", "Done",
     "User rejected T068's 'latest-N-logs' weight computation — wants every pill computed STRICTLY over its designated rolling window (PRs 30d, food 14d, HR/sleep/hydration 7d, weight = its own designated window). And instead of HIDING a pill when its window has no data, ALWAYS show the pill with a 'no recent data' placeholder (also keeps the fixed 3+3+1 / 7-pill layout stable). Weight window was not specified — assistant designating 30 days (shown as '· 30d'), easily changed.",
     "MOBILE DONE + reloaded 2026-06-03: weight = change over rolling 30d (latest − earliest weigh-in in the 30d window; reverted the latest-5-logs hack), shown '±N lb · 30d' or 'no recent weight'; all pills always render — count pills (strength/cardio/food) show 0, measurement pills (HR/weight/sleep/water) show 'no recent <metric>' when empty. WEB (admin + coach) MIRRORED 2026-06-04 — built + deployed.",
     "mobile dashboard.tsx; (then) web AdminUserDetail.jsx, CoachClientDetail.jsx",
     "2026-06-03"),

    ("T068", "Dashboard weight pill missing (change-only edge)", "Dashboard", "Cross", "Done",
     "After T064 made the weight pill change-only (no current-weight fallback), it vanished from the user's dashboard. Cause: the change needs 2 weigh-ins inside the dashboard's fetch window, and the user has <2 in range (their previous weigh-in is older than the window). User flagged the missing pill.",
     "DONE 2026-06-03: change now computes from the 2 most-recent weigh-ins regardless of date — mobile reuses the latest-5 `bw` fetch; web admin + coach dropped the 14-day gte on their bodyweight query. Still change-only (hides only when there's genuinely one weigh-in ever). Built + deployed.",
     "mobile dashboard.tsx; web AdminUserDetail.jsx, CoachClientDetail.jsx",
     "2026-06-03"),

    ("T071", "Dashboard weight pill color — distinct from heart", "Dashboard", "Mobile", "Done",
     "User: the weight pill (teal, from T070) and the heart pill (emerald) read as almost the same color. REVISED direction: give the weight pill the SAME green the Bodyweight page uses (so the pill matches its page), and recolor the HEART pill to something else (off green) so the two are clearly distinguishable.",
     "DONE 2026-06-03: weight pill now uses EMERALD (the Bodyweight page's accent green — reuses statChipEmerald); heart pill moved to FUCHSIA (new statChipFuchsia). Clearly distinguishable now. WEB (admin + coach) mirrored 2026-06-04 (weight = emerald, heart = fuchsia).",
     "mobile dashboard.tsx",
     "2026-06-03"),

    ("T072", "Dashboard pills — sort by tier + gate by user's tier", "Dashboard", "Cross", "Done",
     "User: re-sort the stat pills by app tier and only show the pills for the user's tier (and below). Free = 2 pills (Strength, Cardio); CoreRX adds Weight + Heart + Food; FullRX adds Sleep + Hydration. Tier->pill mapping mirrors RadialNav's NAV_BY_HREF: strength/cardio = free; bodyweight(weight)/heart/calories(food) = corerx; sleep/hydration = fullrx. Uneven count -> last pill centers (already handled by statsRow justify-center + 48% fixed width).",
     "MOBILE DONE 2026-06-03: replicated RadialNav's resolveTier + TIER_RANK in dashboard.tsx; pills reordered to tier/nav order (Strength, Cardio = free; Weight, Heart, Food = corerx; Sleep, Hydration = fullrx) and each gated on `tierRank >= TIER_RANK[pillTier]`. Free user sees 2 pills, corerx 5, fullrx 7; uneven count centers the lone last pill (statsRow justify-center + 48% width). WEB (admin + coach) MIRRORED 2026-06-04 — replicated resolveTier on both files (client tier comes from the admin get_user_for_admin RPC + the coach profiles select('*'), both already return b2c_subscription_tier); reorder + tier-gate + icons + colors + no-recent states + 30d weight; built + deployed. So a coach-attached client (fullrx) shows all 7; a free B2C client shows 2.",
     "mobile dashboard.tsx; (then) web AdminUserDetail.jsx, CoachClientDetail.jsx",
     "2026-06-03"),

    ("T073", "Web empty-state pills lost their accent color (all grey)", "Dashboard", "Web", "Done",
     "Screenshot (admin client view): the empty 'no recent weight / HR / sleep / water' pills all render grey (zinc) instead of keeping their accent tint, so weight (emerald) and heart (fuchsia) are indistinguishable when empty AND it doesn't match mobile — where an empty pill keeps its accent chip border/bg with muted icon + muted text. The T069-T072 web-mirror agent used color='zinc' for all empty states.",
     "DONE 2026-06-04 (background agent + assistant deploy): added a `muted` prop to SnapshotBadge — keeps the accent border/bg from `color` but forces text to `!text-muted-foreground`; the 4 empty pills now pass their accent color + muted (weight=green, heart=fuchsia, sleep=indigo, hydration=cyan) with muted icon+text, matching mobile. Both web files; built + deployed.",
     "web AdminUserDetail.jsx, CoachClientDetail.jsx",
     "2026-06-04"),

    ("T074", "Phantom water_logs appearing for users who never logged water", "Hydration", "Backend", "Done",
     "User: water_logs rows exist for accounts that never logged water. rasp_86 (Test Client, id 3d41a692-7c59-4636-bfee-96b76d1bcf3e) shows 2 days logged with NO visible log entries; motaz.jarrah shows 1 day, no log history. The mobile hydration pill reads '0 days' (NOT 'no recent') — so the count path SEES water_logs rows but counts 0 days hitting goal, i.e. rows exist but don't surface in the page's log list and the user doesn't know what's inserting them.",
     "RESOLVED 2026-06-04 — NOT a phantom inserter, it's real test taps. The rows' amounts are EXACTLY 236.588 / 354.882 / 473.176 ml = 8 / 12 / 16 oz (the hydration page's oz quick-add/picker presets run through ozToMl()); logged_at = created_at, clustered in bursts seconds apart on 2026-05-29 + 05-30 (rasp_86, ~18 rows) and 06-03 22:53 (motaz, 1 row) — i.e. manual button-mashing during the hydration-page dev/test sessions. The ONLY water_logs INSERT in the whole codebase is hydration.tsx addDrink (tap-triggered; grep confirmed nothing auto-inserts — sync/HC/Samsung/migrations don't write water_logs, they only DELETE in delete-user). 'No log showing' = the page's log list shows TODAY only (these are past days); 'pill says 0' = computeHydrationDaysHit counts days HITTING the daily goal in 7d (=0, each day was under goal) but rows exist so it reads 0 not 'no recent'. CLEANUP DONE 2026-06-04 (user said yes): deleted all 19 junk rows (rasp_86 18 + motaz 1) scoped by user_id; both accounts now read 0 water_logs. No code change — nothing was auto-inserting.",
     "mobile hydration.tsx; src/lib/healthConnect.ts, integrations/*; supabase functions/migrations",
     "2026-06-04"),

    ("T075", "Stat pills: a 0 count should read 'no recent', not '0'", "Dashboard", "Cross", "Done",
     "User (admin screenshot): count pills show '0 cardio PRs · 30d' and '0 food days · 14d' while measurement pills show 'no recent weight/HR/sleep/water' — inconsistent. The count-vs-measurement split I built shows the number even at 0 for PRs/food. Make the count pills read the muted 'no recent ...' empty state (accent-tinted, like the measurement empties) when the count is 0, and show the number only when > 0. (0 PRs = no PR set in the window — the client may still have trained — so 'no recent strength/cardio PRs' is accurate.)",
     "DONE 2026-06-04: count pills now show muted 'no recent strength PRs / no recent cardio PRs / no recent food' (accent-tinted, like the measurement empties) when the count is 0, and the number+label only when > 0. Mobile + admin + coach; built + deployed. All 7 pills now read consistently — a value when there's something recent, 'no recent <metric>' otherwise.",
     "mobile dashboard.tsx; web AdminUserDetail.jsx, CoachClientDetail.jsx",
     "2026-06-04"),

    ("T076", "Mobile 'Coached by' shows full name, should be first only", "Dashboard", "Mobile", "Done",
     "On the client's mobile dashboard the 'Coached by …' badge showed the coach's FULL name (e.g. 'Coached by Taz Jarrah'); should show first name only ('Coached by Taz'). The coachInfo handler already stripped to split(' ')[0], but a stale cached coachInfo (hydrated from dataCache before the stripping shipped) rendered the full name.",
     "Fixed: the badge now strips to the first word at RENDER time — (coachInfo.full_name ?? '').trim().split(' ')[0] — so it shows first name only regardless of what's cached. (Athlete dashboard is mobile-only, no web mirror.)",
     "mobile dashboard.tsx",
     "2026-06-04"),

    ("T077", "Mobile sign-in heading 'Sign in to MyRX' -> 'Sign in'", "Auth", "Mobile", "Done",
     "Mobile sign-in page had an eyebrow 'Sign in' + a title 'Sign in to MyRX' (redundant); user wants just 'Sign in'.",
     "Fixed: removed the redundant eyebrow and set the title to 'Sign in' (single clean heading). Web Auth.jsx differs — heading is 'Welcome back' + subtitle 'Sign in to continue to MyRX', not the same string — so left as-is (mobile-only fix).",
     "mobile/app/(auth)/sign-in.tsx; web Auth.jsx",
     "2026-06-04"),

    ("T078", "Finalize the Admin client-detail view", "Coach/Admin", "Web", "Done",
     "User 2026-06-04 re-scoped the parity task: before mirroring admin->coach (T015), finalize the ADMIN client-detail view first so we have a full understanding of what exists + what needs mirroring. Admin client page = web/src/pages/admin/AdminUserDetail.jsx (shell) with 6 tabs: Dashboard, Efforts (AdminUserActivity), Bodyweight (AdminUserBody), Calories (AdminUserCalories - Food Log/Manual Logs/Macro Plan sub-tabs), Billing, Activity Feed (timeline).",
     "Inventory done 2026-06-04. FIXES 2,3,4,5,7 DONE + deployed 2026-06-04 (bundle index-IUBSQ3c3): (2) added missing Loader2 import in AdminUserBody.jsx — weigh-in save no longer crashes; (3) deleted dead tabs/AdminUserProfile.jsx (530 lines) + its import; (4) added 'billing' to validTabs — Billing tab deep-linkable; (5) Efforts cardio 'Best' now direction-aware across all formats (added parseCardioBest mirror), not just /km; (7) fixed stale ClientSettingsDrawer docstring + unified the Calories manual-log save button to the Loader2 spinner pattern. DECISIONS DISPOSED 2026-06-04: (1) Dashboard -> KEEP as placeholder for now (it's the holistic rollup of the other pages; build the underlying pages first) -> folded into T079; (6) Food-log add-on-behalf -> deferred to the coach-INPUT track (coach input is a different approach, TBD). Finalize-pass complete.",
     "web/src/pages/admin/AdminUserDetail.jsx; web/src/pages/admin/tabs/*.jsx; components/{BillingView,MacroPlanEditor,AccountSettings,ClientSettingsDrawer}.jsx",
     "2026-06-04"),

    ("T079", "Mirror athlete pages read-only in the admin client view", "Coach/Admin", "Web", "In progress",
     "User 2026-06-04: build the client-view tabs so each athlete page is reflected READ-ONLY — a coach must first see exactly what the athlete sees (per-page coaching add-ons come later, one page at a time). Coach side has NO inputs (coach-input is a separate approach, TBD). Dashboard stays placeholder (holistic rollup of the others, built last). Athlete data pages: Strength, Cardio, Bodyweight, Calories, Sleep, Heart, Hydration. Admin already covers Strength/Cardio (Efforts), Bodyweight, Calories. GAPS added now: Sleep, Hydration, Heart.",
     "Phase 1 (build pages): add read-only mirror tabs AdminUserSleep / AdminUserHydration / AdminUserHeart (styled like existing admin tabs: Recharts, design tokens), wired into AdminUserDetail. Built via parallel agents 2026-06-04. Phase 2 (per-page, one at a time): align every tab first-view to the athlete exactly + add coaching elements.",
     "web/src/pages/admin/tabs/AdminUser{Sleep,Hydration,Heart}.jsx; web/src/pages/admin/AdminUserDetail.jsx; mirror mobile/app/(app)/{sleep,hydration,heart}.tsx",
     "2026-06-04"),

    ("T080", "Efforts page — align coach view to athlete + coaching add-ons", "Coach/Admin", "Web", "In progress",
     "Per-page pass (Phase 2 of T079), starting with Efforts at user request 2026-06-04. Goal: make the admin Efforts tab's FIRST view show what the athlete sees on their Strength + Cardio pages (read-only, no inputs — coach input is a separate track), then propose coaching add-ons. Open structural question: athlete has SEPARATE Strength + Cardio pages; admin combines them into one 'Efforts' tab (AdminUserActivity.jsx: filter pills + per-movement cards -> detail route). Decide split-vs-combined + remove the Add-effort input.",
     "STRUCTURE DECIDED 2026-06-04: keep ONE combined Efforts tab with an internal Strength|Cardio segmented toggle (dropped 'All', default Strength). LIST-VIEW read-only mirror DONE + deployed 2026-06-04 — removed the Add-effort form + backdate + button; AdminUserActivity is now read-only (207->193 lines). USER 2026-06-04: FINISH THE MIRROR BEFORE ADD-ONS. MAIN REMAINING = (1) DETAIL-SCREEN rebuild to mirror the athlete coaching surfaces, incremental: Strength Weighted Standard FIRST (validate with user), then other strength variants (bodyweight/assisted/carry/iso/sled), then cardio surfaces (pace/airbike/swim/ruck/stairmill/ergs/duration); (2) LIST variant-collapsing (bodyweight tiers / swim strokes / sled push-drag -> one row+badge like the athlete; coupled to the detail routes). DEFERRED until mirror complete: coaching add-ons (recency, stalled flag, frequency, PR timeline, balance, adherence). Detail rebuild notes: admin detail routes (AdminEffortDetail = bare 10-tile grid; AdminCardioDetail = old pace model w/ banned 'lower=faster') don't mirror the athlete coaching surfaces; KEEP the per-effort swipe-delete there. USER CORRECTION 2026-06-04: the coach is NOT fully read-only — they CAN delete efforts; only the athlete-style ADD/log form is excluded (that's the separate coach-input track). List view stays add-free; individual-effort delete lives in the detail screens. PROGRESS 2026-06-04: Strength Weighted Standard detail mirror DONE + deployed (new web/src/pages/admin/detail/AdminStrengthWeightedDetail.jsx, dispatched from AdminEffortDetail for barbell/dumbbell/kettlebell/machine/strongman; read-only + delete kept; rep-max projections + adp-zone pills + 20 tiles + next-target hero + chart + log). AWAITING user validation of this first variant before rolling out remaining strength variants (bodyweight/assisted/carry/iso/sled) + all cardio surfaces + list variant-collapsing.",
     "web/src/pages/admin/tabs/AdminUserActivity.jsx; web AdminEffortDetail/AdminCardioDetail; mirror mobile/app/(app)/{strength,cardio}.tsx + effort detail pages",
     "2026-06-04"),

    ("T081", "Roll out Efforts detail mirrors — all strength + cardio variants + list collapsing", "Coach/Admin", "Web", "Done",
     "User approved the Weighted Standard pattern (T080) 2026-06-04 and said 'proceed with all' (multiple agents OK if no issues). Roll the AdminStrengthWeightedDetail read-only+delete pattern out to every remaining Efforts detail variant. Strength dispatch (in StrengthDetail, by movement looked up by BASE name after stripping ' [Band + Knee]'/' [Band]'/' [Knee]'): isSledWorkConsolidated (exercise==='Sled Work'), isIsometric (strength_type==='isometric'), isAssistedMachine (equipment==='assisted'), isCarry (equipment==='carry'), band/knee suffix && base equipment!=='bodyweight' -> RepsOnly, isBodyweightExercise (equipment==='bodyweight') -> bodyweight consolidated, else weighted (done). Cardio route is SEPARATE: /admin/user/:id/effort/cardio/:slug -> AdminCardioDetail (variants: pace/Group-A incl Concept2 ergs, air bike, swimming consolidated, rucking, stairmill, duration, beat-your-best).",
     "WAVES (verify each): W1 strength variants (AdminStrength{Bodyweight,Assisted,Carry,Isometric,RepsOnly}Detail) -> wire AdminEffortDetail base-name dispatch -> build/deploy/QA; W2 cardio variants -> wire AdminCardioDetail; W3 list variant-collapsing in AdminUserActivity (bodyweight tiers / swim strokes / sled push-drag -> one row+badge). W1 DONE + deployed 2026-06-04: 5 strength variants (bodyweight/assisted/carry+sled/isometric/repsonly) built via parallel agents + wired into AdminEffortDetail base-name dispatch (Sled Work handled by exercise name; isometric via strength_type; band/knee suffix + non-bodyweight -> repsonly); build clean. W2 (cardio) DONE + deployed 2026-06-04: 6 variants (pace incl Concept2 Row/Ski erg; air bike; swimming consolidated; rucking; stairmill; beat-your-best) built via parallel agents; AdminCardioDetail rewritten as a slim dispatcher mirroring the athlete categorizeActivity (Bike Erg -> stationary_bike -> BeatYourBest; Row/Ski Erg -> Pace); build clean. W3 DONE + deployed 2026-06-04: AdminUserActivity collapses bodyweight tiers (base + tier badge FULL RX/BAND/KNEE/B+K via useMovements) / swim strokes (Swimming + FREE/BACK/BREAST/FLY badge) / sled push-drag (Sled Work + PUSH/DRAG badge) into one row each, navigating to the consolidated detail. ALL WAVES COMPLETE — the entire Efforts page (read-only list + every strength & cardio detail variant) now mirrors the athlete with per-effort delete kept. Next for Efforts: the deferred coaching add-ons (recency, stalled flag, frequency, PR timeline, balance, adherence).",
     "web/src/pages/admin/detail/AdminStrength*.jsx + AdminCardio*.jsx; AdminEffortDetail.jsx; AdminCardioDetail.jsx; AdminUserActivity.jsx; mirror mobile [exercise].tsx + [activity].tsx",
     "2026-06-04"),

    ("T082", "Cross-check ALL Efforts detail mirrors vs ACTUAL mobile render + fix", "Coach/Admin", "Web", "Done",
     "ROOT CAUSE (user 2026-06-04): the detail mirrors were built partly from the CLAUDE.md 'locked design spec' sections, which LAG the actual mobile code. StairMill was built to the stale zone-pill spec instead of the current mobile tile/plan-queue render, so it doesn't look like mobile. The mobile COMPONENT CODE is the sole source of truth for the visual — CLAUDE.md is context/intent only, not the render spec. Likely affects other variants too.",
     "Verify+fix agent per surface (6 strength + 6 cardio): read the ACTUAL mobile render for the variant, diff vs the web mirror, fix the web to match the mobile render (Recharts/click-pills substitutions OK; read-only + delete kept). StairMill = full rebuild to the current mobile model. Then build/deploy + user QAs every move. PROGRESS 2026-06-04: agent BURST limit when spawning 7 at once (single agents work) — running the sweep in small batches of ~3. StairMill REBUILT to the plan-queue tile model (matching actual mobile: 8-step queue tile row + selected-step hero, not the stale zone-pill) + deployed. ALL 12 surfaces cross-checked + fixed 2026-06-04 (batches of 3, mobile code = truth). Real stale-spec divergences corrected: weighted (restored hero title spec said removed), bodyweight (removed phantom band-progression strip), assisted (removed invented fade-edge gradients), isometric (removed phantom in-card phase-pill + info-panel; restored hero title + min-height), repsonly (removed invented STRENGTH pill + chart placeholder; added log header), pace + stairmill (FIXED hero/tiles ORDER inversion — mobile is tiles->hero, not hero->queue; removed invented 'COMING UP' heading + tile rest line + fade-edges), rucking + beatyourbest (chart headings, removed/added captions + log header). AirBike + Swimming were near-clean (copy only). All eslint+esbuild clean, deployed.",
     "web/src/pages/admin/detail/AdminStrength*.jsx + AdminCardio*.jsx; mirror mobile [exercise].tsx + [activity].tsx",
     "2026-06-04"),

    ("T083", "Mobile dashboard pills don't refresh on effort delete", "Dashboard", "Mobile", "Done",
     "User 2026-06-04: on the mobile dashboard, deleting an effort does NOT update the stat pills in real time — the user must navigate away and back (which remounts + refetches) to see the corrected pill values. The pills derive from efforts data that isn't refetched/recomputed after an on-dashboard delete.",
     "FIXED 2026-06-04 (done directly — agent quota was rate-limited): handleDelete optimistically updated the recent-activity list but never re-ran fetchDashboard, which computes the pill aggregates (PRs/streak/HR/weight/hydration/sleep/food) from their own queries. Added fetchDashboard() right after the delete; it never sets loading=true so there's no skeleton flash. tsc clean.",
     "mobile/app/(app)/dashboard.tsx",
     "2026-06-04"),

    ("T084", "Cardio copy: 'per rep' -> 'per interval' (StairMill + sweep)", "Cardio", "Mobile+Web", "Done",
     "User 2026-06-04: StairMill (and other cardio) display copy says 'per rep' / 'between reps' / 'Each rep' — cardio prescriptions are INTERVALS, not reps. Fix the user-facing copy to interval terminology where it makes sense (hero row descriptors + cue lines + zone whyText). Internal code identifiers (reps, repKm, repSecs fields) stay. Pace already uses 'per interval' for its descriptors.",
     "Mobile [activity].tsx: StairMill hero rows (4455/4461) + cues (4155/4158) + VO2 whyText (4090), Air Bike hero (3609) + cues (3327/3330) + threshold whyText (3259), Swimming cue (693) -> interval wording. Mirror to web AdminCardio{StairMill,AirBike,Swimming}Detail.jsx. tsc + build/deploy + reload. DONE 2026-06-04: hero descriptors 'per rep'->'per interval' + 'est. per rep'->'est. per interval'; cues 'Each rep'->'Each interval', 'between reps'->'between intervals'; whyText 'Sustained hard reps'->'intervals' + 'Longer reps than sprint'->'Longer intervals'. Pace/Duration already used interval wording. Internal field names (reps/repKm/repSecs) + code comments left as-is. tsc clean; web deployed; mobile Fast-Refresh.",
     "mobile/app/(app)/effort/cardio/[activity].tsx; web AdminCardio*Detail.jsx",
     "2026-06-04"),

    ("T085", "Hydration: verify BHI weighting wired into cups + attribution states the science", "Hydration", "Mobile", "Done",
     "User 2026-06-04: (1) we researched the Beverage-Hydration-Index weighting for how much hydration each drink gives (milk ~1.5x, water/coffee/tea/diet-soda ~1.0x). Is it ACTUALLY wired into the CUPS number displayed, or are cups computed from raw mL? (2) The attribution/science line on the hydration page doesn't state the science the way other pages do (strength 'Epley · Brzycki · Lombardi'; cardio 'Riegel · Daniels'...'). Verify both + fix.",
     "Read mobile/app/(app)/hydration.tsx: confirm cups = round(effectiveBhiMl/250) of round(targetMl/250) using the BHI multiplier (not raw mL); check the attribution line content + styling. Fix the cups wiring if wrong + add/correct the BHI attribution line to match other pages. VERIFIED+FIXED 2026-06-04: (1) BHI IS wired correctly — DRINKS registry sets multipliers (milk 1.5, rest 1.0, Maughan 2016), todayEffectiveMl = sum(amount_ml × multiplierFor(drink_type)), cupsDone = round(todayEffectiveMl/250); milk even shows a ×1.5 badge. No cups fix needed. (2) Attribution credited ONLY the target formula (35 mL/kg, National Academies/Mayo/EFSA), not the BHI — added 'Maughan 2016 · ... hydration-weighted' to mobile hydration.tsx + mirrored to web AdminUserHydration.jsx. tsc clean, deployed.",
     "mobile/app/(app)/hydration.tsx",
     "2026-06-04"),

    ("T086", "Strip athlete-only explanatory copy from web coach mirrors", "Coach/Admin", "Web", "Done",
     "User 2026-06-04 (viewing web AdminUserHydration): athlete-facing explanatory text has no value to a coach + clutters the coach view. Flagged examples: motivational 'Steady sips beat chugging it all at once.', eligibility note 'Only no- and low-calorie... milk included.', and the attribution 'National Academies · ... Maughan 2016 ...'. Sweep ALL web coach-mirror pages, judge coach-meaningful vs athlete-only, REMOVE the unnecessary explanations. Coach view = client DATA + actionable prescription; strip athlete motivation, how/what-to-log instructions, eligibility notes, feature help-text, science attributions. These STAY on the MOBILE athlete pages — only the WEB coach mirrors get stripped (so this intentionally DIVERGES the mirrors from mobile; the T082 cross-check no longer applies to these prose blocks).",
     "REMOVE (always-visible athlete prose): attribution footers, motivational/encouragement lines, eligibility/instruction notes, feature help-text subtitles. KEEP: data (charts/logs/best/PR/current), prescription numbers + the cue line, headers/badges/tiles, and the opt-in 'why this zone' info pills (collapsed, not clutter — flag if user wants those gone too). Surfaces: tabs AdminUser{Sleep,Hydration,Heart}; detail AdminStrength* + AdminCardio*. Hydration done as exemplar, then agent-sweep the rest in batches of 3; build/deploy. DONE 2026-06-04: Hydration (me) + Batch 1 (Sleep/Heart/Weighted-strength) + a background Workflow over the remaining 11 detail surfaces (5 strength + 6 cardio) in chunks of 3. Removed: attribution/citation footers, feature help-text subtitles ('Pick an adaptation zone...', 'Pick a training focus...'), tier-criteria methodology subtitles, motivational/eligibility prose; neutralized athlete-instruction empty states ('Log your first...' -> 'No X logged yet') + removed PartyPopper/emoji decoration. KEPT: all data (charts/tiles/logs/PRs/hero numbers), the prescription cue lines, section titles, badges, and the opt-in 'why this zone' info pills. User confirmed keeping the info pills. All eslint+esbuild clean (agents self-verified); web build clean; deployed. Coach mirrors now intentionally diverge from mobile on these prose blocks (mobile athlete pages keep them).",
     "web/src/pages/admin/tabs/AdminUser{Sleep,Hydration,Heart}.jsx; web/src/pages/admin/detail/AdminStrength*.jsx + AdminCardio*.jsx",
     "2026-06-04"),

    ("T087", "Strength cue: build-up/progression science to reach a new rep-max PR", "Strength", "Mobile+Web", "In progress",
     "User 2026-06-04 (thinking out loud, viewing Bench Press 5RM): the weighted-standard coaching cue 'Do 4-5 sets of 5 reps at 140 lb' is off. 140 is the NEXT 5RM target (current 5RM=135). Two problems: (1) a 5RM means 5 reps is your 1-set MAX, so you physically can't do 4-5 SETS of 5 at your 5RM — multi-set work must be submaximal (reps in reserve); (2) you don't hit a new PR cold — you build up over sessions. User senses there should be progression science (progressive overload / RIR / double-progression) + a separate PR-attempt, not one impossible cue. Asked 'am I right?' -> YES.",
     "BRAINSTORM stage — do NOT implement until the model is agreed. Design a progression-aware strength cue that separates (a) the WORKING prescription (submaximal, RIR-based, day-to-day), (b) the PROGRESSION rule (when to bump load — e.g. double progression: add reps to top of range across all sets, then +smallest jump), (c) the PR ATTEMPT (one top set at the target, periodically). Affects the locked Weighted Standard next-target card + ADP_ZONE defaults (sets/RIR/rest). Plain-English, one decision at a time. RESEARCH DONE 2026-06-04 (deep-research harness returned empty; did it via WebSearch): EVIDENCE-DICTATED answer -> working sets must be SUBMAXIMAL with ~2-3 reps in reserve, NOT the rep-max. (a) Strength gains are ~independent of proximity-to-failure (Refalo 2023 dose-response meta-regression; Schoenfeld) so no need to grind to max — leave reps in reserve. (b) Prilepin's table caps reps-per-set BELOW the single-set max at each %1RM (~85% -> 2-4 reps/set, not 5) — the direct 'why you can't do 4-5 sets of 5 at your 5RM'. (c) Double progression confirmed (Plett/Schoenfeld PeerJ 2022 — rep AND load progression both work). (d) Dose zones match the ACSM position stand (strength >=80% 1RM low reps long rest; hypertrophy 6-12RM 1-2min; endurance <60% >15 reps <90s). SO: working weight ~= the 7-8RM (a couple reps in reserve), double-progress up to the 140 target, 140 = periodic TEST. Attribution credit for the prescription side: 'Prilepin · Helms/Zourdos RIR · ACSM' (+ optionally Schoenfeld). Cue redesign pending user go-ahead.",
     "mobile/app/(app)/effort/strength/[exercise].tsx (Weighted Standard cue + ADP_ZONE_CONFIG); web AdminStrengthWeightedDetail.jsx; CLAUDE.md Weighted Standard spec",
     "2026-06-04"),
    ("T088", "Science-audit EVERY progression model (strength + cardio) for next-target correctness", "Strength+Cardio", "Mobile (+Web mirror)", "In progress",
     "User 2026-06-04/05 (after T087): asked to cross-check every movement category, not just barbell lifts — read all moves per category for strength AND all cardio, and research the science so each category's NEXT PROGRESSION is actually correct. Pulled the live movements table: 11 distinct progression models (6 strength, 5 cardio). Ran 11 parallel evidence audits via WebSearch/WebFetch.",
     "VERDICTS by severity. TIER A (correctness/safety): (A1) Olympic & ballistic lifts (snatch, C&J, power clean, KB swing/snatch) must NOT use rep-max math — a 20RM snatch is meaningless + suggests a dangerous practice; route to low-rep %1RM/velocity model. (A2) Rucking has NO bodyweight cap + mislabeled tiers (Tough=35lb is actually GoRuck HEAVY dry; true Tough=20/30lb by BW); 80lb ladder top = 53% BW for a 150lb user, past the ~1/3 BW safe limit — make tiers BW-relative + add ~1/3 BW hard cap. (A3) Plyometrics/jumps/jump-rope treated as max-reps — wrong; progress on OUTPUT (jump height/distance, box height) + contact budget + rest. (A4) StairMill: FPM is a machine SETTING not earned capacity (+ floor=16 vs 22 steps across models, peak-of-any-effort inflates); anchor on sustained held SPM / vertical m-per-min over a continuous block. TIER B (anchors mis-set): (B5) Bodyweight graduation 10 reps -> ~5-8 (10 reps = endurance, ~70-75% 1RM). (B6) Isometric: tag axis time|load|leverage — time holds keep 2-min cap, loadable holds add load after ~30-60s, skill-leverage holds (planche/levers/flag/L-sit) drop the 120s grid -> variant ladder. (B7) Pace cardio: anchor on Critical Speed (2-3 efforts) not single best pace; zones as %threshold per modality (power for ergs/bike, HR/RPE elliptical) not fixed running s/km offsets. (B8) Swim: replace Riegel single-point CSS with 2-point linear slope per stroke (1.06 is a running exponent, worse for fly/breast); kill 'fastest projection' (biased high); deepen VO2 to CSS-8..-10. (B9) Air bike: watts=cal/min*17.4 is arithmetically ok at 25% GE but 25% is optimistic AND doesn't match the console (~24.75 W/cal-min on Assault) so 'hold >=X W' is unreadable — drop overlay or calibrate to console + name brand. (B10) Weighted 15/20RM tiles are guesses (cap grid ~10-12RM); + the T087 cue fix (submaximal + RIR + double progression). TIER C (copy/labels): hypertrophy '6-12' too strict (growth 5-30+ near failure); carry 'load+distance same session' (zones already do one-axis); assisted 'BW-assist->1RM' is a relative index not a true 1RM (counterweight is constant through ROM, pull-up hardest at bottom) — relabel + add reps/negatives; soften 'polarized 80/20 is best' (pyramidal = for recreational); CLAUDE.md air-bike '17.4 standard Assault/Echo conversion' claim is FALSE; StairMill cites Honda 2014 (a GLUCOSE study, drop it), Allison was 20s not 60s, Boreham was accumulated bouts not 20-min continuous; StairMill gender baseline 12 FPM ~= 192 SPM EXCEEDS the machine max (units error, real ~4-5 FPM). SOUND AS-IS: loaded carry (one copy tweak), the swim/pace 3-zone + easy-floor STRUCTURE. CONSOLIDATED CITEABLE SOURCES: strength = Epley/Brzycki/Lombardi + Reynolds-Robergs 2006 + Schoenfeld rep-continuum 2021 + NSCA Haff&Triplett + Prilepin + Helms/Zourdos RIR + ACSM; bodyweight = Ebben 2012 + Nuzzo 2023 (ECC:CON 1.41) + Steven Low; plyo/skill = NSCA plyo + Markovic&Mikulic 2010 + Ramirez-Campillo + McGill + GMB; iso = Oranchuk 2019 + Schott 1995 + Kitai&Sale 1989 + McGill; carry = McGill strongman EMG + Winwood/Keogh/Cronin + NSCA + Dan John; pace = Daniels VDOT + Seiler 80/20 + Critical Power (Jones/Vanhatalo) + Coggan FTP + ACSM; swim = Wakayoshi 1992 + Maglischo + Costill + Gonjo 2024; air bike = Peronnet&Massicotte 1991 + Concept2 formula + Gibala SIT + Seiler; ruck = Knapik + US Army ATP 3-21.18 + GoRuck + EIB/Ranger 12mi; stairmill = Allison 2017 + Boreham 2000 + ACSM + Ainsworth Compendium + Seiler. STATUS: research complete, awaiting user TRIAGE on what to implement — most surfaces are locked, but math/bug fixes (StairMill baseline, air-bike watts, ruck BW cap, Olympic-lift rep-max) are allowed on locked surfaces per the finalized-surfaces rule; design changes need explicit unlock. PROGRESS (full model-by-model sweep, user chose 'full sweep' 2026-06-05): Model 1 Fix 1.1 (submaximal working-weight cue 'a weight you could do K+reserve, do K' + corrected per-zone RIR 2/2/1 + iterative 'add X each time, work up to PR') SHIPPED mobile+web. Fix 1.2 (Olympic barbell lifts routed OFF the rep-max grid to a NEW Layout 9 %-of-best card: added movements.lift_type column [olympic|ballistic], tagged 22 barbell snatch/clean/jerk-family lifts; mobile OlympicLiftDetail + web AdminStrengthOlympicDetail mirror + dispatch-before-weighted + Layout Design.xlsx row 9 + CLAUDE.md Layout-9 spec) SHIPPED. Fix 1.3 (capped the rendered rep-max grid at 15RM, dropped 16-20 as noise, flagged 13-15RM with a leading '≈' + caveat note; kept the 3-zone model since Endurance needs 13+) SHIPPED mobile+web. Fix 1.4 (hypertrophy whyText softened — growth spans ~5-30+ reps near failure, 6-12 is just the efficient middle, not an exclusive window) SHIPPED mobile+web. Fix 1.2b (ballistic kettlebell -> NEW Layout 10 bell-ladder card: tagged 13 KB ballistic moves lift_type='ballistic'; mobile BallisticLiftDetail + web AdminStrengthBallisticDetail mirror + dispatch-before-weighted + Layout xlsx row 10 + CLAUDE.md spec) SHIPPED. MODEL 1 COMPLETE (weighted cue + Olympic Layout 9 + ballistic Layout 10 + grid cap + hypertrophy copy). Model 2 (bodyweight rep-strength + assist tiers) IN PROGRESS: Fix 2.1 (graduation trigger BW_GRADUATION_REPS 10->8 so tiers/bands graduate in the strength range not endurance; mobile+web+CLAUDE.md) SHIPPED. Fix 2.2 (band/knee): on inspection there is NO false 'band is harder' copy — only the carousel order asserts it, which is load-bearing (graduation path + badge + landing) and the science says the order is AMBIGUOUS not wrong; user chose #1 = LEAVE the order (deliberate no-op). Fix 2.3a (bodyweight in the weighted 1RM): VERIFIED ALREADY HANDLED — the log form stores estimate1RM(profileBodyWeight + addedWeight, reps) [strength.tsx ~L414/424] and the detail projection derives added = projection - bodyweight [L4733]; the audit's 'ignores bodyweight' claim is FALSE for our code. (Only nuance: uses FULL bodyweight = 100% share, so push-ups over-count vs pull-ups — an optional minor per-move-share tweak, not a bug.) NET: Model 2 CORRECTNESS DONE — the graduation 10->8 (2.1) was the one real fix. The rest of the audit's Model-2 items are FEATURE additions, not correctness bugs: an eccentric/negatives tier + leverage/tempo/unilateral alternatives (both optional, additive builds touching the locked card + log form). META-LESSON: this is the 3rd audit claim our code already handles correctly (band/knee, BW-in-1RM) — the audit agents researched GENERIC science and assumed naive implementations; our code is more sophisticated. So VERIFY every audit claim against the actual code before 'fixing' it. Model 3 (isometric) IN PROGRESS: VERIFIED real (unlike Model 2) — all 38 iso moves were on one 10-120s time grid incl. skill holds where 120s is meaningless. LEVERAGE family SHIPPED: added movements.hold_type column (time/load/leverage), tagged 18 leverage holds; NEW Layout 11 -> mobile LeverageHoldDetail + web AdminStrengthLeverageDetail mirror; short milestones 5/10/15/20/30s + skill-ladder (planche tuck->straddle->full, levers tuck->full, handstand wall->free; gate = 30s clean -> next variant; L-Sit/V-Sit standalone per user); dispatch before isometric; Layout xlsx row 11 + CLAUDE.md Layout-11 spec + iso spec updated to exclude leverage. Model 3 (isometric) COMPLETE: leverage family (Layout 11) + LOAD family (Layout 12) both SHIPPED. Load: tagged 7 loadable holds (wall sit, calf-raise, glute-bridge x2, dead hang, split-squat, squat hold) hold_type='load'; mobile LoadHoldDetail + web AdminStrengthLoadDetail mirror; two phases (time milestones 15-60s -> add-load; gate 60s, loaded target 30s; load inc 5lb/2.5kg); LOG FORM gained an Added-weight wheel for load-holds (label 'Name · W unit x D sec', value stays 'D sec' so parseDurationSecs unchanged); chart adaptive (load over time once weighted, else hold time); dispatch before isometric; Layout xlsx row 12 + CLAUDE.md Layout-12 spec + iso spec updated. The ~13 TIME holds (plank/side-plank/hollow/etc.) stay on IsometricDetail (correct). NEXT: Model 4 (assisted machine) with verify-first discipline, then carry + cardio models 5-11. OPTIONAL DEFERRED: BW eccentric/negatives tier + per-move-share BW 1RM (Model 2). === ROUND 2 QA FEEDBACK (2026-06-05), 6 items + overarching, NOT YET DONE: (1) Bench/weighted cue — each line must be SINGLE-LINE; restructure into bullet STEPS (work set / could-do-X-do-Y / add-each-time / work-up-to-PR / rest). (2) Power Clean / Olympic (Layout 9) cue — RESEARCH proper Olympic loading: express 'do X at ~50% (~Y lb)' or 'X x Y lb (~50% 1RM)' + a build-up-to-weight per-session progression; resolve how to pick the weight for a given rep/intensity (technique vs build vs peak) beyond 3 reps. (3) KB Swing / ballistic (Layout 10) — NEVER put attribution in the cue; make REST specific (research required rest / 'full rest', maybe = work time); does the bell-ladder benchmark apply to ALL bells incl 40kg? (4) Pull Up / BW consolidated (Layout 1) — shared chart wrongly mixes 13 light-band reps + 5 full-RX on ONE curve; SPLIT chart PER PILL/tier + update Layout xlsx. (5) Planche / leverage (Layout 11) — user wants the SWIMMING consolidation pattern: variant pill-swipe + tile grid (5 tiles), NOT separate per-variant pages. (6) Wall Sit / load (Layout 12) — 'Add load' is a CUE shown as a TITLE (fix title); NO target tile (build a 5-10 tile grid like Layout 1 but TIME-UNDER-TENSION tiles, added weight per tile); attribution-line format inconsistent. OVERARCHING (LOCKED DIRECTION): EVERY move that is a variation of another must follow the SWIMMING pattern end-to-end — DB naming 'Move [Variant]', consolidated detail w/ variant pill carousel, collapsed index/log row w/ variant badge. Read swimming impl end-to-end + replicate exactly (per-move design). ROUND-2 PROGRESS: [DONE] #1 cue format — user REJECTED bullets, so instead LOCKED ONE prose cue format SYSTEM-WIDE (one flowing sentence, commas not em-dashes, no bullets, no attribution-in-cue, numbers auto-emphasized w/ weights blue) behind a shared CueText component (mobile/src/components/CueText.tsx + web/src/components/CueText.jsx); swept EVERY coaching cue mobile+web (strength + cardio) onto it — zero em-dashes remain — then migrated the 4 hand-done cues (bench/ballistic/leverage/load) off inline spans onto CueText so there is literally ONE cue code path (commit db85128, web deployed). [DONE] #3 KB Swing -> attribution removed from cue/benchmark (credit on its own line), concrete rest 'at least as long as the set takes (~1:1, power needs full recovery)', benchmark scoped to <=32kg/70lb (hidden past the S&S standard), last-bell now shows a 'top bell' state instead of a broken move-up-to-same. RESEARCH BAKED IN: Olympic = 1-3 reps @ %1RM w/ warm-up ramp (Catalyst/NSCA/BarBend); KB power rest = 1:1 to 1:3 work:rest. REMAINING ROUND-2: #2 Power Clean ramp cue (start empty bar -> 50% (~Ylb) -> 60% -> working set; loadable weights from actual best), #4 Pull-Up per-tier chart split + Layout-1 xlsx, #5 Planche/leverage swimming-consolidation, #6 Wall Sit (title-not-cue + TUT tile grid w/ per-tile added weight + attribution consistency), OVERARCHING swimming pass on all true variant families (enumerate from DB first, do NOT merge distinct lifts).",
     "mobile/app/(app)/effort/strength/[exercise].tsx; mobile/app/(app)/effort/cardio/[activity].tsx; mobile/src/lib/formulas.ts + movements.ts; web admin detail mirrors; CLAUDE.md locked specs; how-we-compute copy",
     "2026-06-05"),
]

# ─────────── build ──────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Task Pipeline"

    # Title band
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    t = ws.cell(row=1, column=1, value="MyRX — Task Pipeline")
    t.font = TITLE_FONT
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLS))
    gen = _dt.date.today().isoformat()
    s = ws.cell(row=2, column=1,
                value=f"Cross-session ledger of every task (done + pending). Stable IDs — refer back with 'pick up T0xx'. "
                      f"Generated {gen} by scripts/build_task_pipeline_xlsx.py — edit that script's TASKS list + re-run to update.")
    s.font = SUB_FONT
    s.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 26

    # Header
    HEAD_ROW = 3
    for ci, (name, width) in enumerate(COLS, start=1):
        c = ws.cell(row=HEAD_ROW, column=ci, value=name)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[HEAD_ROW].height = 26

    # Rows
    r = HEAD_ROW + 1
    for row in TASKS:
        for ci, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=ci, value=val)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border = BORDER
            c.font = Font(size=10, bold=(ci == 1))
            if ci % 2 == 0:
                pass
        # status colour (col 5)
        status = str(ws.cell(row=r, column=5).value or "")
        if status in STATUS_FILLS:
            ws.cell(row=r, column=5).fill = STATUS_FILLS[status]
        # subtle zebra on the long-text columns for readability
        if (r - HEAD_ROW) % 2 == 0:
            for ci in (2, 6, 7, 8):
                if ws.cell(row=r, column=ci).fill.fgColor.rgb in (None, "00000000"):
                    ws.cell(row=r, column=ci).fill = ZEBRA
        r += 1

    ws.freeze_panes = f"A{HEAD_ROW + 1}"
    ws.sheet_view.showGridLines = False

    # ── How to use sheet ──
    hw = wb.create_sheet("How to use")
    hw.column_dimensions["A"].width = 110
    lines = [
        ("MyRX — Task Pipeline · how to use", TITLE_FONT, 28),
        ("", None, 8),
        ("WHY THIS EXISTS", Font(bold=True, size=12, color="111721"), 22),
        ("We fork across many sessions and lose track of what's open / where we stopped. This is the single "
         "remembered list. Every task has a stable numeric ID (T001, T002, ...). Refer to one with "
         "\"pick up T021\" and this sheet says exactly where we left off and what's next.", Font(size=11), 60),
        ("", None, 8),
        ("CAPTURE-FIRST RULE (locked 2026-06-03)", Font(bold=True, size=12, color="111721"), 22),
        ("Every time a point comes up that needs fixing or discussion — ANY size — it gets logged here as a "
         "Pending task with a new T### id BEFORE the assistant replies, then it's answered. It stays Pending "
         "until the work is actually done/decided, then flips to Done (or Deferred/Parked/Reverted). Small "
         "items are exactly what gets lost, so nothing is too small to capture.", Font(size=11), 60),
        ("", None, 8),
        ("STATUS VALUES", Font(bold=True, size=12, color="111721"), 22),
        ("Done — shipped (and usually deployed/committed).", Font(size=11), 18),
        ("In progress — actively being iterated.", Font(size=11), 18),
        ("Pending — agreed/known but not started.", Font(size=11), 18),
        ("Deferred — intentionally postponed (waiting on something or a user decision).", Font(size=11), 18),
        ("Parked — discussed then set aside; needs a user decision to re-open.", Font(size=11), 18),
        ("Reverted — was tried, then undone.", Font(size=11), 18),
        ("Closed — dropped / won't do for now (can be reopened).", Font(size=11), 18),
        ("", None, 8),
        ("HOW TO UPDATE (the .xlsx is GENERATED, do not hand-edit)", Font(bold=True, size=12, color="111721"), 22),
        ("1. Edit the TASKS list in scripts/build_task_pipeline_xlsx.py (add a row / flip a status / update the "
         "'where we left off' + 'next' text).", Font(size=11), 34),
        ("2. Re-run:  python scripts/build_task_pipeline_xlsx.py", Font(size=11), 18),
        ("3. Commit BOTH the script and docs/TASK_PIPELINE.xlsx.", Font(size=11), 18),
        ("New tasks get the next free T### id. Never reuse or renumber an id.", Font(size=11), 18),
        ("", None, 8),
        ("SCOPE NOTE", Font(bold=True, size=12, color="111721"), 22),
        ("Seeded 2026-06-03 from the current CLAUDE.md + recent sessions. Older sessions weren't fully "
         "transcribed, so the 'prior' rows are best-effort reconstructions of completed features. Treat this "
         "as the authoritative living record from here forward — keep it current.", Font(size=11), 60),
    ]
    rr = 1
    for text, font, height in lines:
        c = hw.cell(row=rr, column=1, value=text)
        if font:
            c.font = font
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        hw.row_dimensions[rr].height = height
        rr += 1
    hw.sheet_view.showGridLines = False

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"[task-pipeline] wrote {OUT}  ({len(TASKS)} tasks)")

if __name__ == "__main__":
    main()
